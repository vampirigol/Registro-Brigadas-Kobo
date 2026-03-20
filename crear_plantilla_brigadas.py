#!/usr/bin/env python3
"""
Script para crear plantilla Excel para Brigadas de Salud y Atención en Clínicas.
Incluye validación de datos y filas con valores por defecto.
"""

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def crear_plantilla_brigadas():
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Brigadas de Salud"

    # Estilos
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    required_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    # Encabezados y opciones por defecto
    columnas = [
        ("Fecha de atención*", "yyyy-mm-dd", "Formato: 2025-03-11"),
        ("Toma consentimiento inicial*", ["Sí", "No"], "Buenos días. Mi nombre es... ¿Está usted de acuerdo?"),
        ("Lugar de atención", "", "Texto libre"),
        ("Modalidad de la atención*", ["Albergues", "Centros Comunitarios", "Clínica Adventista", "Escuelas", "Móvil"], None),
        ("Estado*", ["Baja California", "Baja California Sur", "Chihuahua", "Nuevo León", "Sonora", "Otro"], None),
        ("Latitud", "", "Formato: x.y °"),
        ("Longitud", "", "Formato: x.y °"),
        ("Altitud (m)", "", "metros"),
        ("Precisión (m)", "", "metros"),
        ("Primera vez o Seguimiento*", ["Primera vez", "Seguimiento", "Atención Única"], None),
        ("Entrega de Insumos*", ["ninguno seleccionado", "Sí", "No"], "Módulos antes de la consulta"),
        ("Servicio que se brinda*", ["Medicina General", "Dental", "Fisioterapia", "Oftalmología", "Laboratorios"], None),
        ("Nacionalidad*", "", "Donde nació el paciente"),
        ("Estatus migratorio*", [
            "En tránsito",
            "Retornado",
            "Solicitante de asilo",
            "Refugiado",
            "Residente temporal",
            "Residente permanente",
            "Prefiero no responder",
            "Ciudadano Mexicano"
        ], None),
        ("Sexo*", ["Masculino", "Femenino", "Otro", "Prefiero no responder"], None),
        ("Fecha de nacimiento*", "yyyy-mm-dd", "Formato: 2025-03-11"),
        ("Edad*", "", "0 si menor de 1 año"),
        ("Talla (cm)*", "", ""),
        ("Peso (kg)*", "", ""),
        ("Padecimiento médico actual*", "", ""),
        ("Motivo de la consulta", "", ""),
        ("Entrega de tratamiento*", ["Sí", "No"], None),
        ("¿Se hizo referencia?*", ["Sí", "No"], None),
        ("Consentimiento informado verbal*", ["Sí", "No"], "Sr/a NOMBRE DE PACIENTE..."),
    ]

    # Escribir encabezados (fila 1)
    for col_idx, (header, _, _) in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Crear filas con valores por defecto
    valores_default = [
        "2025-03-11",
        "Sí",
        "",
        "Centros Comunitarios",
        "Chihuahua",
        "",
        "",
        "",
        "",
        "Primera vez",
        "ninguno seleccionado",
        "Medicina General",
        "México",
        "Ciudadano Mexicano",
        "Masculino",
        "1990-01-15",
        "34",
        "170",
        "75",
        "",
        "",
        "Sí",
        "No",
        "Sí",
    ]

    # Fila 2: ejemplo con valores por defecto
    for col_idx, valor in enumerate(valores_default, 1):
        ws.cell(row=2, column=col_idx, value=valor)

    # Fila 3 en adelante: filas vacías para llenar (con primera columna como guía)
    for row_idx in range(3, 53):  # 50 filas adicionales para datos
        for col_idx in range(1, len(columnas) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx == 1:
                cell.value = "yyyy-mm-dd"

    # Agregar validaciones de datos
    validaciones = [
        (2, ["Sí", "No"]),           # Toma consentimiento
        (5, ["Albergues", "Centros Comunitarios", "Clínica Adventista", "Escuelas", "Móvil"]),
        (6, ["Baja California", "Baja California Sur", "Chihuahua", "Nuevo León", "Sonora", "Otro"]),
        (10, ["Primera vez", "Seguimiento", "Atención Única"]),
        (11, ["ninguno seleccionado", "Sí", "No"]),
        (12, ["Medicina General", "Dental", "Fisioterapia", "Oftalmología", "Laboratorios"]),
        (14, ["En tránsito", "Retornado", "Solicitante de asilo", "Refugiado", "Residente temporal",
              "Residente permanente", "Prefiero no responder", "Ciudadano Mexicano"]),
        (15, ["Masculino", "Femenino", "Otro", "Prefiero no responder"]),
        (21, ["Sí", "No"]),
        (22, ["Sí", "No"]),
        (23, ["Sí", "No"]),
    ]

    for col_idx, opciones in validaciones:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(opciones)}"',
            allow_blank=True,
        )
        dv.error = "Seleccione un valor de la lista"
        dv.errorTitle = "Valor no válido"
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}102")

    # Ajustar ancho de columnas
    for col_idx in range(1, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # Hoja 2: Referencia del texto de consentimiento
    ws_ref = wb.create_sheet("Referencia Consentimiento", 1)
    ws_ref["A1"] = "Texto de consentimiento inicial (copiar si se necesita)"
    ws_ref["A1"].font = Font(bold=True, size=12)
    ws_ref.merge_cells("A1:D1")
    texto_consentimiento = (
        "Buenos días. Mi nombre es NOMBRE COMPLETO, soy el/la AREA PROFESIONAL de Salud. "
        "En esta consulta le preguntaré información personal sobre el estado de salud y/o el de su hijx. "
        "Realizaré un examen físico en el cual utilizaré instrumentos como estetoscopio o termómetro, "
        "o haré exámenes con muestras de sangre. Todo con la finalidad de determinar un diagnóstico "
        "y escoger el mejor tratamiento para su padecimiento. En cualquier momento usted puede negarse "
        "a recibir cualquiera de los procedimientos. Adicional a lo anterior, solicitaré información "
        "personal como Nombre, apellidos, fecha y lugar de nacimiento, entre otros. Dicha captura es "
        "con fines estadísticos y de seguimiento, no se compartirá con nadie ajeno a ADRA a menos que "
        "usted lo exprese de forma escrita, puede negarse en cualquier momento. Es importante recordarle "
        "que nuestros servicios no están condicionados, por lo que nadie puede forzarle a hacer algo "
        "que no desee ni pedirle algo a cambio del mismo. ¿Está usted de acuerdo?"
    )
    ws_ref["A2"] = texto_consentimiento
    ws_ref["A2"].alignment = Alignment(wrap_text=True)
    ws_ref.column_dimensions["A"].width = 100
    ws_ref.row_dimensions[2].height = 200

    # Guardar
    output_path = "Plantilla_Brigadas_Salud_Clinicas.xlsx"
    wb.save(output_path)
    print(f"✓ Plantilla creada: {output_path}")
    return output_path


if __name__ == "__main__":
    crear_plantilla_brigadas()
