"""
Servidor Flask independiente para gestión de archivos Excel/PDF.
Subir, descargar, listar y marcar como validados.
"""

import csv
import difflib
import hashlib
import io
import importlib.util
import json
import logging
import os
import re
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
try:
    from dotenv import load_dotenv
except Exception:  # noqa: BLE001
    load_dotenv = None

from file_store import (
    EDIT_LOCK_TTL_SECONDS,
    PENDING_DIR,
    REFERENCES_DIR,
    VALID_STATUSES,
    VALIDATED_DIR,
    acquire_edit_lock,
    add_file_record,
    add_ref_record,
    count_file_rows,
    delete_file_record,
    delete_ref_record,
    ensure_validated_location,
    get_file_path,
    get_file_record,
    get_record_stats,
    get_ref_file_path,
    get_ref_record,
    get_uploader_stats,
    get_validator_stats,
    has_validated_replacement,
    init_files_db,
    kobo_submission_stats_by_file_ids,
    list_file_records,
    list_kobo_submission_logs,
    list_ref_locations,
    list_ref_records,
    log_kobo_submission,
    mark_file_validated,
    refresh_edit_lock,
    release_edit_lock,
    supersede_matching_files,
    supersede_specific_file,
    update_row_count,
    update_status,
)

try:
    from treatment_suggest import cohort_stats, suggest_treatment
    from cohort_index_builder import build_and_write, schedule_rebuild_if_quiet
except Exception:  # noqa: BLE001
    suggest_treatment = None
    cohort_stats = None
    build_and_write = None
    schedule_rebuild_if_quiet = None

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
if load_dotenv:
    try:
        load_dotenv(BASE_DIR / ".env")
    except Exception:
        pass

# Editor de hoja (tabla completa) vía web
MAX_SHEET_ROWS = 8_000
MAX_SHEET_COLS = 200
KOBO_SUBMIT_PASSWORD = "vamoskobo"

# Esquema visible para usuarios (encabezados amigables Kobo)
KOBO_VISIBLE_COLUMNS = [
    {
        "label": "Nombre del Paciente",
        "required": True,
        "aliases": ["Nombre del Paciente", "Nombre del paciente", "Nombre"],
        "hint": "Nombre completo del paciente.",
    },
    {
        "label": "Fecha de atención",
        "required": True,
        "aliases": ["Fecha de atención", "Fecha de atencion", "Fecha de Atención", "Fecha"],
        "hint": "Formato recomendado: AAAA-MM-DD.",
    },
    {
        "label": "Sexo",
        "required": True,
        "aliases": ["Sexo", "SEX"],
        "options": ["F", "M"],
    },
    {
        "label": "Toma de consentimiento antes de iniciar la consulta",
        "internal": "CONS1",
        "required": True,
        "aliases": [
            "Toma de consentimiento antes de iniciar la consulta",
            "Toma consentimiento inicial",
            "Consent.",
            "CONS1",
            "Consentimiento",
        ],
        "options": ["Sí", "No"],
        "hint": "Primer consentimiento, antes de iniciar la consulta. No es el mismo campo que el consentimiento informado verbal (final).",
    },
    {
        "label": "¿Se tomó consentimiento informado de forma verbal?",
        "internal": "CONS",
        "required": True,
        "aliases": [
            "¿Se tomó consentimiento informado de forma verbal?",
            "Se tomó consentimiento informado de forma verbal",
            "Se tomó consentimiento informado verbal",
            "CONS",
            "Consentimiento informado verbal",
        ],
        "options": ["Sí", "No"],
        "hint": "Obligatorio en Kobo. Indique Sí o No según se haya obtenido el consentimiento informado verbal.",
    },
    {
        "label": "Servicio que se brinda",
        "required": True,
        "aliases": ["Servicio que se brinda", "Servicio", "Especialidad"],
        "options": ["Medicina General", "Dental", "Fisioterapia", "Oftalmología", "Laboratorios"],
    },
    {"label": "Edad", "required": False, "aliases": ["Edad", "AGE"], "hint": "Edad en años."},
    {"label": "Fecha de nacimiento", "required": False, "aliases": ["Fecha de nacimiento", "DOB", "Fecha nacimiento"], "hint": "Formato AAAA-MM-DD."},
    {
        "label": "Estado",
        "required": False,
        "aliases": [
            "Estado",
            "Estado brigada",
            "Estado_brigada",
            "Estado de origen",
            "Estado paciente",
            "Estado_paciente",
            "Originario",
        ],
        "options": ["Baja California Sur", "Chihuahua", "Sonora", "Baja California", "Nuevo León"],
    },
    {
        "label": "Lugar de atención",
        "required": False,
        "aliases": ["Lugar de atención", "Lugar de atencion", "Lugar"],
        "hint": (
            "Opciones por estado. Nuevo León usa Montemorelos por defecto si está vacío. "
            "Baja California: Valle de la Trinidad, San Matías, Santa Catalina, Comunidad Kiliwa, Tijuana, Otro. "
            "Baja California Sur: Santa Rosalía, Mulege, Loreto, Ciudad Constitución, Vizcaíno, "
            "Bahía Tortuga, Bahía Asunción, Punta Abreojos, La Bucana, Otro. "
            "Chihuahua: Ciudad Juárez, Otro."
        ),
        "options": [
            "Montemorelos",
            "Valle de la Trinidad",
            "San Matías",
            "Santa Catalina",
            "Comunidad Kiliwa",
            "Tijuana",
            "Santa Rosalía",
            "Mulege",
            "Loreto",
            "Ciudad Constitución",
            "Vizcaíno",
            "Bahía Tortuga",
            "Bahía Asunción",
            "Punta Abreojos",
            "La Bucana",
            "Ciudad Juárez",
            "Otro",
        ],
    },
    {
        "label": "Modalidad",
        "required": False,
        "aliases": ["Modalidad", "Modalidad de la atención", "Modalidad_de_la_atenci_n", "Modalidad de atención"],
        "options": ["Albergues", "Centros Comunitarios", "Clínica Adventista", "Escuelas", "Móvil"],
    },
    {
        "label": "Especificar colegio o comunidad (SCH)",
        "internal": "SCH",
        "required": False,
        "aliases": [
            "SCH",
            "Especificar nombre de colegio o comunidad",
            "Especificar colegio o comunidad (SCH)",
            "Lugar de atención: Escuelas",
            "Nombre de colegio o comunidad",
        ],
        "hint": "Use esta columna cuando «Modalidad» = Escuelas. Nombre del plantel o comunidad; no replicar el campo «Lugar de atención» en otros modos.",
    },
    {
        "label": "Primera vez o Seguimiento",
        "required": False,
        "aliases": ["Primera vez o Seguimiento", "Primera vez o seguimiento", "followup"],
        "options": ["Primera vez", "Seguimiento", "Atención Única", "Entrega de Insumos"],
    },
    {
        "label": "Nacionalidad",
        "internal": "NAT",
        "required": False,
        "aliases": ["Nacionalidad", "NAT"],
    },
    {
        "label": "Especificar (nacionalidad)",
        "internal": "NATOT",
        "required": False,
        "aliases": [
            "Especificar",
            "NATOT",
            "Nacionalidad (especificar)",
        ],
        "hint": "En Kobo, junto a Nacionalidad el campo de texto se llama «Especificar»; es el detalle (NATOT), no el de diagnóstico de medicina. Recomendable solo si la nacionalidad no es una opción estándar; si no, vacío.",
    },
    {
        "label": "Estatus migratorio",
        "required": False,
        "aliases": ["Estatus migratorio", "estatus_migra"],
        "options": [
            "Ciudadano Mexicano",
            "Solicitante de asilo",
            "Refugiado",
            "Migrante",
            "Retornado Mexicano",
            "Desplazado Interno",
            "Apátrida",
            "Comunidad de Acogida",
            "Otro",
        ],
    },
    {
        "label": "Minoría",
        "required": False,
        "aliases": [
            "Minoría",
            "Minoria",
            "_Pertenece_a_alguna_minor_a_t",
            "¿Pertenece a alguna minoría étnica?",
            "Pertenece a alguna minoría étnica",
        ],
        "options": ["Sí", "No"],
    },
    {"label": "Talla (cm)", "required": False, "aliases": ["Talla (cm)", "HEI"]},
    {"label": "Peso (kg)", "required": False, "aliases": ["Peso (kg)", "WEI"]},
    {"label": "Padecimiento médico actual", "required": False, "aliases": ["Padecimiento médico actual", "Padecimiento medico actual", "Diagnostico_Motivo", "DX"]},
    {"label": "Entrega Tratamiento", "required": False, "aliases": ["Entrega Tratamiento", "¿Entrega Tratamiento?", "entrega_tx"], "options": ["Sí", "No"]},
    {
        "label": "Tratamiento",
        "required": False,
        "aliases": [
            "Tratamiento",
            "Insumos Entregados",
            "Insumos Entregados (Categoría general)",
            "Insumos Entregados (Categoria general)",
            "Resultados_Lab_Insumos",
            "Resultados Lab / Insumos",
            "Insumos",
        ],
        "hint": "Si Oftalmología brinda lentes especificar graduación de ojo derecho/izquierdo.",
    },
    {"label": "Se hizo referencia", "required": False, "aliases": ["Se hizo referencia", "¿Se hizo referencia?", "Referencia"], "options": ["Sí", "No"]},
    {
        "label": "A dónde",
        "required": False,
        "aliases": ["A dónde", "A donde", "¿A dónde?", "Referencia_donde"],
        "options": ["Clínica", "Segundo Nivel", "Hospital", "Laboratorio", "ONG", "Ministerio público", "Otro"],
    },
    {
        "label": "Motivo Ref",
        "internal": "Motivo_referencia",
        "required": False,
        "aliases": ["Motivo Ref", "Motivo referencia", "Motivo_referencia", "Motivo Referido", "Motivo referido"],
    },
    {
        "label": "Especificar (motivo referido)",
        "internal": "Motivo_especificar",
        "required": False,
        "aliases": [
            "Especificar (motivo referido)",
            "Especificar motivo referencia",
            "Especificar m. ref. fisioterapia",
            "Especificar m. ref. fisio",
            "Motivo_especificar",
        ],
        "hint": "Detalle obligatorio en fisioterapia si hay referencia y «Motivo Ref» = «Otro» (Kobo: SPREFMOTMED).",
    },
    {"label": "Acompañante", "required": False, "aliases": ["Acompañante", "Acompanante", "CGR"]},
    {
        "label": "Indicar si el paciente tiene alguna de las siguientes discapacidades",
        "required": False,
        "aliases": [
            "Indicar si el paciente tiene alguna de las siguientes discapacidades",
            "Discapacidad",
            "Tipo de discapacidad",
        ],
        "options": ["Motriz", "Visual", "Auditiva", "Intelectual", "Otra"],
    },
    {
        "label": "¿Mujer embarazada o en periodo de lactancia?",
        "internal": "ME_ML",
        "required": False,
        "aliases": [
            "¿Mujer embarazada o en periodo de lactancia?",
            "Mujer embarazada o en periodo de lactancia",
            "¿Embarazada / Lactancia?",
            "Embarazada / Lactancia",
            "ME_ML",
            "Embarazada o lactancia",
            "Embarazo/Lactancia",
            "Embarazo / Lactancia",
            "Embarazada/Lactancia?",
        ],
        "options": ["Embarazada", "Lactancia", "No Aplica"],
        "hint": "Solo aplica si el sexo es femenino; deje vacío o use No Aplica si no aplica.",
    },
    {
        "label": "Fisioterapia",
        "internal": "Fisio_Diagnostico",
        "required": False,
        "aliases": [
            "Fisioterapia",
            "Diagnóstico Fisioterapia",
            "Diagnostico Fisioterapia",
            "Diagnóstico en Fisioterapia",
            "Diagnóstico fisioterapia",
        ],
        "options": [
            "Revisión",
            "Artrosis",
            "Artritis",
            "Lesiones musculoesqueléticas",
            "Dolor crónico",
            "Enfermedades neurológicas",
            "Problemas respiratorios",
            "Dolor",
            "Contractura",
            "Otro",
        ],
        "hint": "Texto o lista de diagnósticos; KoboUp convierte a 0/1 en columnas Diagnóstico/... en la API. Export: Diagnóstico/Revisión, …/Dolor, …/Especificar.",
    },
    {
        "label": "Plan de Tratamiento",
        "internal": "Plan_de_Tratamiento",
        "required": False,
        "aliases": [
            "Plan de Tratamiento",
            "Plan de Tratamiento (Fisioterapia u otros)",
            "Plan de Tratamiento (Fisioterapia u otro)",
            "Plan de tratamiento fisioterapia u otros",
            "PlanTratamiento",
            "Plan tratamiento",
        ],
        "hint": "Texto o plan; en fisioterapia suele describirse el plan terapéutico.",
    },
    {
        "label": "Diagnóstico Medicina General",
        "required": False,
        "aliases": [
            "Diagnóstico Medicina General",
            "Diagnostico Medicina General",
            "Diagnósticos Medicina General",
            "Diagnosticos Medicina General",
            "Diagnósticos",
            "Diagnosticos",
        ],
        "options": [
            "Ninguno seleccionado",
            "Consulta de rutina",
            "Diarrea aguda",
            "Dermatosis",
            "Bronquitis aguda",
            "Embarazo",
            "Parasitosis",
            "Dolor abdominal",
            "Amigdalitis",
            "Desnutrición",
            "Anemia",
            "Síndrome febril",
            "Asma",
            "Cefalea",
            "Pediculosis",
            "Deshidratación",
            "Rinofaringitis",
            "Rinitis alérgica",
            "Conjuntivitis",
            "Escabiosis",
            "Dermatitis alérgica",
            "Estreñimiento",
            "Amenorrea",
            "Carie dental",
            "Cistitis",
            "Consulta de seguimiento",
            "Dermatitis irritante primaria del pañal",
            "Faringoamigdalitis aguda",
            "Gastroenteritis o colitis de origen infeccioso (sin especificación del agente infeccioso)",
            "Herida",
            "Otitis media",
            "Pielonefritis aguda",
            "Quemadura",
            "Sospecha de dengue (Con signos de alarma)",
            "Sospecha de dengue (Sin signos de alarma)",
            "Sospecha de neumonía",
            "Sospecha de rubeola",
            "Sospecha de sarampión",
            "Sospecha de malaria",
            "Traumatismos",
            "Vaginitis",
            "Vaginosis",
            "Varicela",
            "Sinusitis",
            "Otro",
        ],
        "hint": "Diagnósticos — Marcar todos los que apliquen. Puede marcar múltiples (solo Medicina General).",
    },
    {
        "label": "Especificar diagnóstico (Medicina General)",
        "internal": "dxesp",
        "required": False,
        "aliases": [
            "Especificar diagnóstico (Medicina General)",
            "Especificar diagnóstico",
            "Especificar Diagnóstico Medicina General",
            "Especificar Diagnostico Medicina General",
            "dxesp",
        ],
        "hint": "Campo Kobo «dxesp» en bloque de medicina: texto cuando en diagnósticos se incluye «Otro». No es el «Especificar» de nacionalidad (NATOT) ni el de fisioterapia.",
    },
    {
        "label": "Diagnóstico Odontología",
        "required": False,
        "aliases": [
            "Diagnóstico Odontología",
            "Diagnostico Odontología",
            "Diagnostico Odontologia",
            "Diagnóstico Odontologia",
        ],
        "options": [
            "Caries",
            "Cálculo",
            "Sarro",
            "Gingivitis",
            "Periodontitis",
            "Pulpitis Reversible",
            "Pulpitis Irreversible",
            "Fractura",
            "Otro",
        ],
    },
    {
        "label": "¿Se realiza procedimiento odontológico?",
        "required": False,
        "aliases": [
            "¿Se realiza procedimiento odontológico?",
            "Se realiza procedimiento odontológico?",
            "¿Se realiza procedimiento odontológico? (ej. Limpieza, Extracción, Resina)",
        ],
        "options": ["Si", "No"],
        "hint": "Solo odontología. Si elige Si, registre el o los procedimientos en la columna siguiente.",
    },
    {
        "label": "Qué procedimiento se realiza",
        "required": False,
        "aliases": [
            "Qué procedimiento se realiza",
            "Que procedimiento se realiza",
            "¿Qué procedimiento se realiza?",
            "¿Que procedimiento se realiza?",
            "Qué procedimiento",
            "Que procedimiento",
        ],
        "options": [
            "Resina",
            "Limpieza dental",
            "Endodoncia",
            "Extracción",
            "Cirugía",
            "Otro",
        ],
        "hint": "Opciones del formulario Kobo (puede haber varias). Si usa Otro, detalle según las notas del formulario.",
    },
    {
        "label": "Síntomas que presenta a la fecha de consulta",
        "required": False,
        "aliases": [
            "Síntomas que presenta a la fecha de consulta",
            "Sintomas que presenta a la fecha de consulta",
            "Síntomas",
            "Sintomas",
        ],
        "options": [
            "Ninguno",
            "Ardor",
            "Comezón",
            "Irritación",
            "Lagrimeo",
            "Fotofobia",
            "Dificultad para leer",
            "Otro",
        ],
    },
    {
        "label": "¿Ha recibido algún diagnóstico previo?",
        "required": False,
        "aliases": [
            "¿Ha recibido algún diagnóstico previo?",
            "Ha recibido algún diagnóstico previo",
            "Ha recibido algun diagnostico previo",
        ],
        "options": ["Ninguno", "Catarata", "Glaucoma", "Estrabismo", "Retinopatía", "Pterigión", "Otro"],
    },
    {
        "label": "Diagnóstico Actual",
        "required": False,
        "aliases": ["Diagnóstico Actual", "Diagnostico Actual"],
        "options": ["Ametropía", "Miopía", "Astigmatismo", "Hipermetropía", "Estabismo", "Otro"],
    },
    {
        "label": "¿Se hizo entrega de tratamiento/artículos al beneficiario o beneficiaria?",
        "required": False,
        "aliases": [
            "¿Se hizo entrega de tratamiento/artículos al beneficiario o beneficiaria?",
            "Se hizo entrega de tratamiento/artículos al beneficiario o beneficiaria",
            "Se hizo entrega de tratamiento al beneficiario",
        ],
        "options": ["Si", "No"],
        "hint": "Especifica si se le brindó tratamiento al paciente.",
    },
    {
        "label": "Tratamiento",
        "required": False,
        "aliases": ["Tratamiento"],
        "hint": "Si es Oftalmología y brinda lentes, especifique graduación de ojo derecho/izquierdo.",
    },
    {
        "label": "Especifique qué se entrega",
        "required": False,
        "aliases": [
            "Especifique qué se entrega",
            "Especifique que se entrega",
            "Especifique qué se entrega (detalle del insumo)",
            "Especifique que se entrega (detalle del insumo)",
            "Especifique_qu_se_entrega",
        ],
        "options": [
            "Anteojos",
            "Medicamento/suplemento",
            "Plan de Tratamiento (solo fisio o entrega de receda sin medicamento)",
            "Resultados de Laboratorio",
            "Otro",
        ],
    },
    {
        "label": "Especificar lo que se entrega al beneficiario",
        "internal": "Especificar_lo_que_se_entrega_",
        "required": False,
        "aliases": [
            "Especificar lo que se entrega al beneficiario",
            "Especificar lo que se entrega",
            "Especificar_lo_que_se_entrega_",
        ],
        "hint": "Obligatorio en Kobo si «Especifique qué se entrega» = Otro. Describa el insumo o el detalle entregado.",
    },
    {
        "label": "¿Se le dio suplementos que contengan ácido fólico?",
        "internal": "SUP_Acido_folico",
        "required": False,
        "aliases": [
            "¿Se le dio suplementos que contengan ácido fólico?",
            "Se le dio suplementos que contengan ácido fólico",
            "Se le dio suplementos que contengan acido folico",
        ],
        "options": ["Sí", "No"],
    },
    {
        "label": "¿Se le dio suplementos que contengan hierro?",
        "internal": "SUP_Hierro",
        "required": False,
        "aliases": [
            "¿Se le dio suplementos que contengan hierro?",
            "Se le dio suplementos que contengan hierro",
        ],
        "options": ["Sí", "No"],
    },
    {
        "label": "Unidades entregadas",
        "required": False,
        "aliases": ["Unidades entregadas", "Unidades_entregadas"],
        "hint": "Especifica el número de dosis/unidades entregadas.",
    },
    {
        "label": "¿Se le ha brindado asesoría en uno de los módulos el día de hoy?",
        "internal": "ASESPREV",
        "required": False,
        "aliases": [
            "¿Se le ha brindado asesoría en uno de los módulos el día de hoy?",
            "Se le ha brindado asesoría en uno de los módulos el día de hoy",
            "ASESPREV",
            "Asesoría del día",
            "Asesoria del dia",
        ],
        "options": ["Medicina General", "Oftalmología", "Dental", "Fisioterapia", "Laboratorios", "No Aplica"],
        "hint": "Seleccione el módulo donde recibió asesoría hoy.",
    },
    {
        "label": "Laboratorio Clínico",
        "required": False,
        "aliases": [
            "Laboratorio Clínico",
            "Laboratorio clinico",
            "LABORATORIO CLÍNICO",
            "LABORATORIO CLINICO",
        ],
    },
    {
        "label": "Diagnóstico / Resu",
        "required": False,
        "aliases": [
            "Diagnóstico / Resu",
            "Diagnostico / Resu",
            "Diagnóstico/Resu",
            "Diagnostico/Resu",
            "Diagnóstico / Resultados Laboratorio",
            "Diagnostico / Resultados Laboratorio",
            "Diagnóstico Resultados Laboratorio",
            "Diagnostico Resultados Laboratorio",
        ],
    },
    {
        "label": "Coordenadas",
        "required": False,
        "aliases": [
            "Coordenadas",
            "Ubicación geográfica de la atención",
            "Ubicación geográfica de la atencion",
            "Ubicación geográfica",
            "Ubicacion geografica de la atencion",
        ],
    },
    {"label": "Latitud", "required": False, "aliases": ["Latitud", "LAT", "latitud"]},
    {"label": "Longitud", "required": False, "aliases": ["Longitud", "LON", "LNG", "longitud"]},
]

ALWAYS_REQUIRED_SHEET_COLUMNS = [
    "Toma de consentimiento antes de iniciar la consulta",
    # CONS: obligatoria en el formulario Kobo; se inserta al abrir/guardar si falta.
    "¿Se tomó consentimiento informado de forma verbal?",
    # Estas columnas son críticas para edición clínica básica.
    "Diagnóstico Medicina General",
    "Diagnóstico Odontología",
    "Tratamiento",
]

SERVICE_CONDITIONAL_COLUMNS: dict[str, list[str]] = {
    "Medicina General": [
        "Padecimiento médico actual",
        "Indicar si el paciente tiene alguna de las siguientes discapacidades",
        "Diagnóstico Medicina General",
        "Especificar diagnóstico (Medicina General)",
    ],
    "Dental": [
        "Diagnóstico Odontología",
        "¿Se realiza procedimiento odontológico?",
        "Qué procedimiento se realiza",
    ],
    "Fisioterapia": [
        "Fisioterapia",
        "Plan de Tratamiento",
    ],
    "Oftalmología": [
        "Síntomas que presenta a la fecha de consulta",
        "¿Ha recibido algún diagnóstico previo?",
        "Diagnóstico Actual",
        "Requiere anteojos",
    ],
    "Laboratorios": [
        "Laboratorio Clínico",
        "Diagnóstico / Resu",
    ],
}


def _merge_kobo_schema_with_template(schema: list[dict]) -> tuple[list[dict], str]:
    """
    Enriquece el esquema visible con metadata del XLSForm/plantilla Kobo:
    - required real
    - options (choices)
    - hint (nota)
    """
    template_candidates = [
        BASE_DIR / "_release_kpi" / "crear_plantilla_kobo.py",
        BASE_DIR / "crear_plantilla_kobo.py",
        BASE_DIR.parent / "_release_kpi" / "crear_plantilla_kobo.py",
        BASE_DIR.parent / "crear_plantilla_kobo.py",
    ]
    module_path = next((p for p in template_candidates if p.exists()), None)
    if not module_path:
        return schema, "embedded"

    try:
        spec = importlib.util.spec_from_file_location("kobo_template_schema_loader", module_path)
        if not spec or not spec.loader:
            return schema, "embedded"
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        cols = getattr(mod, "COLUMNS", None)
        if not isinstance(cols, list) or not cols:
            return schema, "embedded"
    except Exception:
        return schema, "embedded"

    def _hdrnorm(value: str) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFD", text)
        text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
        return re.sub(r"[^a-z0-9]+", " ", text).strip()

    # Índices para matching flexible entre esquema actual y template Kobo.
    label_map: dict[str, dict] = {}
    internal_map: dict[str, dict] = {}
    for item in cols:
        if not isinstance(item, (list, tuple)) or len(item) < 6:
            continue
        t_label, t_internal, t_required, t_options, _, t_note = item[:6]
        meta = {
            "label": str(t_label or "").strip(),
            "internal": str(t_internal or "").strip(),
            "required": bool(t_required),
            "options": [str(v).strip() for v in (t_options or []) if str(v).strip()] if isinstance(t_options, list) else [],
            "hint": str(t_note or "").strip(),
        }
        lbl_key = _hdrnorm(meta["label"])
        if lbl_key:
            label_map[lbl_key] = meta
        internal_key = _hdrnorm(meta["internal"])
        if internal_key:
            internal_map[internal_key] = meta

    enriched: list[dict] = []
    for col in schema:
        cur = dict(col)
        aliases = [str(a) for a in cur.get("aliases", []) if str(a).strip()]
        if cur.get("label"):
            aliases.append(str(cur["label"]))
        cur_keys = {_hdrnorm(a) for a in aliases if _hdrnorm(a)}
        match = label_map.get(_hdrnorm(str(cur.get("label", ""))))
        if not match:
            for k in cur_keys:
                if k in internal_map:
                    match = internal_map[k]
                    break
        if match:
            cur["required"] = bool(match["required"])
            if match.get("hint"):
                cur["hint"] = match["hint"]
            if match.get("options"):
                existing_opts = [str(v).strip() for v in cur.get("options", []) if str(v).strip()]
                seen_opt: set[str] = set()
                merged_opts: list[str] = []
                for opt in list(match["options"]) + existing_opts:
                    key = _hdrnorm(opt)
                    if not key or key in seen_opt:
                        continue
                    seen_opt.add(key)
                    merged_opts.append(opt)
                cur["options"] = merged_opts
            if match.get("internal"):
                cur["internal"] = match["internal"]
                internal_name = str(match["internal"]).strip()
                if internal_name and internal_name not in aliases:
                    aliases.append(internal_name)
            # Mantener aliases sin duplicados preservando orden.
            seen_alias: set[str] = set()
            uniq_aliases: list[str] = []
            for a in aliases:
                key = _hdrnorm(a)
                if not key or key in seen_alias:
                    continue
                seen_alias.add(key)
                uniq_aliases.append(a)
            cur["aliases"] = uniq_aliases
        enriched.append(cur)
    return enriched, f"template:{module_path.name}"


KOBO_VISIBLE_COLUMNS, KOBO_SCHEMA_SOURCE = _merge_kobo_schema_with_template(KOBO_VISIBLE_COLUMNS)

MEXICO_STATES: list[str] = [
    "Aguascalientes",
    "Baja California",
    "Baja California Sur",
    "Campeche",
    "Chiapas",
    "Chihuahua",
    "Ciudad de México",
    "Coahuila",
    "Colima",
    "Durango",
    "Estado de México",
    "Guanajuato",
    "Guerrero",
    "Hidalgo",
    "Jalisco",
    "Michoacán",
    "Morelos",
    "Nayarit",
    "Nuevo León",
    "Oaxaca",
    "Puebla",
    "Querétaro",
    "Quintana Roo",
    "San Luis Potosí",
    "Sinaloa",
    "Sonora",
    "Tabasco",
    "Tamaulipas",
    "Tlaxcala",
    "Veracruz",
    "Yucatán",
    "Zacatecas",
]

# Overrides funcionales solicitados para UX/operación.
KOBO_OPTIONS_OVERRIDE: dict[str, list[str]] = {
    "Sexo": ["F", "M", "Femenino", "Masculino"],
    "Estado": MEXICO_STATES,
    "A dónde": [
        "Clínica",
        "Segundo Nivel",
        "Hospital",
        "Laboratorio",
        "ONG",
        "Ministerio público",
        "Otro",
        "No Aplica",
        "N/D",
    ],
    "Lugar de atención": [
        "Montemorelos",
        "Valle de la Trinidad",
        "San Matías",
        "Santa Catalina",
        "Comunidad Kiliwa",
        "Tijuana",
        "Santa Rosalía",
        "Mulege",
        "Loreto",
        "Ciudad Constitución",
        "Vizcaíno",
        "Bahía Tortuga",
        "Bahía Asunción",
        "Punta Abreojos",
        "La Bucana",
        "Ciudad Juárez",
        "Otro",
    ],
    "Modalidad": ["Albergues", "Centros Comunitarios", "Clínica Adventista", "Escuelas", "Móvil"],
    "Toma de consentimiento antes de iniciar la consulta": ["Sí", "No"],
    "¿Se tomó consentimiento informado de forma verbal?": ["Sí", "No"],
    "¿Se le ha brindado asesoría en uno de los módulos el día de hoy?": [
        "Medicina General",
        "Oftalmología",
        "Dental",
        "Fisioterapia",
        "Laboratorios",
        "No Aplica",
    ],
    "Indicar si el paciente tiene alguna de las siguientes discapacidades": [
        "Motriz",
        "Visual",
        "Auditiva",
        "Intelectual",
        "Otra",
    ],
    "Fisioterapia": [
        "Artrosis",
        "Artritis",
        "Lesiones musculoesqueléticas",
        "Dolor crónico",
        "Enfermedades neurológicas",
        "Problemas respiratorios",
        "Otro",
    ],
}
for _col in KOBO_VISIBLE_COLUMNS:
    lbl = str(_col.get("label") or "")
    if lbl in KOBO_OPTIONS_OVERRIDE:
        _col["options"] = list(KOBO_OPTIONS_OVERRIDE[lbl])
ROOT_PROJECT_DIR = BASE_DIR.parent
ROOT_UPLOADS_DIR = ROOT_PROJECT_DIR / "uploads"
ROOT_LOGS_DIR = ROOT_PROJECT_DIR / "logs"
ROOT_SUBMITTED_FILE = ROOT_LOGS_DIR / "filas_enviadas.json"
KPI_EXCLUSIONS_FILE = BASE_DIR / "logs" / "kpi_excluded_files.json"
PRIORITY_VALIDATED_DIR = Path(
    "/Users/luciodelacruz/Desktop/2026/Llenado Kobo tools.bak_20260320/archivos_validados_20260411_004650"
)
EXTRA_KPI_FILES = [
    ROOT_UPLOADS_DIR / "01_FERIA_DE_SALUD-_PLOMO_CD.JUAREZ.xlsx",
    ROOT_UPLOADS_DIR / "02_FERIA_DE_SALUD-_PLOMO_CD.JUAREZ.xlsx",
    ROOT_UPLOADS_DIR / "03_FERIA_DE_SALUD-_PLOMO_CD.JUAREZ.xlsx",
    ROOT_UPLOADS_DIR / "04_FERIA_DE_SALUD-_PLOMO_CD.JUAREZ_121-160.xlsx",
    ROOT_UPLOADS_DIR / "05_FERIA_DE_SALUD-_PLOMO_CD.JUAREZ_161_al_200.xlsx",
]

app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="/static")
CORS(app)

ALLOWED_EXTENSIONS = {"xlsx", "xls", "csv", "pdf"}
MAX_UPLOAD_MB = 50

init_files_db()


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _active_edit_lock_from_entry(entry: dict) -> dict | None:
    locked_by = str(entry.get("edit_locked_by") or "").strip()
    lock_at_raw = entry.get("edit_lock_at")
    if not locked_by or not lock_at_raw:
        return None
    try:
        lock_at = datetime.fromisoformat(str(lock_at_raw))
    except Exception:
        return None
    age_seconds = (datetime.utcnow() - lock_at).total_seconds()
    if age_seconds > EDIT_LOCK_TTL_SECONDS:
        return None
    return {
        "locked_by": locked_by,
        "locked_at": str(lock_at_raw),
        "age_seconds": int(max(0, age_seconds)),
    }


def _augment(entry: dict) -> dict:
    if not entry:
        return entry
    entry = dict(entry)
    fid = entry.get("id")
    if fid is not None:
        entry["download_url"] = f"api/files/{fid}/download"
    active_lock = _active_edit_lock_from_entry(entry)
    entry["is_editing"] = bool(active_lock)
    entry["editing_by"] = active_lock.get("locked_by") if active_lock else None
    entry["editing_at"] = active_lock.get("locked_at") if active_lock else None
    entry["editing_age_seconds"] = active_lock.get("age_seconds") if active_lock else None
    edited_flag = entry.get("edited_validated")
    entry["edited_validated"] = bool(edited_flag) if edited_flag is not None else False
    return entry


def _load_kpi_exclusions() -> set[str]:
    try:
        if not KPI_EXCLUSIONS_FILE.exists():
            return set()
        data = json.loads(KPI_EXCLUSIONS_FILE.read_text(encoding="utf-8"))
        files = data.get("excluded_files", [])
        return {str(name).strip() for name in files if str(name).strip()}
    except Exception:
        return set()


def _save_kpi_exclusions(excluded_files: set[str]) -> None:
    KPI_EXCLUSIONS_FILE.parent.mkdir(parents=True, exist_ok=True)
    payload = {"excluded_files": sorted(excluded_files)}
    KPI_EXCLUSIONS_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _norm_text(value: str) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def _normalize_header(value: str) -> str:
    text = _norm_text(value)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _build_kobo_alias_map() -> dict[str, str]:
    out: dict[str, str] = {}
    for col in KOBO_VISIBLE_COLUMNS:
        label = str(col["label"])
        aliases = list(col.get("aliases", [])) + [label]
        for a in aliases:
            key = _normalize_header(str(a))
            if key and key not in out:
                out[key] = label
    return out


def _read_headers_quick(path: Path) -> list[str]:
    """Lee solo encabezados (primera fila) para xlsx/csv sin dependencias pesadas."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                with open(path, "r", encoding=encoding, errors="replace", newline="") as fh:
                    sample = fh.read(4096)
                    fh.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
                    except Exception:
                        dialect = csv.excel
                    reader = csv.reader(fh, dialect)
                    headers = next(reader, None) or []
                    if len(headers) == 1 and ";" in str(headers[0]):
                        headers = str(headers[0]).split(";")
                    return [str(h).strip() for h in headers if str(h).strip()]
            except Exception:
                continue
        return []

    if suffix in {".xlsx", ".xlsm"}:
        try:
            with zipfile.ZipFile(path) as zf:
                wb = ET.fromstring(zf.read("xl/workbook.xml"))
                wb_ns = {
                    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                }
                first_sheet = wb.find("m:sheets/m:sheet", wb_ns)
                if first_sheet is None:
                    return []
                rel_id = str(first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id") or "").strip()
                if not rel_id:
                    return []
                rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                rels_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
                target = ""
                for rel in rels.findall("r:Relationship", rels_ns):
                    if str(rel.attrib.get("Id") or "") == rel_id:
                        target = str(rel.attrib.get("Target") or "").strip()
                        break
                if not target:
                    return []
                sheet_path = f"xl/{target.lstrip('/')}"
                ws = ET.fromstring(zf.read(sheet_path))
                ws_ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                row = ws.find("m:sheetData/m:row", ws_ns)
                if row is None:
                    return []
                shared: list[str] = []
                if "xl/sharedStrings.xml" in zf.namelist():
                    sst = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                    for si in sst.findall("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"):
                        parts = []
                        for tn in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
                            parts.append(tn.text or "")
                        shared.append("".join(parts))
                headers: list[str] = []
                for cell in row.findall("m:c", ws_ns):
                    t = str(cell.attrib.get("t") or "")
                    v_node = cell.find("m:v", ws_ns)
                    txt = ""
                    if v_node is not None and v_node.text is not None:
                        raw = str(v_node.text)
                        if t == "s":
                            try:
                                txt = shared[int(raw)]
                            except Exception:
                                txt = raw
                        else:
                            txt = raw
                    inlined = cell.find("m:is/m:t", ws_ns)
                    if inlined is not None and inlined.text is not None:
                        txt = str(inlined.text)
                    txt = txt.strip()
                    if txt:
                        headers.append(txt)
                return headers
        except Exception:
            return []
    return []


def _enrich_kobo_aliases_from_validated(schema: list[dict]) -> list[dict]:
    """
    Aprende aliases reales desde archivos en validados para mejorar matching Kobo.
    """
    base_map = _build_kobo_alias_map()
    alias_keys = list(base_map.keys())
    counts: dict[str, int] = {}
    try:
        files = [
            p
            for p in sorted(VALIDATED_DIR.glob("*"))
            if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm", ".csv"}
        ][:220]
        for p in files:
            for h in _read_headers_quick(p):
                counts[h] = counts.get(h, 0) + 1
    except Exception:
        return schema

    if not counts:
        return schema

    seed_label_by_alias: dict[str, str] = {
        "diagnosticos": "Diagnóstico Medicina General",
        "diagnosticos medicina general": "Diagnóstico Medicina General",
        "diagnostico fisio": "Fisioterapia",
        "diag fisio": "Fisioterapia",
        "diag fisio terapia": "Fisioterapia",
        "diag lab": "Diagnóstico / Resultados Laboratorio",
        "diag resultados lab": "Diagnóstico / Resultados Laboratorio",
        "primera vez seg": "Primera vez o Seguimiento",
        "primera vez / seg": "Primera vez o Seguimiento",
        "asesoria previa hoy": "¿Se le ha brindado asesoría en uno de los módulos el día de hoy?",
        "referencia": "Se hizo referencia",
        "referencia?": "Se hizo referencia",
        "a donde": "A dónde",
        "motivo referencia": "Motivo Ref",
        "motivo ref": "Motivo Ref",
        "entrega de tratamiento": "¿Se hizo entrega de tratamiento/artículos al beneficiario o beneficiaria?",
        "entrega trat": "¿Se hizo entrega de tratamiento/artículos al beneficiario o beneficiaria?",
        "insumos entregados categoria general": "Tratamiento",
        "especifique que se entrega detalle del insumo": "Especifique qué se entrega",
        "unidades entregadas cantidad en numero": "Unidades entregadas",
        "lugar de atencion chihuahua": "Lugar de atención",
        "lugar de atencion baja california": "Lugar de atención",
        "lugar de atencion baja california sur": "Lugar de atención",
        "lugar de atencion nuevo leon": "Lugar de atención",
        "lugar de atencion sonora": "Lugar de atención",
        "lugar de atencion otro": "Lugar de atención",
        "lugar bc": "Lugar de atención",
        "lugar b c": "Lugar de atención",
        "lugar b c s": "Lugar de atención",
        "lugar nl": "Lugar de atención",
        "toma de consentimiento antes de iniciar la consulta": "Toma de consentimiento antes de iniciar la consulta",
        "pertenece a alguna minora etnica": "Minoría",
        "pertence a alguna minora etnica": "Minoría",
        "originario": "Estado",
        "motivo referido": "Motivo Ref",
        "plan de tratamiento fisio": "Plan de Tratamiento",
        "laboratorio clinico": "Laboratorio Clínico",
        "coordenadas": "Coordenadas",
    }
    seed_label_by_alias = {_normalize_header(k): v for k, v in seed_label_by_alias.items() if _normalize_header(k)}

    labels_to_aliases: dict[str, list[str]] = {}
    for raw_header, freq in counts.items():
        if freq < 2:
            continue
        norm = _normalize_header(raw_header)
        if not norm or norm in base_map:
            continue
        target_label = seed_label_by_alias.get(norm)
        if not target_label:
            close = difflib.get_close_matches(norm, alias_keys, n=1, cutoff=0.9)
            if close:
                target_label = base_map.get(close[0])
        if target_label:
            labels_to_aliases.setdefault(target_label, []).append(raw_header)

    if not labels_to_aliases:
        return schema

    enriched: list[dict] = []
    for col in schema:
        cur = dict(col)
        label = str(cur.get("label") or "")
        extra = labels_to_aliases.get(label, [])
        if extra:
            aliases = list(cur.get("aliases") or [])
            aliases.extend(extra)
            seen: set[str] = set()
            uniq: list[str] = []
            for a in aliases:
                nk = _normalize_header(str(a))
                if not nk or nk in seen:
                    continue
                seen.add(nk)
                uniq.append(str(a))
            cur["aliases"] = uniq
        enriched.append(cur)
    return enriched


KOBO_VISIBLE_COLUMNS = _enrich_kobo_aliases_from_validated(KOBO_VISIBLE_COLUMNS)
KOBO_ALIAS_TO_LABEL = _build_kobo_alias_map()

# Encabezados frecuentes en plantillas regionales (p. ej. columnas "Lugar …" partidas) → nombre Kobo
# unificado SOLO para la UI del editor; el archivo y las claves de fila siguen usando el nombre real.
_SHEET_HEADER_DISPLAY_OVERRIDES: dict[str, str] = {
    "lugar sonora": "Lugar de atención",
    "lugar nuevo leon": "Lugar de atención",
    "lugar chihuahua": "Lugar de atención",
    "lugar otro": "Lugar de atención",
    "lugar b c": "Lugar de atención",
    "lugar b c s": "Lugar de atención",
    "lugar bc": "Lugar de atención",
    "lugar nl": "Lugar de atención",
    "primera vez seg": "Primera vez o Seguimiento",
    "asesoria previa hoy": "¿Se le ha brindado asesoría en uno de los módulos el día de hoy?",
    "referencia": "Se hizo referencia",
    "referencia?": "Se hizo referencia",
    "a donde": "A dónde",
    "motivo referencia": "Motivo Ref",
    "motivo ref": "Motivo Ref",
    "entrega de tratamiento": "¿Se hizo entrega de tratamiento/artículos al beneficiario o beneficiaria?",
    "entrega trat": "¿Se hizo entrega de tratamiento/artículos al beneficiario o beneficiaria?",
    "entrega de trat": "¿Se hizo entrega de tratamiento/artículos al beneficiario o beneficiaria?",
    "insumos entregados categoria general": "Tratamiento",
    "especifique que se entrega detalle del insumo": "Especifique qué se entrega",
    "unidades entregadas cantidad en numero": "Unidades entregadas",
    "diag lab": "Diagnóstico / Resultados Laboratorio",
    "diag resultados lab": "Diagnóstico / Resultados Laboratorio",
    "diagnostico fisio": "Fisioterapia",
    "diag fisio": "Fisioterapia",
    "diag fisio terapia": "Fisioterapia",
    "diagnosticos": "Diagnóstico Medicina General",
    "diagnosticos medicina general": "Diagnóstico Medicina General",
}


def _sheet_column_display_map(columns: list[str]) -> dict[str, str]:
    """
    Mapa nombre real de columna (en archivo) → etiqueta canónica Kobo para mostrar en cabeceras.
    Si no hay mapeo o el nombre ya coincide, la clave no entra (el cliente usa el nombre real).
    """
    out: dict[str, str] = {}
    for col in columns:
        c = str(col or "").strip()
        if not c:
            continue
        key = _normalize_header(c)
        label = KOBO_ALIAS_TO_LABEL.get(key) or _SHEET_HEADER_DISPLAY_OVERRIDES.get(key)
        if label and label != c:
            out[c] = label
    return out


def _value_looks_missing(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    return _norm_text(text) in {
        "n d",
        "nd",
        "n/d",
        "n.d",
        "n.d.",
        "na",
        "n a",
        "s d",
        "s/d",
        "no aplica",
        "no disponible",
        "sin dato",
        "none",
        "null",
    }


def _label_to_present_columns(columns: list[str]) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for raw in columns:
        col = str(raw or "").strip()
        if not col:
            continue
        key = _normalize_header(col)
        label = KOBO_ALIAS_TO_LABEL.get(key)
        if not label:
            continue
        out.setdefault(label, []).append(col)
    return out


def _is_valid_date_yyyy_mm_dd(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    if re.match(r"^\d{4}-\d{2}-\d{2}(?:\s+00:00:00)?$", text):
        return True
    return False


def _normalize_set(values: list[str]) -> set[str]:
    return {_normalize_header(v) for v in values if _normalize_header(v)}


def _tokenize_option_value(value: str) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    tokens = [p.strip() for p in re.split(r"\|\|\||[,;]+", raw) if p.strip()]
    return tokens if tokens else [raw]


def _display_service_label(service_key: str) -> str:
    labels = {
        "medicina general": "Medicina General",
        "dental": "Dental",
        "fisioterapia": "Fisioterapia",
        "oftalmologia": "Oftalmología",
        "laboratorios": "Laboratorios",
    }
    return labels.get(service_key, service_key.title())


def _build_simple_check(columns: list[str], rows: list[dict[str, str]]) -> dict:
    labels_required = [c["label"] for c in KOBO_VISIBLE_COLUMNS if c.get("required")]
    matched_by_label = _label_to_present_columns(columns)
    missing_required_columns = [lbl for lbl in labels_required if lbl not in matched_by_label]

    rows_missing_required: list[dict] = []
    for i, row in enumerate(rows):
        missing_labels: list[str] = []
        for lbl in labels_required:
            col_names = matched_by_label.get(lbl, [])
            has_value = False
            for col in col_names:
                if not _value_looks_missing(row.get(col, "")):
                    has_value = True
                    break
            if not has_value:
                missing_labels.append(lbl)
        if missing_labels:
            rows_missing_required.append({
                "row_index": i,
                "excel_row": i + 2,
                "missing_labels": missing_labels,
            })

    invalid_date_rows: list[int] = []
    date_col = _find_column_name_by_alias(
        columns,
        ["Fecha de atención", "Fecha_de_atenci_n", "Fecha", "Fecha Atención", "Fecha atencion"],
    )
    if date_col:
        for i, row in enumerate(rows):
            v = str(row.get(date_col, "") or "").strip()
            if not v:
                continue
            if not _is_valid_date_yyyy_mm_dd(v):
                invalid_date_rows.append(i + 2)

    invalid_coords_rows: list[int] = []
    coords_col = _find_column_name_by_alias(
        columns,
        ["Coordenadas", "Ubicación geográfica de la atención", "Ubicaci_n_geogr_fica_de_la_atenci_n"],
    )
    if coords_col:
        for i, row in enumerate(rows):
            v = str(row.get(coords_col, "") or "").strip()
            if not v:
                continue
            lat, lon = _parse_lat_lon_from_text(v)
            if not lat or not lon:
                invalid_coords_rows.append(i + 2)

    reference_destination_missing_rows: list[int] = []
    ref_col = _find_column_name_by_alias(
        columns,
        ["Se hizo referencia", "¿Se hizo referencia?", "Referencia"],
    )
    where_col = _find_column_name_by_alias(
        columns,
        ["A dónde", "A donde", "¿A dónde?", "Referencia_donde"],
    )
    if ref_col and where_col:
        for i, row in enumerate(rows):
            ref_val = _normalize_header(str(row.get(ref_col, "") or ""))
            if ref_val in {"si", "sí"}:
                if _value_looks_missing(row.get(where_col, "")):
                    reference_destination_missing_rows.append(i + 2)

    def _motivo_ref_value_is_otro(value: str) -> bool:
        raw = str(value or "").strip()
        if not raw:
            return False
        for tok in _tokenize_option_value(raw):
            if _normalize_header(tok) == "otro":
                return True
        return _normalize_header(raw) == "otro"

    fisio_motivo_ref_detail_rows: list[int] = []
    motivo_ref_col = _find_column_name_by_alias(
        columns,
        ["Motivo Ref", "Motivo referencia", "Motivo_referencia", "Motivo Referido", "Motivo referido"],
    )
    motivo_esp_col = _find_column_name_by_alias(
        columns,
        [
            "Especificar (motivo referido)",
            "Especificar motivo referencia",
            "Especificar m. ref. fisioterapia",
            "Especificar m. ref. fisio",
            "Motivo_especificar",
        ],
    )
    service_col_simple = _find_column_name_by_alias(
        columns,
        ["Servicio que se brinda", "Servicio", "Especialidad", "Servicio_que_se_brinda"],
    )
    if ref_col and motivo_ref_col and service_col_simple and motivo_esp_col:
        for i, row in enumerate(rows):
            raw_svc = str(row.get(service_col_simple, "") or "").strip()
            if not raw_svc or _normalize_service(raw_svc) != "Fisioterapia":
                continue
            ref_val2 = _normalize_header(str(row.get(ref_col, "") or ""))
            if ref_val2 not in {"si", "sí"}:
                continue
            if not _motivo_ref_value_is_otro(str(row.get(motivo_ref_col, "") or "")):
                continue
            if _value_looks_missing(row.get(motivo_esp_col, "")):
                fisio_motivo_ref_detail_rows.append(i + 2)

    invalid_option_rows: list[dict] = []
    ref_col_for_options = _find_column_name_by_alias(
        columns,
        ["Se hizo referencia", "¿Se hizo referencia?", "Referencia"],
    )
    where_aliases = {_normalize_header(v) for v in ["A dónde", "A donde", "¿A dónde?", "Referencia_donde"]}
    def _preferred_columns_for_label(label: str, candidate_cols: list[str]) -> list[str]:
        """
        Si existe columna con nombre exacto del label, usar solo esa para validar opciones.
        Evita falsos positivos cuando hay columnas alias históricas (ej. Diagnósticos).
        """
        target = _normalize_header(label)
        exact = [c for c in candidate_cols if _normalize_header(c) == target]
        if exact:
            return exact[:1]
        return candidate_cols

    for schema_col in KOBO_VISIBLE_COLUMNS:
        options = schema_col.get("options") or []
        if not options:
            continue
        label = str(schema_col.get("label") or "")
        candidate_cols = matched_by_label.get(label, [])
        if not candidate_cols:
            continue
        candidate_cols = _preferred_columns_for_label(label, candidate_cols)
        valid_set = _normalize_set([str(o) for o in options])
        if not valid_set:
            continue
        for col in candidate_cols:
            bad_rows: list[int] = []
            for i, row in enumerate(rows):
                # Regla condicional: "A dónde" solo se valida cuando "Se hizo referencia" = Sí.
                if _normalize_header(col) in where_aliases and ref_col_for_options:
                    ref_val = _normalize_header(str(row.get(ref_col_for_options, "") or ""))
                    if ref_val not in {"si", "sí"}:
                        continue
                val = str(row.get(col, "") or "").strip()
                if _value_looks_missing(val):
                    continue
                tokens = _tokenize_option_value(val)
                if not tokens:
                    continue
                if any(_normalize_header(t) not in valid_set for t in tokens):
                    bad_rows.append(i + 2)
            if bad_rows:
                invalid_option_rows.append({
                    "label": label,
                    "column": col,
                    "rows": bad_rows[:20],
                    "count": len(bad_rows),
                })

    services = _services_present_in_rows(columns, rows)
    service_missing_columns: list[dict] = []
    existing_keys = {_normalize_header(c) for c in columns}
    for service in sorted(services):
        expected = SERVICE_CONDITIONAL_COLUMNS.get(service, [])
        missing_for_service = [
            lbl for lbl in expected if _normalize_header(lbl) not in existing_keys
        ]
        if missing_for_service:
            service_missing_columns.append({
                "service": _display_service_label(service),
                "missing_columns": missing_for_service,
            })

    ready = (
        len(missing_required_columns) == 0
        and len(rows_missing_required) == 0
        and len(invalid_date_rows) == 0
        and len(invalid_coords_rows) == 0
        and len(invalid_option_rows) == 0
        and len(reference_destination_missing_rows) == 0
        and len(fisio_motivo_ref_detail_rows) == 0
        and len(service_missing_columns) == 0
    )

    return {
        "ready_to_submit": ready,
        "missing_required_columns": missing_required_columns,
        "rows_with_missing_required": {
            "count": len(rows_missing_required),
            "sample": rows_missing_required[:20],
        },
        "invalid_formats": {
            "date_rows": invalid_date_rows[:30],
            "coordinates_rows": invalid_coords_rows[:30],
            "option_rows": invalid_option_rows[:20],
            "reference_destination_rows": reference_destination_missing_rows[:30],
            "fisio_motivo_ref_detail_rows": fisio_motivo_ref_detail_rows[:30],
        },
        "service_conditional_missing": service_missing_columns,
        "human_message": (
            "Listo para enviar a Kobo."
            if ready
            else "Faltan datos por corregir antes de enviar a Kobo."
        ),
    }


def _check_columns_against_kobo(columns: list[str], rows: list[dict[str, str]] | None = None) -> dict:
    labels_required = [c["label"] for c in KOBO_VISIBLE_COLUMNS if c.get("required")]
    labels_all = [c["label"] for c in KOBO_VISIBLE_COLUMNS]
    matched: dict[str, list[str]] = {}
    unknown: list[str] = []
    suggestions: dict[str, str] = {}

    alias_keys = list(KOBO_ALIAS_TO_LABEL.keys())
    for raw in columns:
        col = str(raw or "").strip()
        if not col:
            continue
        key = _normalize_header(col)
        label = KOBO_ALIAS_TO_LABEL.get(key)
        if label:
            matched.setdefault(label, []).append(col)
            continue
        unknown.append(col)
        if key:
            # 0,72 confundía CONS1 (inicial) con CONS (verbal) por compartir «consentimiento».
            close = difflib.get_close_matches(key, alias_keys, n=1, cutoff=0.88)
            if close:
                sugg_lbl = KOBO_ALIAS_TO_LABEL[close[0]]
                n_sugg = _normalize_header(sugg_lbl)
                # No cruzar los dos campos de consentimiento.
                if "inici" in key and "verbal" in n_sugg:
                    pass
                elif "verbal" in key and "inici" in n_sugg:
                    pass
                else:
                    suggestions[col] = sugg_lbl

    missing_required = [lbl for lbl in labels_required if lbl not in matched]
    duplicates = [
        {"label": lbl, "columns": cols}
        for lbl, cols in matched.items()
        if len(cols) > 1
    ]

    rows_for_simple = rows or []
    simple_check = _build_simple_check(columns, rows_for_simple)

    return {
        "ok_required": len(missing_required) == 0,
        "required_total": len(labels_required),
        "required_present": len(labels_required) - len(missing_required),
        "missing_required_columns": missing_required,
        "unknown_columns": unknown,
        "duplicates": duplicates,
        "recognized_labels": sorted(matched.keys()),
        "rename_suggestions": suggestions,
        "kobo_schema": KOBO_VISIBLE_COLUMNS,
        "kobo_schema_source": KOBO_SCHEMA_SOURCE,
        "kobo_labels": labels_all,
        "simple_check": simple_check,
    }


def _parse_simple_mapping_yaml(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    if not path.exists():
        return out
    try:
        for raw in path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if ":" not in line:
                continue
            key, val = line.split(":", 1)
            k = key.strip()
            v = val.strip()
            if not k or not v:
                continue
            if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
                v = v[1:-1]
            out[k] = v
    except Exception:
        return {}
    return out


def _kobo_internal_alias_map() -> dict[str, str]:
    """
    Mapa de alias visibles -> clave interna para construcción de payload Kobo.
    """
    out: dict[str, str] = {}
    for col in KOBO_VISIBLE_COLUMNS:
        internal = str(col.get("internal") or "").strip()
        if not internal:
            continue
        aliases = list(col.get("aliases") or [])
        aliases.append(str(col.get("label") or ""))
        aliases.append(internal)
        for a in aliases:
            key = _normalize_header(str(a))
            if key and key not in out:
                out[key] = internal
    return out


KOBO_INTERNAL_ALIAS_TO_KEY = _kobo_internal_alias_map()

# Subcolumnas tipo export Kobo: «... discapacidades/Motriz» = 0 o 1
_DIS_TAIL_TO_INTERNAL: dict[str, str] = {
    "motriz": "DIS_motriz",
    "visual": "DIS_visual",
    "auditiva": "DIS_auditiva",
    "intelectual": "DIS_intelectual",
    "otra": "DIS_otra",
}
# Subcolumnas export: «Diagnóstico/Artrosis» = 0/1; misma lógica que en filling_rules (claves FISIO_*)
_FISIO_DIAG_TAIL_TO_INTERNAL: dict[str, str] = {
    "revision": "FISIO_Revision",
    "artrosis": "FISIO_Artrosis",
    "artritis": "FISIO_Artritis",
    "lesionesmusculoesqueleticas": "FISIO_Lesiones_musculoesqueleticas",
    "dolorcronico": "FISIO_Dolor_cronico",
    "dolorcronic": "FISIO_Dolor_cronico",
    "enfermedadesneurologicas": "FISIO_Enf_neurologicas",
    "problemasrespiratorios": "FISIO_Problemas_respiratorios",
    "dolor": "FISIO_Dolor",
    "contractura": "FISIO_Contractura",
    "otro": "FISIO_Otro",
}
# Fisi: «Localización de la lesión/Cabeza» → 0/1
_LOC_LESION_TAIL_TO_INTERNAL: dict[str, str] = {
    "cabeza": "LOC_Cabeza",
    "cuello": "LOC_Cuello",
    "torax": "LOC_Torax",
    "abdomen": "LOC_Abdomen",
    "cadera": "LOC_Cadera",
    "miembrosuperiorizquierdo": "LOC_MSI",
    "miembrosuperiorderecho": "LOC_MSD",
    "miembroinferiorizquierdo": "LOC_MII",
    "miembroinferiorderecho": "LOC_MID",
    "espalda": "LOC_Espalda",
    "otro": "LOC_Otro",
}
# Dental: «Diagnóstico/Caries» (colas disjuntas con fisi; «otro» ambiguo 2 tramos → DENT; ruta 3+ con fisi/odont)
_DENT_DIAG_TAIL_TO_INTERNAL: dict[str, str] = {
    "revisionderutina": "DENT_Revision_rutina",
    "caries": "DENT_Caries",
    "calculo": "DENT_Calculo",
    "sarro": "DENT_Sarro",
    "gingivitis": "DENT_Gingivitis",
    "periodontitis": "DENT_Periodontitis",
    "pulpitisreversible": "DENT_Pulpitis_reversible",
    "pulpitisirreversible": "DENT_Pulpitis_irreversible",
    "fractura": "DENT_Fractura",
    "filtracion": "DENT_Filtracion",
    "desgaste": "DENT_Desgaste",
    "infeccion": "DENT_Infeccion",
    "otro": "DENT_Otro",
}
_OFT_SINTOMAS_TAIL_TO_INTERNAL: dict[str, str] = {
    "ninguno": "OFT_SX_Ninguno",
    "ardor": "OFT_SX_Ardor",
    "comezon": "OFT_SX_Comezon",
    "irritacion": "OFT_SX_Irritacion",
    "lagrimeo": "OFT_SX_Lagrimeo",
    "fotofobia": "OFT_SX_Fotofobia",
    "dificultadparaleer": "OFT_SX_Dif_leer",
    "disminuciondevision": "OFT_SX_Dism_vision",
    "vistacansada": "OFT_SX_Vista_cansada",
    "dolor": "OFT_SX_Dolor",
    "visionborrosa": "OFT_SX_Vision_borrosa",
    "otro": "OFT_SX_Otro",
}
_OFT_DXPREV_TAIL_TO_INTERNAL: dict[str, str] = {
    "ninguno": "OFT_PREV_Ninguno",
    "catarata": "OFT_PREV_Catarata",
    "glaucoma": "OFT_PREV_Glaucoma",
    "estrabismo": "OFT_PREV_Estrabismo",
    "retinopatia": "OFT_PREV_Retinopatia",
    "pterigion": "OFT_PREV_Pterigion",
    "presbicia": "OFT_PREV_Presbicia",
    "otro": "OFT_PREV_Otro",
}
_OFT_DXACTUAL_TAIL_TO_INTERNAL: dict[str, str] = {
    "revision": "OFT_ACT_Revision",
    "conjuntivitis": "OFT_ACT_Conjuntivitis",
    "ametropia": "OFT_ACT_Ametropia",
    "miopia": "OFT_ACT_Miopia",
    "astigmatismo": "OFT_ACT_Astigmatismo",
    "hipermetropia": "OFT_ACT_Hipermetropia",
    "presbicia": "OFT_ACT_Presbicia",
    "estabismo": "OFT_ACT_Estrabismo",
    "estrabismo": "OFT_ACT_Estrabismo",
    "cataratas": "OFT_ACT_Cataratas",
    "glaucoma": "OFT_ACT_Glaucoma",
    "pterigion": "OFT_ACT_Pterigion",
    "ojoseco": "OFT_ACT_Ojo_seco",
    "otro": "OFT_ACT_Otro",
}
# «¿Qupe/Qué procedimiento se realiza?/Resina» — mezcla dental/oft; mismo grupo kobo
_PROCEDIMIENTO_TAIL_TO_INTERNAL: dict[str, str] = {
    "resina": "PROC_Resina",
    "limpiezadental": "PROC_Limpieza_dental",
    "endodoncia": "PROC_Endodoncia",
    "extraccion": "PROC_Extraccion",
    "cirugia": "PROC_Cirugia",
    "presionarterial": "PROC_Presion_arterial",
    "rx": "PROC_RX",
    "ortodoncia": "PROC_Ortodoncia",
    "colocaciondeprotesis": "PROC_Protesis",
    "colocaciondefluor": "PROC_Fluor",
    "colocaciondeeugenol": "PROC_Eugenol",
    "tomadeimpresion": "PROC_Toma_impresion",
    "chalazion": "PROC_Chalazion",
    "dilataciondepupila": "PROC_Dilatacion_pupila",
    "otro": "PROC_Otro",
}
# «Especifique qué se entrega/Medicamento/suplemento» = 0/1 (toda la ruta bajo el padre)
_ESPEC_ENT_SUB_TO_INTERNAL: dict[str, str] = {
    "medicamentossuplemento": "ESPEC_ENT_Med_sup",
    "anteojos": "ESPEC_ENT_Anteojos",
    "plandetratamiento": "ESPEC_ENT_Plan",
    "insumosdewash": "ESPEC_ENT_WASH",
    "insumosdehigienedental": "ESPEC_ENT_Higiene_dental",
    "insumosdehigienebucal": "ESPEC_ENT_Higiene_dental",  # alias
    "resultadosdelaboratorio": "ESPEC_ENT_Lab",
    "otro": "ESPEC_ENT_Otro",
}
_FISIO_ONLY_DIAG_TAILS: frozenset[str] = frozenset(
    {
        "artrosis",
        "artritis",
        "lesionesmusculoesqueleticas",
        "dolorcronico",
        "dolorcronic",
        "enfermedadesneurologicas",
        "problemasrespiratorios",
        "contractura",
    }
)
_DENT_ONLY_DIAG_TAILS: frozenset[str] = frozenset(
    {
        "revisionderutina",
        "caries",
        "calculo",
        "sarro",
        "gingivitis",
        "periodontitis",
        "pulpitisreversible",
        "pulpitisirreversible",
        "filtracion",
        "desgaste",
        "infeccion",
        "fractura",
    }
)


def _path_last_seg_norm(raw_key: str) -> tuple[str, str, str, list[str]]:
    """
    (full_key_norm, parent_joined_norm, last_seg_norm, parts) para enrutar subcolumnas.
    """
    s = str(raw_key or "").strip()
    parts = [p.strip() for p in s.split("/") if p.strip()]
    if len(parts) < 2:
        return "", "", "", parts
    last = parts[-1]
    parent_joined = "/".join(parts[:-1])
    full = s
    return (
        re.sub(r"[^a-z0-9]+", "", _norm_text(full)),
        re.sub(r"[^a-z0-9]+", "", _norm_text(parent_joined)),
        re.sub(r"[^a-z0-9]+", "", _norm_text(last)),
        parts,
    )


def _disability_subcolumn_key_to_internal(raw_key: str) -> str:
    """Mapea cabecera estilo Kobo `.../Motriz` a clave interna `DIS_motriz` (valores 0/1 o Sí/No)."""
    s = str(raw_key or "").strip()
    if "/" not in s:
        return ""
    last = s.rsplit("/", 1)[-1].strip()
    t = re.sub(r"[^a-z0-9]+", "", _norm_text(last))
    return _DIS_TAIL_TO_INTERNAL.get(t, "")


def _cell_affirmative_dis_yes(v: object) -> bool:
    """Misma lógica que filling_rules._cell_affirmative_dis (Sí/1 en flags DIS_*)."""
    s = str(v or "").strip().lower()
    return s in ("1", "sí", "si", "yes", "y", "true", "x", "s")


def _disability_multiselect_cell_is_placeholder(v: object) -> bool:
    """Valores de export/histórico que equivalen a «sin respuesta» en la columna multiselect DIS."""
    t = str(v or "").strip()
    if not t:
        return True
    nk = _normalize_header(t)
    return nk in {
        "0",
        "no",
        "n",
        "false",
        "nd",
        "n d",
        "na",
        "n a",
        "noaplica",
        "no aplica",
        "ninguno",
        "ninguna",
        "sin dato",
        "s d",
    }


def _is_disability_multiselect_sheet_column(col: str) -> bool:
    nk = _normalize_header(str(col or ""))
    return nk in {
        "indicar si el paciente tiene alguna de las siguientes discapacidades",
        "discapacidad",
        "tipo de discapacidad",
        "dis",
    }


def _sanitize_disability_sheet_display(
    columns: list[str], rows: list[dict[str, str]]
) -> tuple[list[str], list[dict[str, str]]]:
    """
    Muestra discapacidad sin «relleno» de export Kobo: subcolumnas …/Motriz en 0/No → vacío;
    en la columna multiselect, solo placeholders (0, N/D, ninguno, etc.) → vacío.
    Así el usuario ve celdas vacías y captura solo si aplica.
    """
    if not rows:
        return columns, rows
    targets = [
        c
        for c in columns
        if _is_disability_multiselect_sheet_column(str(c)) or _disability_subcolumn_key_to_internal(str(c))
    ]
    if not targets:
        return columns, rows
    out_rows: list[dict[str, str]] = []
    for row in rows:
        nr = dict(row)
        for c in targets:
            cur = nr.get(c, "")
            if _disability_subcolumn_key_to_internal(str(c)):
                if not _cell_affirmative_dis_yes(cur):
                    nr[c] = ""
            elif _disability_multiselect_cell_is_placeholder(cur):
                nr[c] = ""
        out_rows.append(nr)
    return columns, out_rows


def _fisio_diag_tail_fuzzy(t: str) -> str:
    if t.startswith("revisi") and "rutina" not in t:
        return "FISIO_Revision"
    if t.startswith("lesion") and "musc" in t:
        return "FISIO_Lesiones_musculoesqueleticas"
    if "dolor" in t and "cron" in t:
        return "FISIO_Dolor_cronico"
    if t in ("dolor", "dolo"):
        return "FISIO_Dolor"
    if "enfer" in t and "neurol" in t:
        return "FISIO_Enf_neurologicas"
    if "proble" in t and "resp" in t:
        return "FISIO_Problemas_respiratorios"
    return ""


def _fisio_diag_path_to_internal(raw_key: str) -> str:
    """Compat. — usa el router completo (incl. Diagnóstico dental + fisi)."""
    return _kobo_binary_subcolumn_to_internal(raw_key)


def _kobo_binary_subcolumn_to_internal(raw_key: str) -> str:
    """
    Mapea export Kobo (subcolumnas 0/1) → clave interna: LOC_*, DENT_*, FISIO_*,
    OFT_*, PROC_*, ESPEC_ENT_*.
    """
    _full, pnorm, t, parts = _path_last_seg_norm(raw_key)
    if len(parts) < 2:
        return ""
    p0n = re.sub(r"[^a-z0-9]+", "", _norm_text(parts[0]))
    if p0n.startswith("especif") and "entreg" in p0n and "se" in p0n:
        subj = re.sub(r"[^a-z0-9]+", "", _norm_text("/".join(parts[1:])))
        if subj:
            return _ESPEC_ENT_SUB_TO_INTERNAL.get(subj, "")
        return ""
    if not t:
        return ""
    nseg = len(parts)

    if "local" in pnorm and "lesi" in pnorm:
        return _LOC_LESION_TAIL_TO_INTERNAL.get(t, "")

    if "sintoma" in pnorm and "present" in pnorm and "fecha" in pnorm:
        return _OFT_SINTOMAS_TAIL_TO_INTERNAL.get(t, "")

    if "recib" in pnorm and "diagn" in pnorm and "prev" in pnorm:
        return _OFT_DXPREV_TAIL_TO_INTERNAL.get(t, "")

    if "diagnostic" in pnorm and "actual" in pnorm:
        return _OFT_DXACTUAL_TAIL_TO_INTERNAL.get(t, "")

    if "proced" in pnorm or "qupe" in pnorm:
        d = _PROCEDIMIENTO_TAIL_TO_INTERNAL.get(t, "")
        if d:
            return d
        if "tom" in t and "imp" in t:
            return "PROC_Toma_impresion"
        return ""

    if "diagn" in pnorm and "prev" not in pnorm and "recib" not in pnorm and "actual" not in pnorm:
        pfull = pnorm
        if t == "otro" and nseg == 2:
            return "DENT_Otro"
        if t == "otro" and nseg >= 3 and "fisio" in pfull:
            return "FISIO_Otro"
        if t == "otro" and nseg >= 3 and ("odont" in pfull or "dental" in pfull):
            return "DENT_Otro"
        if nseg >= 3 and ("odont" in pfull or "dental" in p0n):
            return _DENT_DIAG_TAIL_TO_INTERNAL.get(t, "")
        if nseg >= 3 and "fisio" in pfull:
            return _FISIO_DIAG_TAIL_TO_INTERNAL.get(t, "") or _fisio_diag_tail_fuzzy(t)
        if t in _DENT_ONLY_DIAG_TAILS:
            return _DENT_DIAG_TAIL_TO_INTERNAL.get(t, "")
        if t in _FISIO_ONLY_DIAG_TAILS or (t in _FISIO_DIAG_TAIL_TO_INTERNAL and t != "otro"):
            return _FISIO_DIAG_TAIL_TO_INTERNAL.get(t, "") or _fisio_diag_tail_fuzzy(t)
        f_val = _FISIO_DIAG_TAIL_TO_INTERNAL.get(t, "") or _fisio_diag_tail_fuzzy(t)
        d_val = _DENT_DIAG_TAIL_TO_INTERNAL.get(t, "")
        if t in _DENT_ONLY_DIAG_TAILS:
            return d_val
        if d_val and t != "otro" and t not in _FISIO_DIAG_TAIL_TO_INTERNAL:
            return d_val
        if f_val:
            return f_val
        return d_val or ""
    return ""


def _coerce_cell_binary_01(raw: str) -> str:
    """
    Celdas binarias API/export: vacío → 0; 1, Sí, si, x… → 1; 0, No… → 0.
    """
    t = str(raw or "").strip().lower()
    if not t:
        return "0"
    if t in ("1", "sí", "si", "yes", "y", "true", "x", "s"):
        return "1"
    if t in ("0", "no", "n", "false"):
        return "0"
    if t in ("0", "1"):
        return t
    return "0"


def _to_kobo_internal_record(row: dict[str, str]) -> dict[str, str]:
    rec: dict[str, str] = {}
    for raw_k, raw_v in (row or {}).items():
        k = str(raw_k or "").strip()
        if not k:
            continue
        v = str(raw_v or "").strip()
        nk = _normalize_header(k)
        dint = _disability_subcolumn_key_to_internal(k)
        bint = _kobo_binary_subcolumn_to_internal(k)
        internal = KOBO_INTERNAL_ALIAS_TO_KEY.get(nk) or None
        if not internal:
            if nk in {
                "fisioterapia",
                "diagnostico fisioterapia",
                "diagnostico en fisioterapia",
                "diagnostico fisiopatia",
                "diag fisi",
                "diag fisi terapia",
            }:
                internal = "Fisio_Diagnostico"
            elif nk in {"diagnostico medicina general", "diagnosticos medicina general", "diagnostico odontologia"}:
                internal = "Diagnostico_Motivo"
            elif nk in {
                "plan de tratamiento",
                "plan de tratamiento fisioterapia u otros",
                "plan tratamiento",
                "plan fisio",
                "plan de tratamiento solo fisio o entrega de receda sin medicamento",
            }:
                internal = "Plan_de_Tratamiento"
            elif nk in {"indicar si el paciente tiene alguna de las siguientes discapacidades", "discapacidad", "tipo de discapacidad"}:
                internal = "Discapacidad"
            elif nk in {"se hizo entrega de tratamiento articulos al beneficiario o beneficiaria", "se hizo entrega de tratamiento al beneficiario"}:
                internal = "entrega_tx"
            elif nk in {
                "especifique que se entrega",
                "especifique qu se entrega",
                "especifique qué se entrega",
                "especifique qu se entrega detalle del insumo",
                "especifique que se entrega detalle del insumo",
                "especifique qué se entrega detalle del insumo",
                "especifique_qu_se_entrega",
            }:
                internal = "Especifique_qu_se_entrega"
            elif nk in {
                "especificar lo que se entrega al beneficiario",
                "especificar lo que se entrega",
                "especificar lo que se entrega al benefic",
                "especificar_lo_que_se_entrega_",
            }:
                internal = "Especificar_lo_que_se_entrega_"
            elif nk in {"unidades entregadas", "unidades_entregadas"}:
                internal = "Unidades_entregadas"
            else:
                if dint:
                    internal = dint
                else:
                    if bint:
                        internal = bint
                    elif re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", k):
                        internal = k
        use_binary_default = bool(dint or bint) or (internal in ("SUP_Hierro", "SUP_Acido_folico"))
        if not internal and not v:
            continue
        if not use_binary_default and not v:
            continue
        v_out: str
        if use_binary_default:
            v_out = _coerce_cell_binary_01(v)
        else:
            v_out = v
        if internal and internal not in rec:
            rec[internal] = v_out
    return rec


def _build_row_dict(headers: list[str], values: list[object]) -> dict[str, str]:
    row: dict[str, str] = {}
    for idx, header in enumerate(headers):
        cell = values[idx] if idx < len(values) else ""
        row[header] = _stringify_cell_value(header, cell)
    return row


def _is_age_column(header: str) -> bool:
    h = _normalize_header(header)
    return h in {"edad", "age"}


def _stringify_cell_value(header: str, cell: object) -> str:
    if cell is None:
        return ""
    if _is_age_column(header):
        # Evita mostrar "31.0" en Edad; mantiene decimales reales si existen.
        if isinstance(cell, float) and cell.is_integer():
            return str(int(cell))
        text = str(cell).strip()
        if re.fullmatch(r"\d+\.0+", text):
            return text.split(".", 1)[0]
        return text
    return str(cell).strip()


def _parse_lat_lon_from_text(raw: str) -> tuple[str, str]:
    text = str(raw or "").strip()
    if not text:
        return "", ""
    # Acepta formatos comunes: "lat,lon", "lat;lon", "lat lon"
    parts = [p.strip() for p in re.split(r"[,\s;]+", text) if p.strip()]
    if len(parts) < 2:
        return "", ""
    return parts[0], parts[1]


def _ensure_lat_lon_from_coordinates(columns: list[str], rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    if not columns:
        return columns, rows

    coord_idx = -1
    coord_name = ""
    for i, c in enumerate(columns):
        if _normalize_header(c) in {"coordenadas", "coordenada", "coordenadas gps", "gps"}:
            coord_idx = i
            coord_name = c
            break
    if coord_idx < 0:
        return columns, rows

    has_lat = any(_normalize_header(c) == "latitud" for c in columns)
    has_lon = any(_normalize_header(c) == "longitud" for c in columns)
    if has_lat and has_lon:
        return columns, rows

    out_cols = list(columns)
    insert_at = coord_idx + 1
    if not has_lat:
        out_cols.insert(insert_at, "Latitud")
        insert_at += 1
    if not has_lon:
        out_cols.insert(insert_at, "Longitud")

    out_rows: list[dict[str, str]] = []
    for row in rows:
        new_row = dict(row)
        lat, lon = _parse_lat_lon_from_text(row.get(coord_name, ""))
        if not has_lat:
            new_row["Latitud"] = lat
        if not has_lon:
            new_row["Longitud"] = lon
        out_rows.append(new_row)
    return out_cols, out_rows


def _ensure_lugar_atencion_by_estado(columns: list[str], rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    estado_col = _find_column_name_by_alias(
        columns,
        ["Estado", "Estado brigada", "Estado_brigada", "Estado de origen", "Estado paciente", "Estado_paciente"],
    )
    if not estado_col:
        return columns, rows

    lugar_col = _find_column_name_by_alias(columns, ["Lugar de atención", "Lugar de atencion", "Lugar"])
    out_cols = list(columns)
    out_rows = [dict(r or {}) for r in rows]
    if not lugar_col:
        lugar_col = "Lugar de atención"
        out_cols.append(lugar_col)
        for r in out_rows:
            r.setdefault(lugar_col, "")

    for row in out_rows:
        estado = _normalize_header(str(row.get(estado_col, "") or ""))
        lugar = str(row.get(lugar_col, "") or "").strip()
        if estado == "nuevo leon" and not lugar:
            row[lugar_col] = "Montemorelos"
    return out_cols, out_rows


def _lugar_atencion_alias_keys() -> set[str]:
    keys: set[str] = set()
    for col in KOBO_VISIBLE_COLUMNS:
        if _normalize_header(str(col.get("label") or "")) != "lugar de atencion":
            continue
        aliases = list(col.get("aliases") or []) + [str(col.get("label") or "")]
        for a in aliases:
            nk = _normalize_header(str(a))
            if nk:
                keys.add(nk)
    return keys


LUGAR_ATENCION_ALIAS_KEYS = _lugar_atencion_alias_keys()
LUGAR_ATENCION_EMPTY_VALUES = {"no", "nd", "n d", "n/d", "n.d", "n.d."}


def _is_lugar_atencion_column_name(col_name: str) -> bool:
    raw = str(col_name or "").strip()
    if not raw:
        return False
    nk = _normalize_header(raw)
    if nk in LUGAR_ATENCION_ALIAS_KEYS:
        return True
    # Soporta columnas derivadas por estado, ej. "Lugar de atención: Chihuahua".
    if ":" in raw:
        base = _normalize_header(raw.split(":", 1)[0])
        if base in LUGAR_ATENCION_ALIAS_KEYS:
            return True
    return False


def _sanitize_lugar_atencion_missing_values(
    columns: list[str], rows: list[dict[str, str]]
) -> tuple[list[str], list[dict[str, str]]]:
    """
    Limpia valores no válidos de Lugar de atención:
    - "No"
    - "N/D" y variantes
    Con esto el campo queda vacío para validación Kobo.
    """
    lugar_cols = [c for c in columns if _is_lugar_atencion_column_name(c)]
    if not lugar_cols:
        return columns, rows

    out_rows: list[dict[str, str]] = []
    for row in rows:
        nr = dict(row or {})
        for col in lugar_cols:
            raw = str(nr.get(col, "") or "").strip()
            if not raw:
                continue
            if _norm_text(raw) in LUGAR_ATENCION_EMPTY_VALUES:
                nr[col] = ""
        out_rows.append(nr)
    return columns, out_rows


TRATAMIENTO_SOURCE_ALIAS_KEYS = {
    "medicamento",
    "medicamentos",
    "medicamentos no especificos",
    "medicamentos no especifico",
    "medicamentos no especficos",
    "medicamentos no especfico",
    "insumos entregados",
    "insumos entregados categoria general",
    "insumos entregados categoria",
}


def _is_tratamiento_source_column_name(col_name: str) -> bool:
    nk = _normalize_header(str(col_name or ""))
    if not nk:
        return False
    if nk in TRATAMIENTO_SOURCE_ALIAS_KEYS:
        return True
    # Cobertura para encabezados truncados/variables, p.ej.:
    # "Medicamentos (No..." o "Medicamentos no especificados"
    if "insumos entregados" in nk:
        return True
    if "medicamentos no" in nk:
        return True
    if "medicamento no" in nk:
        return True
    return False


def _merge_tratamiento_columns(
    columns: list[str], rows: list[dict[str, str]]
) -> tuple[list[str], list[dict[str, str]]]:
    """
    Regla KoboUp:
    - "Medicamentos" e "Insumos entregados" no se muestran como columnas separadas.
    - Su contenido se consolida en una única columna "Tratamiento".
    """
    source_cols = [c for c in columns if _is_tratamiento_source_column_name(c)]
    if not source_cols:
        return columns, rows

    out_cols = [c for c in columns if c not in source_cols]
    tratamiento_col = _find_column_name_by_alias(out_cols, ["Tratamiento"])
    if not tratamiento_col:
        tratamiento_col = "Tratamiento"
        out_cols.append(tratamiento_col)

    out_rows: list[dict[str, str]] = []
    for row in rows:
        nr = dict(row or {})
        merged_parts: list[str] = []

        current_tx = str(nr.get(tratamiento_col, "") or "").strip()
        if current_tx:
            merged_parts.append(current_tx)

        for src in source_cols:
            raw = str(nr.get(src, "") or "").strip()
            nr.pop(src, None)
            if not raw:
                continue
            if _norm_text(raw) in {"nd", "n d", "n/d", "n.d", "n.d.", "no"}:
                continue
            merged_parts.append(raw)

        if merged_parts:
            uniq: list[str] = []
            seen: set[str] = set()
            for part in merged_parts:
                key = _norm_text(part)
                if key and key in seen:
                    continue
                if key:
                    seen.add(key)
                uniq.append(part)
            nr[tratamiento_col] = " | ".join(uniq)
        else:
            nr.setdefault(tratamiento_col, "")

        out_rows.append(nr)

    return out_cols, out_rows


def _reorder_treatment_next_to_diagnostico(columns: list[str]) -> list[str]:
    """
    Deja "Tratamiento" junto a la primera columna de diagnóstico encontrada.
    """
    if not columns:
        return columns
    tratamiento_idx = None
    for i, col in enumerate(columns):
        if _normalize_header(str(col)) == "tratamiento":
            tratamiento_idx = i
            break
    if tratamiento_idx is None:
        return columns

    diag_idx = None
    for i, col in enumerate(columns):
        nk = _normalize_header(str(col))
        if nk.startswith("diagnostico"):
            diag_idx = i
            break
    if diag_idx is None:
        return columns

    out = list(columns)
    tratamiento_col = out.pop(tratamiento_idx)
    if tratamiento_idx < diag_idx:
        diag_idx -= 1
    out.insert(diag_idx + 1, tratamiento_col)
    return out


def _diagnostico_mg_options() -> list[str]:
    for col in KOBO_VISIBLE_COLUMNS:
        if _normalize_header(str(col.get("label") or "")) == "diagnostico medicina general":
            return [str(v).strip() for v in (col.get("options") or []) if str(v).strip()]
    return []


def _split_diag_tokens(raw: str) -> list[str]:
    text = str(raw or "").strip()
    if not text:
        return []
    # Soporta formatos como: "Otro: X", "A / B", "A;B", "A|||B"
    parts = [p.strip() for p in re.split(r"\|\|\||[,;/]+", text) if p and p.strip()]
    out: list[str] = []
    for p in parts:
        cleaned = re.sub(r"^\s*otro\s*[:\-]\s*", "", p, flags=re.IGNORECASE).strip()
        if cleaned:
            out.append(cleaned)
    return out if out else [text]


def _normalize_diagnostico_mg_values(columns: list[str], rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    diag_col = _find_column_name_by_alias(
        columns,
        [
            "Diagnóstico Medicina General",
            "Diagnostico Medicina General",
            "Diagnósticos Medicina General",
            "Diagnosticos Medicina General",
            "Diagnósticos",
            "Diagnosticos",
        ],
    )
    if not diag_col:
        return columns, rows

    options = _diagnostico_mg_options()
    if not options:
        return columns, rows
    # No buscar el encabezado genérico «Especificar»: en Kobo es el texto de NATOT
    # (nacionalidad), no el de diagnóstico (dxesp).
    especificar_col = _find_column_name_by_alias(
        columns,
        [
            "Especificar diagnóstico (Medicina General)",
            "Especificar diagnóstico",
            "Especificar Diagnóstico Medicina General",
            "Especificar Diagnostico Medicina General",
            "dxesp",
        ],
    )
    # Fuentes de respaldo: en varios archivos históricos el texto libre quedó en estas columnas.
    fallback_diag_sources: list[str] = []
    for alias in [
        "Padecimiento médico actual",
        "Padecimiento medico actual",
        "Diagnósticos",
        "Diagnosticos",
        "Diagnóstico Medicina General",
        "Diagnostico Medicina General",
    ]:
        c = _find_column_name_by_alias(columns, [alias])
        if c and c not in fallback_diag_sources and c != diag_col:
            fallback_diag_sources.append(c)
    option_map = {_normalize_header(opt): opt for opt in options}
    non_otro_keys = [k for k in option_map.keys() if k not in {"otro", "otra"}]
    placeholder_diag_keys = {
        "ninguno seleccionado",
        "ninguno",
        "sin diagnostico",
        "sin diagnosticos",
        "sin diagnostico seleccionado",
        "sin diagnosticos seleccionados",
        "na",
        "n a",
        "n d",
        "n/d",
        "nd",
    }
    # Alias comunes observados en archivos históricos.
    alias_map = {
        "cefaleas": "Cefalea",
        "cefalea": "Cefalea",
        "migrana": "Cefalea",
        "migraña": "Cefalea",
        "amigdalitis": "Amigdalitis",
        "otitis": "Otitis media",
        "diarrea": "Diarrea aguda",
    }

    out_cols = list(columns)
    out_rows: list[dict[str, str]] = []
    # Clave temporal: no usar «Especificar» (en Kobo es el rotulo de NATOT).
    _tmp_mg_esp = "__koboup_mg_dx_esp__"
    _new_mg_esp_label = "Especificar diagnóstico (Medicina General)"
    needs_especificar = False
    for row in rows:
        nr = dict(row)
        raw = str(nr.get(diag_col, "") or "").strip()
        raw_key = _normalize_header(raw)
        raw_for_parse = raw

        # Si viene vacío o en estado "placeholder" (ej. "Ninguno seleccionado"),
        # intentar recuperar el diagnóstico real desde columnas históricas.
        if not raw or raw_key in placeholder_diag_keys:
            for src_col in fallback_diag_sources:
                src_val = str(nr.get(src_col, "") or "").strip()
                if not src_val or _value_looks_missing(src_val):
                    continue
                src_key = _normalize_header(src_val)
                if not src_key or src_key in placeholder_diag_keys:
                    continue
                raw_for_parse = src_val
                break

        # Si la columna diagnóstica ya fue reducida a "Otro", intentar recuperar
        # el detalle desde columnas históricas de diagnóstico para poblar "Especificar".
        if raw_key in {"otro", "otra"}:
            for src_col in fallback_diag_sources:
                src_val = str(nr.get(src_col, "") or "").strip()
                if not src_val or _value_looks_missing(src_val):
                    continue
                src_norm = _normalize_header(src_val)
                if src_norm in {"otro", "otra"}:
                    continue
                raw_for_parse = src_val
                break
        if not raw_for_parse:
            if especificar_col:
                nr[especificar_col] = str(nr.get(especificar_col, "") or "").strip()
            out_rows.append(nr)
            continue
        tokens = _split_diag_tokens(raw_for_parse)
        normalized: list[str] = []
        especificar_values: list[str] = []
        seen: set[str] = set()
        for tk in tokens:
            key = _normalize_header(tk)
            if not key:
                continue
            mapped = ""
            if key in option_map:
                mapped = option_map[key]
            elif key in alias_map:
                mapped = alias_map[key]
            else:
                close = difflib.get_close_matches(key, non_otro_keys, n=1, cutoff=0.84)
                if close:
                    mapped = option_map[close[0]]
                else:
                    mapped = "Otro"
                    especificar_values.append(tk.strip())
            if mapped and mapped not in seen:
                seen.add(mapped)
                normalized.append(mapped)
        if normalized:
            nr[diag_col] = "|||".join(normalized)
        if "Otro" in normalized:
            if especificar_values:
                needs_especificar = True
                nr[_tmp_mg_esp] = "|||".join(
                    v for i, v in enumerate(especificar_values) if v and v not in especificar_values[:i]
                )
            elif especificar_col:
                # Mantiene lo que ya hubiera escrito en la columna de detalle (dxesp).
                keep = str(nr.get(especificar_col, "") or "").strip()
                if keep:
                    needs_especificar = True
                    nr[_tmp_mg_esp] = keep
            else:
                # Otro en lista pero sin detalle: columna de texto vacía
                needs_especificar = True
                nr[_tmp_mg_esp] = ""
        else:
            # Si ya no hay "Otro", limpiar el detalle de «Otro».
            nr[_tmp_mg_esp] = ""
        out_rows.append(nr)
    if needs_especificar and not especificar_col:
        out_cols.append(_new_mg_esp_label)
        especificar_col = _new_mg_esp_label
    if especificar_col:
        for nr in out_rows:
            if _tmp_mg_esp in nr:
                nr[especificar_col] = str(nr.get(_tmp_mg_esp, "") or "").strip()
                del nr[_tmp_mg_esp]
    return out_cols, out_rows


def _reorder_especificar_next_to_diagnostico_mg(columns: list[str]) -> list[str]:
    """
    Deja la columna «Especificar diagnóstico (Medicina General)» (dxesp) inmediatamente
    a la derecha de la columna de diagnóstico de medicina general.
    """
    if not columns or len(columns) < 2:
        return columns
    diag = _find_column_name_by_alias(
        columns,
        [
            "Diagnóstico Medicina General",
            "Diagnostico Medicina General",
            "Diagnósticos Medicina General",
            "Diagnosticos Medicina General",
            "Diagnósticos",
            "Diagnosticos",
        ],
    )
    esp = _find_column_name_by_alias(
        columns,
        [
            "Especificar diagnóstico (Medicina General)",
            "Especificar diagnóstico",
            "Especificar Diagnóstico Medicina General",
            "Especificar Diagnostico Medicina General",
            "dxesp",
        ],
    )
    if not diag or not esp or diag == esp:
        return columns
    try:
        d_i = columns.index(diag)
        e_i = columns.index(esp)
    except ValueError:
        return columns
    if e_i == d_i + 1:
        return columns
    out = list(columns)
    out.pop(e_i)
    d_i = out.index(diag)
    out.insert(d_i + 1, esp)
    return out


def _find_column_name_by_alias(columns: list[str], aliases: list[str]) -> str:
    alias_keys = {_normalize_header(a) for a in aliases if _normalize_header(a)}
    for c in columns:
        if _normalize_header(c) in alias_keys:
            return c
    return ""


def _ensure_columns_present(columns: list[str], rows: list[dict[str, str]], labels: list[str]) -> tuple[list[str], list[dict[str, str]]]:
    out_cols = list(columns)
    existing_keys = {_normalize_header(c) for c in out_cols}
    # Si la columna ya existe vía alias Kobo, no se vuelve a agregar.
    present_by_label = _label_to_present_columns(out_cols)
    to_add: list[str] = []
    for label in labels:
        if label in present_by_label and present_by_label[label]:
            continue
        key = _normalize_header(label)
        if key and key not in existing_keys:
            existing_keys.add(key)
            to_add.append(label)
            out_cols.append(label)
    if not to_add:
        return columns, rows
    out_rows: list[dict[str, str]] = []
    for row in rows:
        nr = dict(row)
        for label in to_add:
            nr.setdefault(label, "")
        out_rows.append(nr)
    return out_cols, out_rows


def _services_present_in_rows(columns: list[str], rows: list[dict[str, str]]) -> set[str]:
    service_col = _find_column_name_by_alias(
        columns,
        ["Servicio que se brinda", "Servicio", "Especialidad", "Servicio_que_se_brinda"],
    )
    if not service_col:
        return set()
    out: set[str] = set()
    for row in rows:
        raw = str(row.get(service_col) or "").strip()
        if not raw:
            continue
        out.add(_normalize_service(raw))
    return out


def _ensure_service_conditional_columns(columns: list[str], rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    columns, rows = _ensure_columns_present(columns, rows, ALWAYS_REQUIRED_SHEET_COLUMNS)
    services = _services_present_in_rows(columns, rows)
    needed: list[str] = []
    for service in services:
        needed.extend(SERVICE_CONDITIONAL_COLUMNS.get(service, []))
    if not needed:
        return columns, rows
    # Quitar duplicados preservando orden.
    seen: set[str] = set()
    uniq: list[str] = []
    for label in needed:
        key = _normalize_header(label)
        if not key or key in seen:
            continue
        seen.add(key)
        uniq.append(label)
    return _ensure_columns_present(columns, rows, uniq)


def _autofill_asesoria_previa(columns: list[str], rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    """
    Completa automáticamente la columna de asesoría previa por paciente:
    - Primera fila del paciente/día: vacía (no hay módulos previos en el mismo archivo;
      equivale a «sin dato» en Kobo, no a la opción explícita «No Aplica»)
    - Siguientes filas: módulos/especialidades previas acumuladas (separador |||)
    Reemplaza el valor existente para mantener consistencia operativa.
    """
    if not rows:
        return columns, rows

    asesoria_label = "¿Se le ha brindado asesoría en uno de los módulos el día de hoy?"
    columns, rows = _ensure_columns_present(columns, rows, [asesoria_label])

    service_col = _find_column_name_by_alias(
        columns,
        ["Servicio que se brinda", "Servicio", "Especialidad", "Servicio_que_se_brinda"],
    )
    patient_col = _find_column_name_by_alias(
        columns,
        ["Nombre del Paciente", "Nombre del paciente", "Nombre", "NAME"],
    )
    date_col = _find_column_name_by_alias(
        columns,
        ["Fecha de atención", "Fecha_de_atenci_n", "Fecha", "Fecha_atencion"],
    )
    asesoria_col = _find_column_name_by_alias(
        columns,
        [
            "¿Se le ha brindado asesoría en uno de los módulos el día de hoy?",
            "Se le ha brindado asesoría en uno de los módulos el día de hoy",
            "ASESPREV",
            "Asesoría previa hoy",
            "Asesoria previa hoy",
        ],
    )
    if not service_col or not asesoria_col:
        return columns, rows

    # key -> lista ordenada de servicios previos ya vistos
    history: dict[str, list[str]] = {}
    out_rows: list[dict[str, str]] = []

    for idx, row in enumerate(rows):
        nr = dict(row)
        raw_service = str(nr.get(service_col, "") or "").strip()
        current_service = _normalize_service(raw_service) if raw_service else ""

        patient_key = (
            _normalize_header(str(nr.get(patient_col, "") or "").strip())
            if patient_col
            else ""
        )
        date_key = (
            _normalize_header(str(nr.get(date_col, "") or "").strip())
            if date_col
            else ""
        )
        # Si falta paciente, evita mezclar historiales entre filas anónimas.
        group_key = f"{patient_key}|{date_key}" if patient_key else f"__row__{idx}"

        prev_services = history.get(group_key, [])
        if prev_services:
            nr[asesoria_col] = "|||".join(prev_services)
        else:
            nr[asesoria_col] = ""

        if current_service and current_service not in prev_services:
            prev_services = prev_services + [current_service]
        history[group_key] = prev_services
        out_rows.append(nr)

    return columns, out_rows


_ME_ML_COLUMN_ALIASES: list[str] = [
    "¿Mujer embarazada o en periodo de lactancia?",
    "Mujer embarazada o en periodo de lactancia",
    "¿Embarazada / Lactancia?",
    "Embarazada / Lactancia",
    "ME_ML",
    "Embarazada o lactancia",
    "Embarazo/Lactancia",
    "Embarazo / Lactancia",
    "Embarazada/Lactancia?",
]


def _me_ml_cell_value_valid(raw: str) -> bool:
    """Valores de catálogo Kobo: Embarazada, Lactancia, No Aplica (y alias numéricos heredados)."""
    t = _normalize_header(str(raw or "").strip())
    if not t:
        return False
    if t in {"embarazada", "embarazo", "1"}:
        return True
    if t in {"lactancia", "2 1", "2_1"}:
        return True
    if t in {"no aplica", "noaplica", "0", "na", "n a", "n/d", "nd"}:
        return True
    return False


def _sex_value_is_female(raw: str) -> bool:
    t = _normalize_header(str(raw or "").strip())
    return t in {"femenino", "f", "female", "mujer", "2"}


def _autofill_me_ml_by_sexo(columns: list[str], rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    """
    Embarazo/lactancia (ME_ML):
    - Sexo masculino (u otro no femenino): celda vacía si no hay valor válido o sobra texto.
    - Sexo femenino sin dato válido (Embarazada / Lactancia / No Aplica): «No Aplica».
    """
    if not rows:
        return columns, rows
    sex_col = _find_column_name_by_alias(columns, ["Sexo", "SEX"])
    me_col = _find_column_name_by_alias(columns, _ME_ML_COLUMN_ALIASES)
    if not me_col:
        return columns, rows
    out_rows: list[dict[str, str]] = []
    for row in rows:
        nr = dict(row)
        sex_raw = str(nr.get(sex_col, "") or "").strip() if sex_col else ""
        cur = str(nr.get(me_col, "") or "").strip()
        if _sex_value_is_female(sex_raw):
            if not _me_ml_cell_value_valid(cur):
                nr[me_col] = "No Aplica"
        else:
            if cur:
                nr[me_col] = ""
        out_rows.append(nr)
    return columns, out_rows


def _load_records_from_file(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            rows = ws.iter_rows(values_only=True)
            headers_row = next(rows, None)
            if not headers_row:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in headers_row]
            if headers and ";" in headers[0] and not any(h for h in headers[1:]):
                split_headers = [part.strip() for part in headers[0].split(";")]
                parsed_rows: list[dict[str, str]] = []
                for values in rows:
                    raw_parts = [
                        str(cell).strip()
                        for cell in values
                        if cell is not None and str(cell).strip()
                    ]
                    if not raw_parts:
                        continue
                    merged = ",".join(raw_parts)
                    split_values = [part.strip() for part in merged.split(";")]
                    if any(split_values):
                        parsed_rows.append(_build_row_dict(split_headers, split_values))
                return parsed_rows
            return [
                _build_row_dict(headers, list(values))
                for values in rows
                if values and any(cell not in (None, "") for cell in values)
            ]
        finally:
            wb.close()

    if suffix == ".csv":
        for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                with open(path, "r", encoding=encoding, errors="replace", newline="") as fh:
                    sample = fh.read(4096)
                    fh.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
                    except Exception:
                        dialect = csv.excel
                    reader = csv.reader(fh, dialect)
                    headers = next(reader, None)
                    if not headers:
                        return []
                    if len(headers) == 1 and ";" in headers[0]:
                        fh.seek(0)
                        reader = csv.reader(fh, delimiter=";")
                        headers = next(reader, None)
                        if not headers:
                            return []
                    headers = [str(cell).strip() if cell is not None else "" for cell in headers]
                    return [
                        _build_row_dict(headers, row)
                        for row in reader
                        if row and any(str(cell).strip() for cell in row)
                    ]
            except Exception:
                continue
    return []


def _sheet_column_order(records: list[dict[str, str]]) -> list[str]:
    """Orden estable de columnas: orden de aparición al recorrer filas."""
    if not records:
        return []
    columns: list[str] = []
    seen: set[str] = set()
    for rec in records:
        for k in rec.keys():
            if k not in seen:
                seen.add(k)
                columns.append(k)
    return columns


def _sheet_format_key(path: Path) -> tuple[str | None, str | None]:
    """
    (formato, error). formato es 'xlsx' o 'csv' si se admite edición y guardado.
    """
    suf = path.suffix.lower()
    if suf == ".pdf":
        return None, "Los archivos PDF no se pueden editar como tabla. Use la descarga o un PDF editor."
    if suf == ".xls":
        return None, (
            "El formato .xls (Excel antiguo) no se puede reescribir con seguridad. "
            "Convierta a .xlsx, vuelva a subir el archivo o abra y guarde en Excel en formato moderno."
        )
    if suf == ".xlsx":
        return "xlsx", None
    if suf == ".csv":
        return "csv", None
    return None, "No se admite el editor de tabla para este tipo de archivo."


def _write_sheet_xlsx(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if columns:
        ws.append(list(columns))
    for row in rows:
        ws.append([str(row.get(c) or "") for c in columns])
    wb.save(str(path))


def _write_sheet_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        if columns:
            w.writerow(columns)
        for row in rows:
            w.writerow([str(row.get(c) or "") for c in columns])


def _write_sheet_to_path(
    path: Path, format_key: str, columns: list[str], rows: list[dict[str, str]]
) -> None:
    if format_key == "xlsx":
        _write_sheet_xlsx(path, columns, rows)
    elif format_key == "csv":
        _write_sheet_csv(path, columns, rows)
    else:
        raise ValueError("formato de hoja desconocido")


def _normalize_put_columns(raw: object) -> tuple[list[str] | None, str | None]:
    if not isinstance(raw, list):
        return None, "Falta 'columns' (array) con nombres de columna."
    if len(raw) == 0:
        return [], None
    if len(raw) > MAX_SHEET_COLS:
        return None, f"Máximo {MAX_SHEET_COLS} columnas."
    out: list[str] = []
    used: set[str] = set()
    for i, c in enumerate(raw):
        name = str(c if c is not None else "").strip() or f"Columna {i + 1}"
        candidate = name
        n = 1
        while candidate in used:
            n += 1
            candidate = f"{name} ({n})"
        used.add(candidate)
        out.append(candidate)
    return out, None


def _normalize_put_rows(
    raw: object, columns: list[str]
) -> tuple[list[dict[str, str]] | None, str | None]:
    if not isinstance(raw, list):
        return None, "'rows' debe ser un array (lista de filas)."
    if len(raw) > MAX_SHEET_ROWS:
        return None, f"Máximo {MAX_SHEET_ROWS} filas de datos (sin contar el encabezado)."
    out: list[dict[str, str]] = []
    for r in raw:
        if r is not None and not isinstance(r, dict):
            return None, "Cada fila debe ser un objeto (mapa de columna → valor)."
        rd: dict[str, str] = {}
        rec = r or {}
        for c in columns:
            v = rec.get(c, "")
            rd[c] = "" if v is None else str(v)
        out.append(rd)
    return out, None


def _editor_name_from_request(data: dict | None = None) -> str:
    payload = data or {}
    for key in ("editor_name", "editor", "uploaded_by", "validated_by"):
        val = payload.get(key)
        if val and str(val).strip():
            return str(val).strip()
    for key in ("editor_name", "editor", "uploaded_by"):
        val = request.args.get(key)
        if val and str(val).strip():
            return str(val).strip()
    hdr = request.headers.get("X-Editor-Name")
    if hdr and str(hdr).strip():
        return str(hdr).strip()
    return ""


def _sheet_lock_error_response(lock_info: dict | None):
    lock = lock_info or {}
    locked_by = str(lock.get("locked_by") or "").strip() or "otro usuario"
    lock_at = lock.get("locked_at")
    msg = (
        f"Este archivo está siendo editado por {locked_by}. "
        "Intente más tarde o pídale cerrar el editor."
    )
    return (
        jsonify(
            {
                "ok": False,
                "error": msg,
                "locked_by": locked_by,
                "locked_at": lock_at,
                "lock_ttl_seconds": EDIT_LOCK_TTL_SECONDS,
            }
        ),
        423,
    )


def _pick_value(row: dict[str, str], aliases: list[str]) -> str:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_normalize_header(alias), "")
        if str(value).strip():
            return str(value).strip()
    for alias in aliases:
        alias_norm = _normalize_header(alias)
        for key_norm, value in normalized.items():
            if not str(value).strip():
                continue
            if alias_norm and (alias_norm in key_norm or key_norm in alias_norm):
                return str(value).strip()
    return ""


def _normalize_service(raw: str) -> str:
    service_norm = _norm_text(raw)
    alias_map = [
        ("Dental", ["dental", "odontologia", "odontologia"]),
        ("Fisioterapia", ["fisioterapia", "fisio", "rehabilitacion"]),
        ("Oftalmología", ["oftalmologia", "optica", "vision", "lentes"]),
        ("Laboratorios", ["laboratorio", "laboratorios", "lab", "examenes"]),
        ("Medicina General", ["medicina general", "medicina", "consulta", "medico"]),
    ]
    for label, keywords in alias_map:
        if any(keyword in service_norm for keyword in keywords):
            return label
    return "Medicina General"


def _contains_any(text: str, keywords: list[str]) -> bool:
    return bool(text) and any(keyword in text for keyword in keywords)


def _compute_supply_flags(supply_norm: str, service_label: str) -> dict[str, bool]:
    med_keywords = [
        "medicamento", "medicamentos", "medicina", "pastilla", "pastillas", "tableta",
        "tabletas", "capsula", "capsulas", "jarabe", "crema", "pomada", "gel", "spray",
        "tabs", "tab", "susp", "solucion", "soluciones", "suplemento", "sobres",
        "amoxicilina", "ibuprofeno", "paracetamol", "omeprazol", "metformina",
        "insulina", "salbutamol", "albendazol", "vitamina", "ketorolaco", "loratadina",
        "ambroxol", "losartan", "nifedipino", "metoprolol", "glibenclamida",
        "hipromelosa", "ketoconazol", "vermisen", "cilocid", "convifer", "lumboxen",
        "collifrin",
    ]
    lens_keywords = ["lentes", "anteojos", "armazon", "graduacion", "montura", "bifocal"]
    kit_keywords = [
        "kit dental", "kit de limpieza", "kit limpieza dental", "kit de limpieza dental",
        "kid dental", "kid de limpieza", "kid de limpieza dental", "cepillo dental",
        "pasta dental", "hilo dental",
    ]
    extraction_keywords = ["extraccion", "extracciones"]
    cleaning_keywords = [
        "limpieza", "profilaxis", "detartr", "kit de limpieza",
        "kit limpieza dental", "kit de limpieza dental",
        "kid de limpieza", "kid de limpieza dental",
    ]
    dental_rx_keywords = ["rx", "rayo x", "rayos x", "radiografia"]

    is_dental = service_label == "Dental"
    return {
        "kit_dental": _contains_any(supply_norm, kit_keywords),
        "medicamento": _contains_any(supply_norm, med_keywords),
        "lentes": _contains_any(supply_norm, lens_keywords),
        "extracciones": is_dental and _contains_any(supply_norm, extraction_keywords),
        "limpieza_dental": is_dental and _contains_any(supply_norm, cleaning_keywords),
        "rx_odontologia": is_dental and _contains_any(supply_norm, dental_rx_keywords),
    }


def _extract_extraction_units(supply_norm: str) -> int:
    if "extrac" not in supply_norm:
        return 0

    total = 0
    patterns = [
        r"(\d+)\s*extrac",
        r"extrac\w*\s*[:\-x]?\s*(\d+)",
        r"extrac\w*\s*\((\d+)\)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, supply_norm):
            try:
                total += int(match.group(1))
            except Exception:
                pass
    return total or 1


def _extract_cleaning_units(supply_norm: str) -> int:
    cleaning_present = any(
        keyword in supply_norm
        for keyword in [
            "limpieza", "profilaxis", "detartr", "kit de limpieza",
            "kit limpieza dental", "kit de limpieza dental",
            "kid de limpieza", "kid de limpieza dental",
        ]
    )
    if not cleaning_present:
        return 0

    total = 0
    patterns = [
        r"(\d+)\s*(?:limpieza|profilaxis|detartr)",
        r"(?:limpieza|profilaxis|detartr)\w*\s*[:\-x]?\s*(\d+)",
        r"(?:limpieza|profilaxis|detartr)\w*\s*\((\d+)\)",
        r"(\d+)\s*(?:kit de limpieza|kit limpieza dental|kit de limpieza dental|kid de limpieza|kid de limpieza dental)",
        r"(?:kit de limpieza|kit limpieza dental|kit de limpieza dental|kid de limpieza|kid de limpieza dental)\s*[:\-x]?\s*(\d+)",
        r"(?:kit de limpieza|kit limpieza dental|kit de limpieza dental|kid de limpieza|kid de limpieza dental)\s*\((\d+)\)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, supply_norm):
            try:
                total += int(match.group(1))
            except Exception:
                pass
    return total or 1


def _parse_supply_quantity(raw_value: str) -> int:
    text = _norm_text(raw_value)
    if text in {"", "n/d", "nd", "no", "-", "nan"}:
        return 0
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return 0
    try:
        value = float(match.group(0))
    except Exception:
        return 0
    if value <= 0:
        return 0
    return int(value)


def _get_supply_context(row: dict[str, str]) -> tuple[str, str, int]:
    category_text = _pick_value(
        row,
        [
            "Insumos Entregados (Categoría general)",
            "Insumos Entregados (Categoria general)",
            "Resultados_Lab_Insumos",
            "Resultados Lab / Insumos",
            "Insumos Entregados",
            "Insumos",
        ],
    )
    treatment_text = _pick_value(row, ["Tratamiento", "Medicamentos", "Tx"])
    if not treatment_text:
        treatment_text = _pick_value(row, ["Medicamentos / Procedimiento", "Tratamiento indicado"])
    detail_text = _pick_value(
        row,
        [
            "Especifique_qu_se_entrega",
            "Especifique qué se entrega",
            "Especifique que se entrega",
            "Especifique qué se entrega (detalle del insumo)",
            "Especifique que se entrega (detalle del insumo)",
            "Especifique",
        ],
    )
    quantity_text = _pick_value(
        row,
        [
            "Unidades_entregadas (Cantidad en número)",
            "Unidades_entregadas (Cantidad en numero)",
            "Unidades_entregadas",
            "Unid.",
            "Unid",
            "Cantidad entregada",
            "Cantidad en número",
            "Cantidad en numero",
            "Unidades",
            "Cantidad",
            "Entregados",
        ],
    )
    supply_text = " | ".join(part for part in [category_text, treatment_text, detail_text] if part)
    quantity_value = _parse_supply_quantity(quantity_text)
    return category_text, supply_text, quantity_value


def _get_supply_increment(
    key: str,
    flags: dict[str, bool],
    category_flags: dict[str, bool],
    quantity_value: int,
) -> int:
    if not flags[key]:
        return 0
    active_supply_flags = sum(
        1 for supply_key in ("kit_dental", "medicamento", "lentes") if flags.get(supply_key)
    )
    if quantity_value > 0 and (category_flags.get(key) or active_supply_flags == 1):
        return quantity_value
    return 1


def _compute_submission_hash(record: dict) -> str:
    key_fields = [
        "NAME",
        "Fecha_de_atenci_n",
        "Servicio_que_se_brinda",
        "SEX",
        "AGE",
        "Diagnostico_Motivo",
    ]
    parts = [str(record.get(field, "")).strip().lower() for field in key_fields]
    return hashlib.md5("|".join(parts).encode("utf-8")).hexdigest()


def _load_root_excel_to_records():
    module_path = ROOT_PROJECT_DIR / "excel_loader.py"
    spec = importlib.util.spec_from_file_location("root_excel_loader", module_path)
    if not spec or not spec.loader:
        raise RuntimeError(f"No se pudo cargar {module_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.load_excel_to_records


def _compute_public_kpis_from_root_history() -> dict | None:
    """Reconstruye KPI desde el historial real de filas enviadas del proyecto principal."""
    if not ROOT_SUBMITTED_FILE.exists():
        return None

    try:
        store = json.loads(ROOT_SUBMITTED_FILE.read_text(encoding="utf-8"))
    except Exception as exc:
        logging.warning("No se pudo leer %s: %s", ROOT_SUBMITTED_FILE, exc)
        return None

    submissions = store.get("submissions", [])
    if not submissions:
        return None

    load_excel_to_records = _load_root_excel_to_records()
    specialties_order = [
        "Medicina General",
        "Dental",
        "Fisioterapia",
        "Oftalmología",
        "Laboratorios",
    ]
    specialties = {name: 0 for name in specialties_order}
    supplies = {"kit_dental": 0, "medicamento": 0, "lentes": 0}
    dental_procedures = {"extracciones": 0, "limpieza_dental": 0, "rx_odontologia": 0}
    submitted_hashes = {str(item.get("hash", "")).strip() for item in submissions if item.get("hash")}
    file_names: list[str] = []
    seen_file_names: set[str] = set()
    for item in submissions:
        file_name = Path(str(item.get("file", "")).strip()).name
        if not file_name or file_name in seen_file_names:
            continue
        seen_file_names.add(file_name)
        file_names.append(file_name)

    seen_hashes: set[str] = set()
    seen_patients: set[tuple[str, str, str, str]] = set()

    for file_name in file_names:
        path = ROOT_UPLOADS_DIR / file_name
        if not path.exists() or path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
            continue
        try:
            records = load_excel_to_records(path)
        except Exception as exc:
            logging.warning("No se pudo leer %s para KPI raíz: %s", path.name, exc)
            continue

        for record in records:
            row_hash = _compute_submission_hash(record)
            if row_hash not in submitted_hashes or row_hash in seen_hashes:
                continue
            seen_hashes.add(row_hash)

            patient_key = (
                _norm_text(record.get("NAME", "")),
                str(record.get("Fecha_de_atenci_n", "")).strip(),
                _norm_text(record.get("SEX", "")),
                str(record.get("AGE", "")).strip(),
            )
            if any(patient_key):
                seen_patients.add(patient_key)

            service_label = _normalize_service(record.get("Servicio_que_se_brinda", ""))
            specialties[service_label] += 1

            category_text, supply_text, quantity_value = _get_supply_context(record)
            supply_norm = _norm_text(supply_text)
            category_norm = _norm_text(category_text)
            flags = _compute_supply_flags(supply_norm, service_label)
            category_flags = _compute_supply_flags(category_norm, service_label)
            for key in supplies:
                supplies[key] += _get_supply_increment(key, flags, category_flags, quantity_value)
            for key in dental_procedures:
                if flags[key]:
                    dental_procedures[key] += (
                        _extract_extraction_units(supply_norm)
                        if key == "extracciones"
                        else _extract_cleaning_units(supply_norm)
                        if key == "limpieza_dental"
                        else 1
                    )

    return {
        "patients_registered": len(seen_patients),
        "total_consultations": sum(specialties.values()),
        "specialties": [
            {"key": _normalize_header(name).replace(" ", "_"), "label": name, "count": specialties[name]}
            for name in specialties_order
        ],
        "supplies": [
            {"key": "kit_dental", "label": "Kit dental", "count": supplies["kit_dental"]},
            {"key": "medicamento", "label": "Medicamento", "count": supplies["medicamento"]},
            {"key": "lentes", "label": "Lentes", "count": supplies["lentes"]},
        ],
        "dental_procedures": [
            {"key": "extracciones", "label": "Extracciones", "count": dental_procedures["extracciones"]},
            {"key": "limpieza_dental", "label": "Limpieza dental", "count": dental_procedures["limpieza_dental"]},
            {"key": "rx_odontologia", "label": "Rx de odontología", "count": dental_procedures["rx_odontologia"]},
        ],
        "total_supplies_delivered": sum(supplies.values()),
        "grand_total": sum(specialties.values()) + sum(supplies.values()),
        "files_downloaded": len(file_names),
        "validated_files": len(file_names),
        "source": "root_history",
    }


def _compute_public_kpis_from_priority_folder() -> dict | None:
    """Lee los Excel/CSV de la carpeta local prioritaria de archivos validados."""
    if not PRIORITY_VALIDATED_DIR.exists():
        return None
    excluded_files = _load_kpi_exclusions()

    files = [
        path for path in PRIORITY_VALIDATED_DIR.iterdir()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls", ".csv"}
    ]

    seen_paths: set[str] = set()
    combined_files: list[Path] = []
    for path in files + EXTRA_KPI_FILES:
        if not path.exists():
            continue
        resolved = str(path.resolve())
        if resolved in seen_paths:
            continue
        seen_paths.add(resolved)
        combined_files.append(path)

    files = sorted(combined_files, key=lambda path: path.name.lower())
    if not files:
        return None

    specialties_order = [
        "Medicina General",
        "Dental",
        "Fisioterapia",
        "Oftalmología",
        "Laboratorios",
    ]
    specialties = {name: 0 for name in specialties_order}
    supplies = {"kit_dental": 0, "medicamento": 0, "lentes": 0}
    dental_procedures = {"extracciones": 0, "limpieza_dental": 0, "rx_odontologia": 0}
    seen_consultations: set[tuple[str, str, str, str, str]] = set()
    seen_patients: set[tuple[str, str, str, str]] = set()
    supplies_by_file: dict[str, dict[str, int]] = {}

    active_files_count = 0

    for path in files:
        try:
            rows = _load_records_from_file(path)
        except Exception as exc:
            logging.warning("No se pudo leer %s para KPI de carpeta prioritaria: %s", path.name, exc)
            continue

        is_excluded = path.name in excluded_files
        file_supply_counts = {"kit_dental": 0, "medicamento": 0, "lentes": 0}
        for row in rows:
            name = _pick_value(row, ["NAME", "Nombre del Paciente", "Nombre"])
            date = _pick_value(row, ["Fecha_de_atenci_n", "Fecha de Atención", "Fecha de atención", "Fecha atención", "Fecha"])
            sex = _pick_value(row, ["SEX", "Sexo"])
            age = _pick_value(row, ["AGE", "Edad"])
            service = _pick_value(row, ["Servicio_que_se_brinda", "Servicio que se brinda", "Servicio", "Especialidad"])
            service_label = _normalize_service(service)

            category_text, supply_text, quantity_value = _get_supply_context(row)
            supply_norm = _norm_text(supply_text)
            category_norm = _norm_text(category_text)
            flags = _compute_supply_flags(supply_norm, service_label)
            category_flags = _compute_supply_flags(category_norm, service_label)
            for key in supplies:
                increment = _get_supply_increment(key, flags, category_flags, quantity_value)
                file_supply_counts[key] += increment

            if is_excluded:
                continue

            patient_key = (_norm_text(name), date.strip(), _norm_text(sex), age.strip())
            if any(patient_key):
                seen_patients.add(patient_key)

            consultation_key = patient_key + (service_label,)
            if consultation_key in seen_consultations:
                continue
            seen_consultations.add(consultation_key)
            specialties[service_label] += 1

            for key in supplies:
                increment = _get_supply_increment(key, flags, category_flags, quantity_value)
                supplies[key] += increment
            for key in dental_procedures:
                if flags[key]:
                    dental_procedures[key] += (
                        _extract_extraction_units(supply_norm)
                        if key == "extracciones"
                        else _extract_cleaning_units(supply_norm)
                        if key == "limpieza_dental"
                        else 1
                    )

        file_total_supplies = sum(file_supply_counts.values())
        supplies_by_file[path.name] = {
            "excluded": is_excluded,
            **file_supply_counts,
            "total_supplies": file_total_supplies,
        }

        if is_excluded:
            continue
        active_files_count += 1

    total_consultations = sum(specialties.values())
    total_supplies_delivered = sum(supplies.values())
    return {
        "patients_registered": len(seen_patients),
        "total_consultations": total_consultations,
        "specialties": [
            {"key": _normalize_header(name).replace(" ", "_"), "label": name, "count": specialties[name]}
            for name in specialties_order
        ],
        "supplies": [
            {"key": "kit_dental", "label": "Kit dental", "count": supplies["kit_dental"]},
            {"key": "medicamento", "label": "Medicamento", "count": supplies["medicamento"]},
            {"key": "lentes", "label": "Lentes", "count": supplies["lentes"]},
        ],
        "dental_procedures": [
            {"key": "extracciones", "label": "Extracciones", "count": dental_procedures["extracciones"]},
            {"key": "limpieza_dental", "label": "Limpieza dental", "count": dental_procedures["limpieza_dental"]},
            {"key": "rx_odontologia", "label": "Rx de odontología", "count": dental_procedures["rx_odontologia"]},
        ],
        "supplies_by_file": [
            {
                "file_name": file_name,
                "excluded": counts["excluded"],
                "kit_dental": counts["kit_dental"],
                "medicamento": counts["medicamento"],
                "lentes": counts["lentes"],
                "total_supplies": counts["total_supplies"],
            }
            for file_name, counts in sorted(
                supplies_by_file.items(),
                key=lambda item: (-item[1]["total_supplies"], item[0].lower()),
            )
        ],
        "total_supplies_delivered": total_supplies_delivered,
        "grand_total": total_consultations + total_supplies_delivered,
        "files_downloaded": active_files_count,
        "validated_files": active_files_count,
        "total_files_analyzed": len(files),
        "excluded_files": sorted(excluded_files),
        "source": "priority_validated_folder_plus_extra_uploads",
    }


def _compute_public_kpis() -> dict:
    folder_kpis = _compute_public_kpis_from_priority_folder()
    if folder_kpis is not None:
        return folder_kpis

    root_kpis = _compute_public_kpis_from_root_history()
    if root_kpis is not None:
        return root_kpis

    files = list_file_records("validado")
    specialties_order = [
        "Medicina General",
        "Dental",
        "Fisioterapia",
        "Oftalmología",
        "Laboratorios",
    ]
    specialties = {name: 0 for name in specialties_order}
    supplies = {"kit_dental": 0, "medicamento": 0, "lentes": 0}
    dental_procedures = {"extracciones": 0, "limpieza_dental": 0, "rx_odontologia": 0}
    seen_consultations: set[tuple[str, str, str, str, str]] = set()
    seen_patients: set[tuple[str, str, str, str]] = set()

    for file_entry in files:
        path = get_file_path(file_entry)
        if not path.exists() or file_entry.get("file_type") == "pdf":
            continue
        try:
            rows = _load_records_from_file(path)
        except Exception as exc:
            logging.warning("No se pudo leer %s para KPI: %s", path.name, exc)
            continue

        for row in rows:
            name = _pick_value(row, ["NAME", "Nombre del Paciente", "Nombre"])
            date = _pick_value(row, ["Fecha_de_atenci_n", "Fecha de Atención", "Fecha de atención", "Fecha atención", "Fecha"])
            sex = _pick_value(row, ["SEX", "Sexo"])
            age = _pick_value(row, ["AGE", "Edad"])
            service = _pick_value(row, ["Servicio_que_se_brinda", "Servicio que se brinda", "Servicio", "Especialidad"])
            service_label = _normalize_service(service)

            patient_key = (_norm_text(name), date.strip(), _norm_text(sex), age.strip())
            if any(patient_key):
                seen_patients.add(patient_key)

            consultation_key = patient_key + (service_label,)
            if consultation_key in seen_consultations:
                continue
            seen_consultations.add(consultation_key)
            specialties[service_label] += 1

            category_text, supply_text, quantity_value = _get_supply_context(row)
            supply_norm = _norm_text(supply_text)
            category_norm = _norm_text(category_text)
            flags = _compute_supply_flags(supply_norm, service_label)
            category_flags = _compute_supply_flags(category_norm, service_label)
            for key in supplies:
                supplies[key] += _get_supply_increment(key, flags, category_flags, quantity_value)
            for key in dental_procedures:
                if flags[key]:
                    dental_procedures[key] += (
                        _extract_extraction_units(supply_norm)
                        if key == "extracciones"
                        else _extract_cleaning_units(supply_norm)
                        if key == "limpieza_dental"
                        else 1
                    )

    total_consultations = sum(specialties.values())
    total_supplies_delivered = sum(supplies.values())
    return {
        "patients_registered": len(seen_patients),
        "total_consultations": total_consultations,
        "specialties": [
            {"key": _normalize_header(name).replace(" ", "_"), "label": name, "count": specialties[name]}
            for name in specialties_order
        ],
        "supplies": [
            {"key": "kit_dental", "label": "Kit dental", "count": supplies["kit_dental"]},
            {"key": "medicamento", "label": "Medicamento", "count": supplies["medicamento"]},
            {"key": "lentes", "label": "Lentes", "count": supplies["lentes"]},
        ],
        "dental_procedures": [
            {"key": "extracciones", "label": "Extracciones", "count": dental_procedures["extracciones"]},
            {"key": "limpieza_dental", "label": "Limpieza dental", "count": dental_procedures["limpieza_dental"]},
            {"key": "rx_odontologia", "label": "Rx de odontología", "count": dental_procedures["rx_odontologia"]},
        ],
        "total_supplies_delivered": total_supplies_delivered,
        "grand_total": total_consultations + total_supplies_delivered,
        "files_downloaded": len(files),
        "validated_files": len(files),
    }


@app.route("/")
def index():
    return send_from_directory(str(STATIC_DIR), "index.html")


@app.route("/api/treatment-suggestions", methods=["POST"])
def api_treatment_suggestions():
    if not suggest_treatment:
        return jsonify({"ok": False, "error": "Módulo de sugerencias no disponible"}), 501
    return jsonify(suggest_treatment(request.get_json(silent=True) or {}))


@app.route("/api/treatment-cohort/stats", methods=["GET"])
def api_treatment_cohort_stats():
    if not cohort_stats:
        return jsonify({"ok": False, "error": "Módulo no disponible"}), 501
    return jsonify(cohort_stats())


@app.route("/api/treatment-cohort/rebuild", methods=["POST"])
def api_treatment_cohort_rebuild():
    if not build_and_write:
        return jsonify({"ok": False, "error": "Módulo no disponible"}), 501
    try:
        data = build_and_write()
        return jsonify({"ok": True, "fila_tto_con_dosis": data.get("fila_tto_con_dosis"), "file_count": data.get("file_count")})
    except Exception as ex:  # noqa: BLE001
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/api/files", methods=["GET"])
def list_files():
    status = (request.args.get("status") or "").strip().lower()
    status_filter = status if status in VALID_STATUSES else None
    raw = list_file_records(status_filter)
    fids = [f["id"] for f in raw if f.get("id") is not None]
    kobo_stats = kobo_submission_stats_by_file_ids(fids)
    files: list[dict] = []
    for f in raw:
        aug = _augment(f)
        fid = aug.get("id")
        st = kobo_stats.get(int(fid)) if fid is not None else None
        aug["kobo_api_sent"] = bool(st and st.get("kobo_api_sent"))
        last_at = (st or {}).get("kobo_api_last_submitted_at")
        aug["kobo_api_last_submitted_at"] = str(last_at) if last_at is not None else None
        files.append(aug)
    return jsonify({"ok": True, "files": files})


@app.route("/api/files", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "No se envió ningún archivo"}), 400
    uploaded = request.files["file"]
    if uploaded.filename == "":
        return jsonify({"error": "Nombre de archivo vacío"}), 400
    if not allowed_file(uploaded.filename):
        return jsonify({"error": "Solo se permiten .xlsx, .xls, .csv o .pdf"}), 400

    ext = uploaded.filename.rsplit(".", 1)[1].lower()
    original_name = uploaded.filename
    safe_name = secure_filename(original_name) or f"archivo.{ext}"
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    stored_name = f"{timestamp}_{safe_name}"

    uploaded_by = (request.form.get("uploaded_by") or "").strip() or None
    notes = (request.form.get("notes") or "").strip() or None

    status_param = (request.form.get("status") or "").strip().lower()
    if status_param in VALID_STATUSES:
        target_status = status_param
    else:
        target_status = "pendiente"

    dest_dir = VALIDATED_DIR if target_status == "validado" else PENDING_DIR
    dest_dir.mkdir(exist_ok=True, parents=True)
    dest_path = dest_dir / stored_name

    uploaded.save(dest_path)
    size_bytes = dest_path.stat().st_size if dest_path.exists() else None

    record = add_file_record(
        original_name=original_name,
        stored_name=stored_name,
        file_type="pdf" if ext == "pdf" else "excel",
        status=target_status,
        size_bytes=size_bytes,
        notes=notes,
        uploaded_by=uploaded_by,
    )

    if ext != "pdf":
        row_count = count_file_rows(dest_path)
        if row_count is not None:
            update_row_count(record["id"], row_count)
            record["row_count"] = row_count

    superseded = []
    if target_status == "validado":
        replaces_id_str = (request.form.get("replaces_id") or "").strip()
        if replaces_id_str:
            try:
                replaces_id = int(replaces_id_str)
                supersede_specific_file(replaces_id, record["id"])
                superseded = [replaces_id]
            except (ValueError, TypeError):
                pass
        if not superseded:
            superseded = supersede_matching_files(record["id"], original_name)

    return jsonify({"ok": True, "file": _augment(record), "superseded": superseded})


@app.route("/api/files/<int:file_id>/download", methods=["GET"])
def download_file(file_id: int):
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    path = get_file_path(entry)
    if not path.exists():
        return jsonify({"error": "El archivo ya no está disponible en el servidor"}), 404
    return send_file(
        str(path),
        as_attachment=True,
        download_name=entry.get("original_name") or entry.get("stored_name"),
    )


@app.route("/api/files/<int:file_id>/register-download", methods=["POST"])
def register_download(file_id: int):
    """Registra quién descargó y cambia estado a por_validar."""
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    data = request.get_json(silent=True) or {}
    downloaded_by = (data.get("downloaded_by") or "").strip() or None

    from file_store import _connect
    from datetime import datetime as _dt
    now = _dt.utcnow().isoformat()
    with _connect() as conn:
        conn.execute(
            "UPDATE files SET downloaded_by = ?, downloaded_at = ?, updated_at = ? WHERE id = ?",
            (downloaded_by, now, now, file_id),
        )

    if entry.get("status") == "pendiente":
        update_status(file_id, "por_validar")

    updated = get_file_record(file_id)
    return jsonify({"ok": True, "file": _augment(updated)})


@app.route("/api/files/<int:file_id>/status", methods=["POST"])
def change_status(file_id: int):
    """Cambia el estado de un archivo: pendiente → por_validar → validado."""
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404

    if entry.get("status") == "reemplazado":
        return jsonify({"error": "Este archivo ya fue reemplazado y no puede cambiar de estado"}), 409

    data = request.get_json(silent=True) or {}
    new_status = (data.get("status") or "").strip().lower()
    if new_status not in VALID_STATUSES:
        return jsonify({"error": f"Estado inválido. Use: {', '.join(VALID_STATUSES)}"}), 400

    validated_by = (data.get("validated_by") or "").strip() or None
    notes = (data.get("notes") or "").strip() or None

    if new_status == "validado":
        if has_validated_replacement(file_id):
            update_status(file_id, "reemplazado")
            updated = get_file_record(file_id)
            return jsonify({
                "ok": True,
                "file": _augment(updated),
                "was_superseded": True,
                "message": "Este archivo fue marcado como reemplazado porque ya existe una versión validada.",
            })
        ensure_validated_location(entry)

    updated = update_status(file_id, new_status, validated_by=validated_by, notes=notes)

    superseded = []
    if new_status == "validado":
        superseded = supersede_matching_files(file_id, entry["original_name"])
        if schedule_rebuild_if_quiet:
            try:
                schedule_rebuild_if_quiet()
            except Exception:
                pass

    return jsonify({"ok": True, "file": _augment(updated), "superseded": superseded})


@app.route("/api/files/<int:file_id>/validate", methods=["POST"])
def validate_file(file_id: int):
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404

    if entry.get("status") == "reemplazado":
        return jsonify({"error": "Este archivo ya fue reemplazado por una versión validada"}), 409

    if has_validated_replacement(file_id):
        update_status(file_id, "reemplazado")
        updated = get_file_record(file_id)
        return jsonify({
            "ok": True,
            "file": _augment(updated),
            "was_superseded": True,
            "message": "Este archivo fue marcado como reemplazado porque ya existe una versión validada.",
        })

    data = request.get_json(silent=True) or {}
    validated_by = (data.get("validated_by") or "").strip() or None
    notes = (data.get("notes") or "").strip() or None
    ensure_validated_location(entry)
    updated = mark_file_validated(file_id, validated_by=validated_by, notes=notes) or entry

    superseded = supersede_matching_files(file_id, entry["original_name"])

    if schedule_rebuild_if_quiet:
        try:
            schedule_rebuild_if_quiet()
        except Exception:
            pass

    return jsonify({"ok": True, "file": _augment(updated), "superseded": superseded})


@app.route("/api/files/<int:file_id>/replace-with", methods=["POST"])
def replace_with_existing(file_id: int):
    """Marca este archivo como reemplazado por otro archivo ya validado en el sistema."""
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404

    if entry.get("status") == "reemplazado":
        return jsonify({"error": "Este archivo ya fue reemplazado"}), 409

    data = request.get_json(silent=True) or {}
    validated_file_id = data.get("validated_file_id")
    if not validated_file_id:
        return jsonify({"error": "Debes indicar el archivo validado"}), 400

    validated_entry = get_file_record(int(validated_file_id))
    if not validated_entry:
        return jsonify({"error": "El archivo validado seleccionado no existe"}), 404

    if validated_entry.get("status") != "validado":
        return jsonify({"error": "El archivo seleccionado no tiene estado 'validado'"}), 400

    supersede_specific_file(file_id, int(validated_file_id))
    updated = get_file_record(file_id)
    return jsonify({"ok": True, "file": _augment(updated)})


@app.route("/api/files/<int:file_id>", methods=["DELETE"])
def delete_file(file_id: int):
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    path = get_file_path(entry)
    try:
        path.unlink(missing_ok=True)
    except Exception:
        pass
    delete_file_record(file_id)
    return jsonify({"ok": True})


@app.route("/api/files/<int:file_id>", methods=["PUT"])
def edit_file(file_id: int):
    """Edita metadatos visibles de un archivo (nombre, responsable y nota)."""
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    if entry.get("status") != "validado":
        return jsonify({"error": "Solo se pueden editar archivos en estado validado"}), 409

    data = request.get_json(silent=True) or {}
    original_name = str(data.get("original_name") or "").strip()
    uploaded_by = str(data.get("uploaded_by") or "").strip()
    notes = str(data.get("notes") or "").strip()

    if not original_name:
        return jsonify({"error": "original_name es obligatorio"}), 400
    if len(original_name) > 255:
        return jsonify({"error": "original_name es demasiado largo"}), 400
    if len(uploaded_by) > 120:
        return jsonify({"error": "uploaded_by es demasiado largo"}), 400
    if len(notes) > 500:
        return jsonify({"error": "notes es demasiado largo"}), 400

    from file_store import _connect
    from datetime import datetime as _dt

    now = _dt.utcnow().isoformat()
    with _connect() as conn:
        conn.execute(
            """
            UPDATE files
            SET original_name = ?, uploaded_by = ?, notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                original_name,
                uploaded_by or None,
                notes or None,
                now,
                file_id,
            ),
        )
    updated = get_file_record(file_id)
    return jsonify({"ok": True, "file": _augment(updated)})


@app.route("/api/files/<int:file_id>/sheet", methods=["GET"])
def get_file_sheet(file_id: int):
    """Devuelve columnas y filas del Excel/CSV para el editor de tabla en el cliente."""
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    if entry.get("status") != "validado":
        return jsonify({"error": "Solo se pueden editar archivos en estado validado"}), 409
    path = get_file_path(entry)
    if not path.exists():
        return jsonify({"error": "El archivo no está en disco."}), 404
    format_key, err = _sheet_format_key(path)
    if err:
        return jsonify({"error": err}), 400
    assert format_key is not None

    editor_name = _editor_name_from_request()
    if not editor_name:
        return jsonify({"error": "Ingrese su nombre para editar el archivo (editor_name)."}), 400
    ok_lock, lock_info = acquire_edit_lock(file_id, editor_name)
    if not ok_lock:
        return _sheet_lock_error_response(lock_info)

    records = _load_records_from_file(path)
    if len(records) > MAX_SHEET_ROWS:
        return (
            jsonify(
                {
                    "error": (
                        f"Demasiadas filas para el editor en línea (máx. {MAX_SHEET_ROWS}). "
                        "Divida el archivo o use Excel localmente."
                    ),
                }
            ),
            413,
        )
    columns = _sheet_column_order(records)
    rows_out = [{c: rec.get(c, "") for c in columns} for rec in records]
    initial_columns = list(columns)
    initial_rows = [{c: r.get(c, "") for c in initial_columns} for r in rows_out]
    columns, rows_out = _ensure_lat_lon_from_coordinates(columns, rows_out)
    columns, rows_out = _ensure_service_conditional_columns(columns, rows_out)
    columns, rows_out = _autofill_asesoria_previa(columns, rows_out)
    columns, rows_out = _autofill_me_ml_by_sexo(columns, rows_out)
    columns, rows_out = _merge_tratamiento_columns(columns, rows_out)
    columns = _reorder_treatment_next_to_diagnostico(columns)
    columns, rows_out = _sanitize_lugar_atencion_missing_values(columns, rows_out)
    columns, rows_out = _ensure_lugar_atencion_by_estado(columns, rows_out)
    columns, rows_out = _normalize_diagnostico_mg_values(columns, rows_out)
    columns = _reorder_especificar_next_to_diagnostico_mg(columns)
    columns, rows_out = _sanitize_disability_sheet_display(columns, rows_out)
    # Persistir auto-completado de columnas para que queden guardadas en el archivo/base.
    if columns != initial_columns or rows_out != initial_rows:
        try:
            _write_sheet_to_path(path, format_key, columns, rows_out)
            st = path.stat()
            from file_store import update_file_size_and_row_count

            update_file_size_and_row_count(file_id, st.st_size, len(rows_out))
            entry = get_file_record(file_id) or entry
        except Exception as exc:  # noqa: BLE001
            logging.warning("No se pudo persistir auto-completado de columnas en %s: %s", path, exc)
    columns_check = _check_columns_against_kobo(columns, rows_out)
    return jsonify(
        {
            "ok": True,
            "file": _augment(entry),
            "format": format_key,
            "editor_name": editor_name,
            "lock": lock_info,
            "lock_ttl_seconds": EDIT_LOCK_TTL_SECONDS,
            "columns": columns,
            "column_display": _sheet_column_display_map(columns),
            "columns_check": columns_check,
            "rows": rows_out,
        }
    )


@app.route("/api/files/<int:file_id>/sheet", methods=["PUT"])
def put_file_sheet(file_id: int):
    """Sobrescribe el Excel/CSV con las filas/columnas editadas y metadatos visibles."""
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    if entry.get("status") != "validado":
        return jsonify({"error": "Solo se pueden editar archivos en estado validado"}), 409
    path = get_file_path(entry)
    if not path.exists():
        return jsonify({"error": "El archivo no está en disco."}), 404
    format_key, err = _sheet_format_key(path)
    if err:
        return jsonify({"error": err}), 400
    assert format_key is not None

    data = request.get_json(silent=True) or {}
    editor_name = _editor_name_from_request(data)
    if not editor_name:
        return jsonify({"error": "editor_name es obligatorio para guardar cambios."}), 400
    ok_lock, lock_info = refresh_edit_lock(file_id, editor_name)
    if not ok_lock:
        return _sheet_lock_error_response(lock_info)

    original_name = str(data.get("original_name") or "").strip()
    notes = str(data.get("notes") or "").strip()
    if not original_name:
        return jsonify({"error": "original_name es obligatorio"}), 400
    if len(original_name) > 255:
        return jsonify({"error": "original_name es demasiado largo"}), 400
    if len(editor_name) > 120:
        return jsonify({"error": "editor_name es demasiado largo"}), 400
    if len(notes) > 500:
        return jsonify({"error": "notes es demasiado largo"}), 400

    columns, cerr = _normalize_put_columns(data.get("columns"))
    if cerr or columns is None:
        return jsonify({"error": cerr or "Solicitud inválida."}), 400
    rows, rerr = _normalize_put_rows(data.get("rows"), columns)
    if rerr or rows is None:
        return jsonify({"error": rerr or "Solicitud inválida."}), 400
    columns, rows = _ensure_service_conditional_columns(columns, rows)
    columns, rows = _autofill_asesoria_previa(columns, rows)
    columns, rows = _autofill_me_ml_by_sexo(columns, rows)
    columns, rows = _merge_tratamiento_columns(columns, rows)
    columns = _reorder_treatment_next_to_diagnostico(columns)
    columns, rows = _sanitize_lugar_atencion_missing_values(columns, rows)
    columns, rows = _normalize_diagnostico_mg_values(columns, rows)
    columns = _reorder_especificar_next_to_diagnostico_mg(columns)
    columns, rows = _sanitize_disability_sheet_display(columns, rows)

    try:
        _write_sheet_to_path(path, format_key, columns, rows)
    except OSError as exc:
        return jsonify({"error": f"No se pudo guardar: {exc}"}), 500

    st = path.stat()
    nrows = len(rows)
    try:
        from file_store import _connect

        now = datetime.utcnow().isoformat()
        with _connect() as conn:
            conn.execute(
                """
                UPDATE files
                SET size_bytes = ?, row_count = ?, updated_at = ?,
                    original_name = ?, uploaded_by = ?, notes = ?,
                    edit_locked_by = ?, edit_lock_at = ?
                WHERE id = ?
                """,
                (
                    st.st_size,
                    nrows,
                    now,
                    original_name,
                    editor_name,
                    notes or None,
                    editor_name,
                    now,
                    file_id,
                ),
            )
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": f"Se guardó el archivo pero no la base de datos: {exc}"}), 500
    updated = get_file_record(file_id)
    return jsonify(
        {
            "ok": True,
            "file": _augment(updated),
            "row_count": nrows,
            "column_display": _sheet_column_display_map(columns),
            "lock": {"locked_by": editor_name, "locked_at": datetime.utcnow().isoformat()},
        }
    )


@app.route("/api/files/<int:file_id>/sheet/unlock", methods=["POST"])
def unlock_file_sheet(file_id: int):
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    if entry.get("status") != "validado":
        return jsonify({"error": "Solo se pueden editar archivos en estado validado"}), 409
    data = request.get_json(silent=True) or {}
    editor_name = _editor_name_from_request(data)
    if not editor_name:
        return jsonify({"error": "editor_name es obligatorio para liberar bloqueo."}), 400
    ok = release_edit_lock(file_id, editor_name=editor_name)
    if not ok:
        lock_info = {
            "locked_by": entry.get("edit_locked_by"),
            "locked_at": entry.get("edit_lock_at"),
        }
        return _sheet_lock_error_response(lock_info)
    return jsonify({"ok": True})


@app.route("/api/files/<int:file_id>/sheet/force-unlock", methods=["POST"])
def force_unlock_file_sheet(file_id: int):
    """Libera un bloqueo de edición aunque pertenezca a otro usuario."""
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    if entry.get("status") != "validado":
        return jsonify({"error": "Solo se pueden editar archivos en estado validado"}), 409

    previous_lock = {
        "locked_by": entry.get("edit_locked_by"),
        "locked_at": entry.get("edit_lock_at"),
    }
    ok = release_edit_lock(file_id, force=True)
    if not ok:
        return jsonify({"error": "No se pudo liberar el bloqueo."}), 500
    return jsonify({"ok": True, "released_lock": previous_lock})


@app.route("/api/files/<int:file_id>/sheet/heartbeat", methods=["POST"])
def heartbeat_file_sheet(file_id: int):
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    if entry.get("status") != "validado":
        return jsonify({"error": "Solo se pueden editar archivos en estado validado"}), 409
    data = request.get_json(silent=True) or {}
    editor_name = _editor_name_from_request(data)
    if not editor_name:
        return jsonify({"error": "editor_name es obligatorio para heartbeat."}), 400
    ok, lock_info = refresh_edit_lock(file_id, editor_name)
    if not ok:
        return _sheet_lock_error_response(lock_info)
    return jsonify({"ok": True, "lock": lock_info, "lock_ttl_seconds": EDIT_LOCK_TTL_SECONDS})


@app.route("/api/files/<int:file_id>/edited-validated", methods=["POST"])
def mark_edited_validated(file_id: int):
    """Marca o desmarca manualmente un archivo validado como 'Editado Validado'."""
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    if entry.get("status") != "validado":
        return jsonify({"error": "Solo se puede marcar en archivos validados"}), 409

    data = request.get_json(silent=True) or {}
    editor_name = _editor_name_from_request(data)
    if not editor_name:
        return jsonify({"error": "editor_name es obligatorio"}), 400
    marked = bool(data.get("edited_validated"))
    now = datetime.utcnow().isoformat()

    from file_store import _connect

    with _connect() as conn:
        if marked:
            conn.execute(
                """
                UPDATE files
                SET edited_validated = 1,
                    edited_validated_by = ?,
                    edited_validated_at = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (editor_name, now, now, file_id),
            )
        else:
            conn.execute(
                """
                UPDATE files
                SET edited_validated = 0,
                    edited_validated_by = NULL,
                    edited_validated_at = NULL,
                    updated_at = ?
                WHERE id = ?
                """,
                (now, file_id),
            )
    updated = get_file_record(file_id)
    return jsonify({"ok": True, "file": _augment(updated)})


@app.route("/api/files/<int:file_id>/sheet/kobo-submit", methods=["POST"])
def submit_file_rows_to_kobo(file_id: int):
    """
    Envía filas seleccionadas del editor de hoja a KoboToolbox (API XML),
    protegido por autorización y contraseña.
    """
    entry = get_file_record(file_id)
    if not entry:
        return jsonify({"ok": False, "error": "Archivo no encontrado"}), 404
    if entry.get("status") != "validado":
        return jsonify({"ok": False, "error": "Solo se puede subir a Kobo desde archivos validados"}), 409

    data = request.get_json(silent=True) or {}
    editor_name = _editor_name_from_request(data)
    if not editor_name:
        return jsonify({"ok": False, "error": "editor_name es obligatorio"}), 400

    ok_lock, lock_info = refresh_edit_lock(file_id, editor_name)
    if not ok_lock:
        return _sheet_lock_error_response(lock_info)

    password = str(data.get("password") or "").strip()
    if password != KOBO_SUBMIT_PASSWORD:
        return jsonify({"ok": False, "error": "Contraseña incorrecta para autorizar el envío a Kobo"}), 403

    token = os.getenv("KOBO_API_TOKEN", "").strip()
    asset_uid = os.getenv("KOBO_ASSET_UID", "").strip()
    kc_url = os.getenv("KOBO_KC_URL", "https://kc.kobotoolbox.org").strip() or "https://kc.kobotoolbox.org"
    if not token or not asset_uid:
        return jsonify({"ok": False, "error": "Faltan KOBO_API_TOKEN o KOBO_ASSET_UID en el servidor"}), 400

    mapping_path_candidates = [
        BASE_DIR / "mapping.yaml",
        BASE_DIR.parent / "mapping.yaml",
    ]
    mapping_path = next((p for p in mapping_path_candidates if p.exists()), None)
    if not mapping_path:
        return jsonify({"ok": False, "error": "No se encontró mapping.yaml para envío a Kobo"}), 500
    mapping = _parse_simple_mapping_yaml(mapping_path)
    if not mapping:
        return jsonify({"ok": False, "error": "mapping.yaml vacío o inválido"}), 500

    # Cargar módulos de transformación/envío de forma dinámica.
    try:
        filling_path = next(
            p for p in [BASE_DIR / "filling_rules.py", BASE_DIR.parent / "filling_rules.py"] if p.exists()
        )
        kobo_api_path = next(
            p for p in [BASE_DIR / "kobo_api.py", BASE_DIR.parent / "kobo_api.py"] if p.exists()
        )
    except StopIteration:
        return jsonify({"ok": False, "error": "Faltan módulos filling_rules.py o kobo_api.py en servidor"}), 500

    try:
        fill_spec = importlib.util.spec_from_file_location("koboup_filling_rules", filling_path)
        fill_mod = importlib.util.module_from_spec(fill_spec)
        assert fill_spec and fill_spec.loader
        fill_spec.loader.exec_module(fill_mod)
        apply_rules = getattr(fill_mod, "apply_rules")
    except Exception as exc:
        return jsonify({"ok": False, "error": f"No se pudo cargar filling_rules.py: {exc}"}), 500

    try:
        api_spec = importlib.util.spec_from_file_location("koboup_kobo_api", kobo_api_path)
        api_mod = importlib.util.module_from_spec(api_spec)
        assert api_spec and api_spec.loader
        api_spec.loader.exec_module(api_mod)
        submit_via_api = getattr(api_mod, "submit_via_api")
    except Exception as exc:
        return jsonify({"ok": False, "error": f"No se pudo cargar kobo_api.py: {exc}"}), 500

    raw_columns = data.get("columns")
    raw_rows = data.get("rows")
    if isinstance(raw_columns, list) and isinstance(raw_rows, list):
        columns, cerr = _normalize_put_columns(raw_columns)
        if cerr or columns is None:
            return jsonify({"ok": False, "error": cerr or "columns inválidas"}), 400
        rows, rerr = _normalize_put_rows(raw_rows, columns)
        if rerr or rows is None:
            return jsonify({"ok": False, "error": rerr or "rows inválidas"}), 400
    else:
        path = get_file_path(entry)
        if not path.exists():
            return jsonify({"ok": False, "error": "El archivo no está en disco"}), 404
        records = _load_records_from_file(path)
        columns = _sheet_column_order(records)
        rows = [{c: rec.get(c, "") for c in columns} for rec in records]
        columns, rows = _ensure_lat_lon_from_coordinates(columns, rows)
        columns, rows = _ensure_service_conditional_columns(columns, rows)
        columns, rows = _autofill_asesoria_previa(columns, rows)

    columns, rows = _autofill_me_ml_by_sexo(columns, rows)

    columns, rows = _merge_tratamiento_columns(columns, rows)
    columns = _reorder_treatment_next_to_diagnostico(columns)
    columns, rows = _sanitize_lugar_atencion_missing_values(columns, rows)
    columns, rows = _ensure_lugar_atencion_by_estado(columns, rows)
    columns, rows = _normalize_diagnostico_mg_values(columns, rows)
    columns = _reorder_especificar_next_to_diagnostico_mg(columns)
    columns, rows = _sanitize_disability_sheet_display(columns, rows)

    if not rows:
        return jsonify({"ok": False, "error": "No hay filas para enviar"}), 400

    submit_all = bool(data.get("submit_all"))
    selected = data.get("row_indices")
    if submit_all:
        selected_idx = list(range(len(rows)))
    else:
        if not isinstance(selected, list) or not selected:
            return jsonify({"ok": False, "error": "Seleccione al menos una fila para enviar"}), 400
        selected_idx = []
        for x in selected:
            try:
                i = int(x)
            except (TypeError, ValueError):
                continue
            if 0 <= i < len(rows):
                selected_idx.append(i)
        selected_idx = sorted(set(selected_idx))
        if not selected_idx:
            return jsonify({"ok": False, "error": "No se seleccionaron filas válidas"}), 400

    if len(selected_idx) > 1200:
        return jsonify({"ok": False, "error": "Máximo 1200 filas por envío para evitar bloqueo del servidor"}), 400

    sent = 0
    failed = 0
    errors: list[dict] = []
    for idx in selected_idx:
        source_row = rows[idx] or {}
        internal_record = _to_kobo_internal_record(source_row)
        try:
            payload_record = apply_rules(internal_record)
        except Exception as exc:
            failed += 1
            errors.append({"row_index": idx, "excel_row": idx + 2, "error": f"Error de reglas: {exc}"})
            continue

        try:
            ok, msg = submit_via_api(
                payload_record,
                mapping,
                api_token=token,
                asset_uid=asset_uid,
                kc_url=kc_url,
            )
        except Exception as exc:
            ok, msg = False, f"Error de envío: {exc}"

        if ok:
            sent += 1
        else:
            failed += 1
            errors.append({"row_index": idx, "excel_row": idx + 2, "error": msg})

    ok_all = failed == 0
    try:
        log_kobo_submission(
            file_id=file_id,
            file_name=str(entry.get("original_name") or entry.get("stored_name") or ""),
            submitted_by=editor_name,
            selected_total=len(selected_idx),
            sent_count=sent,
            failed_count=failed,
            details={
                "partial": (sent > 0 and failed > 0),
                "error_count": len(errors),
            },
        )
    except Exception:
        app.logger.exception("No se pudo registrar bitácora de envío Kobo")
    return jsonify(
        {
            "ok": ok_all,
            "partial": (sent > 0 and failed > 0),
            "sent": sent,
            "failed": failed,
            "selected_total": len(selected_idx),
            "errors": errors[:120],
            "message": (
                f"Envío a Kobo completado. Enviadas: {sent}, fallidas: {failed}."
                if ok_all or sent > 0
                else "No se pudo enviar ninguna fila a Kobo."
            ),
        }
    ), (200 if (ok_all or sent > 0) else 500)


@app.route("/api/kobo-submissions/logs", methods=["GET"])
def get_kobo_submission_logs():
    raw_limit = request.args.get("limit", "100")
    try:
        limit = int(raw_limit)
    except Exception:
        limit = 100
    logs = list_kobo_submission_logs(limit=limit)
    return jsonify({"ok": True, "logs": logs, "count": len(logs)})


# ── Reference PDFs ─────────────────────────────────────────────────


def _augment_ref(entry: dict) -> dict:
    if not entry:
        return entry
    entry = dict(entry)
    rid = entry.get("id")
    if rid is not None:
        entry["download_url"] = f"api/refs/{rid}/download"
    return entry


@app.route("/api/refs", methods=["GET"])
def list_refs():
    location = (request.args.get("location") or "").strip() or None
    refs = [_augment_ref(r) for r in list_ref_records(location)]
    return jsonify({"ok": True, "refs": refs})


@app.route("/api/refs/locations", methods=["GET"])
def get_locations():
    return jsonify({"ok": True, "locations": list_ref_locations()})


@app.route("/api/refs", methods=["POST"])
def upload_ref():
    if "file" not in request.files:
        return jsonify({"error": "No se envió ningún archivo"}), 400
    uploaded = request.files["file"]
    if uploaded.filename == "":
        return jsonify({"error": "Nombre de archivo vacío"}), 400
    if not uploaded.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Solo se permiten archivos PDF"}), 400

    location = (request.form.get("location") or "").strip()
    if not location:
        return jsonify({"error": "La ubicación es obligatoria"}), 400

    original_name = uploaded.filename
    safe_name = secure_filename(original_name) or "referencia.pdf"
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    stored_name = f"{timestamp}_{safe_name}"

    REFERENCES_DIR.mkdir(exist_ok=True, parents=True)
    dest_path = REFERENCES_DIR / stored_name
    uploaded.save(dest_path)
    size_bytes = dest_path.stat().st_size if dest_path.exists() else None

    uploaded_by = (request.form.get("uploaded_by") or "").strip() or None
    notes = (request.form.get("notes") or "").strip() or None

    record = add_ref_record(
        original_name=original_name,
        stored_name=stored_name,
        location=location,
        uploaded_by=uploaded_by,
        notes=notes,
        size_bytes=size_bytes,
    )
    return jsonify({"ok": True, "ref": _augment_ref(record)})


@app.route("/api/refs/<int:ref_id>/download", methods=["GET"])
def download_ref(ref_id: int):
    entry = get_ref_record(ref_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    path = get_ref_file_path(entry)
    if not path.exists():
        return jsonify({"error": "El archivo ya no está disponible"}), 404
    return send_file(
        str(path),
        as_attachment=True,
        download_name=entry.get("original_name") or entry.get("stored_name"),
    )


@app.route("/api/refs/<int:ref_id>", methods=["DELETE"])
def delete_ref(ref_id: int):
    entry = get_ref_record(ref_id)
    if not entry:
        return jsonify({"error": "Archivo no encontrado"}), 404
    path = get_ref_file_path(entry)
    try:
        path.unlink(missing_ok=True)
    except Exception:
        pass
    delete_ref_record(ref_id)
    return jsonify({"ok": True})


BULK_DOWNLOAD_PASSWORD = "vamoscontodo"


@app.route("/api/files/download-validated-zip", methods=["POST"])
def download_validated_zip():
    """Descarga masiva de todos los archivos validados en un .zip protegido por contraseña."""
    data = request.get_json(silent=True) or {}
    password = (data.get("password") or "").strip()
    if password != BULK_DOWNLOAD_PASSWORD:
        return jsonify({"error": "Contraseña incorrecta"}), 403

    validated = list_file_records("validado")
    if not validated:
        return jsonify({"error": "No hay archivos validados para descargar"}), 404

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        seen_names: dict[str, int] = {}
        for entry in validated:
            path = get_file_path(entry)
            if not path.exists():
                continue
            dl_name = entry.get("original_name") or entry.get("stored_name")
            if dl_name in seen_names:
                seen_names[dl_name] += 1
                stem = Path(dl_name).stem
                ext = Path(dl_name).suffix
                dl_name = f"{stem} ({seen_names[dl_name]}){ext}"
            else:
                seen_names[dl_name] = 0
            zf.write(path, dl_name)

    buf.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"archivos_validados_{timestamp}.zip",
    )


@app.route("/api/stats/ranking", methods=["GET"])
def get_ranking():
    validators = get_validator_stats()
    uploaders = get_uploader_stats()
    return jsonify({"ok": True, "validators": validators, "uploaders": uploaders})


@app.route("/api/stats/records", methods=["GET"])
def records_stats():
    """Retorna estadísticas de registros (filas) de archivos validados."""
    stats = get_record_stats()
    return jsonify({"ok": True, **stats})


@app.route("/api/stats/kpis", methods=["GET"])
def public_kpis():
    """Retorna KPI públicos reconstruidos desde archivos validados."""
    try:
        return jsonify({"ok": True, "kpis": _compute_public_kpis()})
    except Exception as exc:
        logging.exception("Error al calcular KPI públicos")
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/stats/kpis/exclusions", methods=["POST"])
def update_kpi_exclusions():
    data = request.get_json(silent=True) or {}
    file_name = str(data.get("file_name") or "").strip()
    excluded = bool(data.get("excluded"))
    if not file_name:
        return jsonify({"ok": False, "error": "Falta file_name"}), 400

    exclusions = _load_kpi_exclusions()
    if excluded:
        exclusions.add(file_name)
    else:
        exclusions.discard(file_name)
    _save_kpi_exclusions(exclusions)
    return jsonify({"ok": True, "excluded_files": sorted(exclusions)})


@app.route("/api/stats/recount", methods=["POST"])
def recount_all():
    """Recuenta filas de todos los archivos Excel/CSV que no tengan row_count."""
    from file_store import _connect

    with _connect() as conn:
        rows = conn.execute(
            "SELECT id, stored_name, status, file_type FROM files WHERE row_count IS NULL AND file_type != 'pdf'"
        ).fetchall()

    updated = 0
    for row in rows:
        entry = {"stored_name": row["stored_name"], "status": row["status"]}
        path = get_file_path(entry)
        if path.exists():
            rc = count_file_rows(path)
            if rc is not None:
                update_row_count(row["id"], rc)
                updated += 1

    return jsonify({"ok": True, "updated": updated, "total_checked": len(rows)})


def _search_needle_in_text(needle: str, text: str) -> bool:
    if not text or not needle:
        return False
    return needle in _norm_text(text)


@app.route("/api/search/validated", methods=["GET"])
def search_validated_in_files():
    """
    Busca texto en archivos en estado *validado* (Excel, CSV) y en metadatos
    (nombre, nota, quien subió). Ignora mayúsculas/acentos vía _norm_text.
    """
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"ok": False, "error": "Escriba al menos 2 caracteres."}), 400
    needle = _norm_text(q)
    if not needle:
        return jsonify({"ok": False, "error": "Búsqueda inválida."}), 400

    try:
        limit = int(request.args.get("limit", 200))
    except (TypeError, ValueError):
        limit = 200
    limit = max(1, min(limit, 1000))

    try:
        per_file = int(request.args.get("per_file", 20))
    except (TypeError, ValueError):
        per_file = 20
    per_file = max(1, min(per_file, 100))

    file_id = request.args.get("file_id", type=int)
    scope = (request.args.get("scope") or "all").strip().lower()
    if scope not in ("all", "rows", "metadata"):
        scope = "all"

    results: list[dict] = []
    files_scanned = 0
    truncated = False

    entries = list_file_records("validado")
    if file_id is not None:
        entries = [e for e in entries if e.get("id") == file_id]
        if not entries:
            return jsonify({"ok": False, "error": "Archivo no encontrado o no es validado."}), 404

    for entry in entries:
        eid = entry.get("id")
        name = str(entry.get("original_name") or entry.get("stored_name") or "")
        stored = str(entry.get("stored_name") or "")
        notes = str(entry.get("notes") or "")
        uploader = str(entry.get("uploaded_by") or "")

        files_scanned += 1
        ftype = str(entry.get("file_type") or "").lower()

        meta_blob = f"{name} {stored} {notes} {uploader}"
        meta_hits = scope in ("all", "metadata") and _search_needle_in_text(needle, meta_blob)

        if scope in ("all", "metadata") and meta_hits:
            results.append(
                {
                    "file_id": eid,
                    "file_name": name,
                    "match_in": "Archivo (nombre, nota o responsable)",
                    "column": "—",
                    "row_index": None,
                    "excel_row": None,
                    "value": name or stored,
                }
            )
            if len(results) >= limit:
                truncated = True
                break

        if len(results) >= limit:
            truncated = True
            break

        if scope == "metadata":
            continue

        if ftype == "pdf":
            continue

        path = get_file_path(entry)
        if not path.exists():
            continue
        suf = path.suffix.lower()
        if suf not in {".xlsx", ".xls", ".csv", ".xlsm"}:
            continue

        try:
            records = _load_records_from_file(path)
        except Exception as exc:  # noqa: BLE001
            logging.warning("Búsqueda: no se pudo leer %s: %s", path, exc)
            continue

        n_hits_file = 0
        for ri, rec in enumerate(records or []):
            if len(results) >= limit or n_hits_file >= per_file:
                if len(results) >= limit:
                    truncated = True
                break
            for col, val in (rec or {}).items():
                if n_hits_file >= per_file or len(results) >= limit:
                    break
                sval = str(val) if val is not None else ""
                if not sval.strip():
                    continue
                if not _search_needle_in_text(needle, sval):
                    continue
                display = sval if len(sval) < 2000 else f"{sval[:2000]}…"
                results.append(
                    {
                        "file_id": eid,
                        "file_name": name,
                        "match_in": "Celda (Excel/CSV)",
                        "column": col,
                        "row_index": ri,
                        "excel_row": ri + 2,
                        "value": display,
                    }
                )
                n_hits_file += 1
                if len(results) >= limit:
                    truncated = True
                    break
            if truncated and len(results) >= limit:
                break

    return jsonify(
        {
            "ok": True,
            "query": q,
            "scope": scope,
            "results": results,
            "files_scanned": files_scanned,
            "result_count": len(results),
            "truncated": truncated,
        }
    )


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s: %(message)s",
    )
    port = int(os.environ.get("PORT", 5002))
    print(f"\n  KoboUp disponible en http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
