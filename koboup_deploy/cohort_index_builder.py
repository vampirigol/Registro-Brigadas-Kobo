# -*- coding: utf-8 -*-
"""
Construye cohort_treatment_index.json a partir de:
- Archivos con estado "validado" en la BD (uploads/validados)
- Además, archivos bajo la carpeta prioritaria (paths_config)
"""
from __future__ import annotations

import csv
import json
import re
import threading
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from treatment_criteria import (
    age_to_band,
    column_looks_dx_espir,
    column_looks_dx_mg_list,
    column_looks_edad,
    column_looks_service,
    column_looks_tratamiento_o_insumo,
    norm_text,
    parse_age,
    treatment_text_has_dose_or_presence,
)
from paths_config import PRIORITY_VALIDATED_DIR

BASE = Path(__file__).resolve().parent
DATA = BASE / "data"
INDEX_PATH = DATA / "cohort_treatment_index.json"
STOP_BUILD_LOCK = threading.Lock()
_LAST_BUILD: float = 0.0


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, encoding=enc, newline="") as f:
                sample = f.read(16384)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
                except csv.Error:
                    dialect = csv.excel
                r = csv.reader(f, dialect)
                rows = [x for x in r]
        except (UnicodeDecodeError, OSError, csv.Error):
            continue
        if not rows:
            continue
        return (rows[0] if rows else []), (rows[1:] if len(rows) > 1 else [])
    return [], []


def _read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return [], []
    out_rows: list[list[str]] = []
    headers: list[str] = []
    try:
        for sheet in wb.worksheets:
            rows = sheet.iter_rows()
            hrow = next(rows, None)
            if not hrow:
                continue
            headers = [str(c.value or "") for c in hrow]
            for row in rows:
                out_rows.append([c.value for c in row])
            if headers and out_rows:
                break
    finally:
        wb.close()
    return headers, out_rows


def _iter_data_paths() -> Iterator[Path]:
    seen: set[str] = set()
    try:
        from file_store import get_file_path, list_file_records

        for rec in list_file_records("validado"):
            p = get_file_path(rec)
            r = str(p.resolve())
            if p.exists() and p.suffix.lower() in (".xlsx", ".xls", ".csv") and r not in seen:
                seen.add(r)
                yield p
    except Exception:
        pass

    if PRIORITY_VALIDATED_DIR.exists():
        for p in sorted(PRIORITY_VALIDATED_DIR.iterdir(), key=lambda x: x.name.lower()):
            if not p.is_file():
                continue
            # Archivos sólo de demostración (p. ej. DEMO_Columnas_API_Kobo.xlsx)
            if p.name.upper().startswith("DEMO_"):
                continue
            r = str(p.resolve())
            if p.suffix.lower() not in (".xlsx", ".xls", ".csv") or r in seen:
                continue
            seen.add(r)
            yield p


def _row_values(headers: list[str], row: list[Any], flags: list[bool]) -> str:
    n = max(len(headers), len(row))
    ext = list(row) + [""] * (n - len(row))
    parts: list[str] = []
    for i, f in enumerate(flags):
        if not f or i >= n:
            continue
        v = ext[i]
        if v is None or str(v).strip() == "":
            continue
        parts.append(str(v).strip())
    return " | ".join(parts)


def build_cohort_treatment_index() -> dict[str, Any]:
    entries: list[dict[str, Any]] = []
    counter: Counter = Counter()  # (dk, band, sk, tx) -> count
    file_count = 0
    row_n = 0
    for path in _iter_data_paths():
        if path.suffix.lower() == ".csv":
            headers, data = _read_csv(path)
        else:
            headers, data = _read_xlsx(path)
        if not headers or not data:
            continue
        hflags_dx = [column_looks_dx_espir(h) for h in headers]
        hflags_mg = [column_looks_dx_mg_list(h) for h in headers]
        hflags_tx = [column_looks_tratamiento_o_insumo(h) for h in headers]
        hflags_age = [column_looks_edad(h) for h in headers]
        hflags_se = [column_looks_service(h) for h in headers]
        if not any(hflags_tx):
            continue
        file_count += 1
        for row in data:
            if not any(row):
                continue
            age_val = _row_values(headers, row, hflags_age)
            age = parse_age(age_val) if age_val else None
            band = age_to_band(age)
            serv = _row_values(headers, row, hflags_se) or "desconocida"
            sk = norm_text(serv)[:60]
            dx1 = _row_values(headers, row, hflags_dx)
            if not str(dx1).strip():
                dx1 = _row_values(headers, row, hflags_mg)
            dk = norm_text(dx1)[:200] or "_sin_diagnostico"
            t_raw = _row_values(headers, row, hflags_tx)
            if not t_raw.strip() or not treatment_text_has_dose_or_presence(t_raw):
                continue
            tnorm = re.sub(r"\s+", " ", t_raw).strip()[:500]
            keyb = f"{dk}\t{band}\t{sk}\t{tnorm}"
            counter[keyb] += 1
            row_n += 1

    for k, c in counter.items():
        parts = k.split("\t", 3)
        if len(parts) < 4:
            continue
        entries.append(
            {
                "dxk": parts[0],
                "edad_banda": parts[1],
                "servk": parts[2],
                "texto": parts[3],
                "c": c,
            }
        )

    return {
        "version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "file_count": file_count,
        "fila_tto_con_dosis": row_n,
        "entries": entries,
    }


def write_cohort_index(data: dict[str, Any], path: Path | None = None) -> Path:
    path = path or INDEX_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=0)
    return path


def build_and_write() -> dict[str, Any]:
    d = build_cohort_treatment_index()
    write_cohort_index(d)
    return d


def schedule_rebuild_if_quiet() -> None:
    """Reconstrucción en segundo plano, como mucho 1 por minuto (anti-spam al validar muchos)."""
    global _LAST_BUILD

    def _run() -> None:
        global _LAST_BUILD
        with STOP_BUILD_LOCK:
            now = time.time()
            if now - _LAST_BUILD < 60:
                return
            _LAST_BUILD = now
        try:
            build_and_write()
        except Exception:  # noqa: BLE001
            pass

    t = threading.Thread(target=_run, daemon=True, name="cohort-index")
    t.start()
