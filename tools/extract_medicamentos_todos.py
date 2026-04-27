#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrae cadenas de medicamentos de CSV y XLSX bajo
Llenado Kobo tools y Llenado Kobo tools.bak_20260320 (excluye venv, node_modules, site-packages).
Solo columnas cuyo encabezado sugiere tratamiento, medicamentos, plan, insumos, especificar, dosis, etc.
Uso: python tools/extract_medicamentos_todos.py [--out salida.txt]
"""
from __future__ import annotations

import csv
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Iterator

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

ROOTS = [
    Path(__file__).resolve().parent.parent,
    Path(__file__).resolve().parent.parent.parent / "Llenado Kobo tools.bak_20260320",
]

SKIP_DIR_PARTS = frozenset(
    {
        "venv",
        "node_modules",
        "site-packages",
        ".git",
        "__pycache__",
    }
)

COL_HINTS = (
    "medicamento",
    "tratamiento",
    "plan de tratamiento",
    "plan_tratamiento",
    "plan tratamiento",
    "especif",
    "especique",
    "insumo",
    "detalle",
    "nombres espec",
    "tx",
    "indicad",
    "recet",
    "dosis",
    "unidades_entregadas",
    "unidades entregad",
    "categoria general",
    "categoría general",
)

STOP_SEGMENTS = re.compile(
    r"^(lentes?|lente|agua(?!.*mg)|cloro|rayos?|rayo x|rx \d|radiograf|"
    r"compres|venda|yeso|vacun|higiene|fluor|protesis|pr[oó]tesis|"
    r"kit dental|kit higiene|kit\b(?!\s*\w*mg)|"
    r"ejercicio|fisioterap$|fisioterap\w*$|kinesio|odonto$|dental$|"
    r"trape|trapo|trepiant|m[oó]vil|referencia$|n/?d$|n/a$|s[ií]$|"
    r"no aplica$|mediciones?)$",
    re.I,
)

# Términos sueltos que nunca son un medicamento
NOISE_EXACT = frozenset(
    s.strip().lower()
    for s in (
        "medicina general",
        "odontología",
        "oftalmología",
        "fisioterapia",
        "laboratorio clínico",
        "laboratorios",
        "laboratorio",
        "dental",
        "medicamento",
        "medicamentos",
        "insumos",
        "insumo",
        "médicamentos",
        "mexicana",
        "mexicano",
        "méxico",
        "mexico",
        "femenino",
        "masculino",
        "atención única",
        "atencion unica",
        "atencion",
        "primera vez",
        "seguimiento",
        "móvil",
        "móbile",
        "mobil",
        "sí",
        "si",
        "no",
        "otro",
        "otros",
        "general",
        "revisión",
        "revision",
        "consulta general",
        "consulta dental",
        "consulta de rutina",
        "chequeo general",
        "revisión general",
        "revision general",
        "cuidadora mujer",
        "solo/a",
        "sola",
        "submitted_via_web",
        "medicamento (",
        "tab c",
    )
    if s
)

# Presentación: si aparece, el fragmento es candidato
HAS_PRESENTATION = re.compile(
    r"(\d{1,4}[\s.,]*\s*mg\b|\bmg\b|\bml\b|\bgr\.?\b|tabs?\.?|"
    r"compr(imidos?)?|c[aá]ps?\.?|amp(ollas?|\.?|ol)|"
    r"u\.?i\.?|\bui\b|frasco|jeringa|dosis|"
    r"\b\d{1,3}\s*tab(s)?\b|diaria\s*x)",
    re.I,
)

# Ampliar DCI / marcas frecuentes (sin grupos anidados para evitar errores de re)
KNOWN_INN = re.compile(
    r"\b("
    r"paracetamol|acetaminofén|acetaminofen|ibuprofeno|naproxeno|ketorolaco|diclofenaco|diclopf|"
    r"amoxicilina|acido clavul|ácido clavul|clavulánico|clavulánico|"
    r"ampicilina|azitromicina|cefalexina|metronidazol|"
    r"albendazol|albendaz|ivermectina|praziquantel|"
    r"clindamicina|doxiciclina|levofloxacino|ofloxacina|ciprofloxacina|"
    r"trimetoprim|sulfametox|nitrofurantoina|norfloxacino|"
    r"omeprazol|pantoprazol|lansoprazol|esomprazol|"
    r"metoclopramida|domperidona|dimeticona|ondansetron|"
    r"loratadina|levocetirizina|difenhidramina|cetirizina|"
    r"salbutamol|ambroxol|budesonida|fluticasona|"
    r"metformina|glibenclamida|gliclazida|glimepirida|acarbosa|"
    r"insulina|losartán|losartan|enalapril|lisinopril|amlodipino|furosemida|"
    r"hidroclorotiazida|isosorbide|dinitrato|atenolol|propranolol|carvedilol|"
    r"tiamazol|metimazol|alprazolam|clonazepam|diazepam|"
    r"ácido fólico|acido folico|acido ursodeoxic|ácido|acido graso|fierro|hierro|complejo b|"
    r"vitamina a|vitamina b|vitamina c|vitamina d|multivitam|"
    r"hipromelosa|artifici|tropicamida|fenilefrina|tetrazolina|"
    r"metocarbamol|orfenadrina|nimesulida|celecoxib|meloxicam|piroxicam|indometacina|"
    r"buscapin|butilhioscina|hioscina|fenspir|benzatin|"
    r"convifer|cilocid|vermisen|vermisan|ligui|miconazol|terbinafina|clotrimazol|"
    r"ketoconazol|nystatin|nistat"
    r")\b",
    re.I,
)

MED_SKELETON = re.compile(
    r"^([a-záéíóúüñ0-9](?:[a-záéíóúüñ0-9\-\.][a-záéíóúüñ0-9\s\-\+\.']{0,50})?)"
    r'(?:\s*[\(]?\d{1,4}[\s.,]?\d{0,3}\s*mg\)?|'
    r"\s*tabs?\.?|\s*compr(imi)?\.?|"
    r"\s*\d{1,3}\s*tab\w*|\s*c[aá]ps?\.?|\s*ub\.?)?$",
    re.I,
)

LAB_LIKE = re.compile(r"^(gluc|colest|trig|hb|hierro|hemoglob|gota|glic)[a-z]*\s*:\s*", re.I)

UUID_LIKE = re.compile(r"^[a-z0-9]{20,}$", re.I)
HASH_KOBO = re.compile(r"^vcu|kobo|formhub", re.I)


def iter_data_files(roots: list[Path]) -> Iterator[Path]:
    for base in roots:
        if not base.is_dir():
            continue
        for p in base.rglob("*"):
            if not p.is_file():
                continue
            parts_lower = {x.lower() for x in p.parts}
            if SKIP_DIR_PARTS & parts_lower:
                continue
            if p.suffix.lower() in (".csv", ".xlsx"):
                yield p


def norm_col(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def column_looks_med_relevant(h: str) -> bool:
    n = norm_col(h)
    return any(hint in n for hint in COL_HINTS)


def split_segments(text: str) -> list[str]:
    if not text or not str(text).strip():
        return []
    t = str(text)
    t = t.replace(";", ",")
    t = t.replace("+", ", ")
    t = t.replace("|", ", ")
    if "/" in t and len(t) < 200 and t.count("/") < 4:
        t = t.replace("/", ", ")
    parts = re.split(r"[,]|\n", t)
    out: list[str] = []
    for p in parts:
        p = p.strip().strip('"').strip("'")
        p = re.sub(r"\s+", " ", p)
        if p:
            out.append(p)
    return out


def clean_segment(s: str) -> str:
    s = s.strip()
    s = re.sub(r"^[\-•\d\.\s]+", "", s)
    s = re.sub(r"^(y|e|and)\s+", "", s, flags=re.I)
    s = re.sub(r"\b\d+-\d+-\d+\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" ,.;")
    return s


def is_obvious_noise(s: str) -> bool:
    sl = s.strip().lower()
    if not sl or len(sl) < 2:
        return True
    if sl in NOISE_EXACT:
        return True
    if UUID_LIKE.match(sl) or HASH_KOBO.search(sl):
        return True
    if LAB_LIKE.search(sl) and "mg" not in sl and "tab" not in sl:
        return True
    if re.search(r"^(fecha|estado|nombre|edad|coord|submitted|meta/)", sl, re.I):
        return True
    if re.search(r"duplicado|kobo|_uuid|deviceid|instanceid", sl, re.I):
        return True
    # Lugares / orgs frecuentes
    if re.search(
        r"^(baja californ(ia)?|chihuahua|sonora|loreto|obreg[oó]n|ju[aá]rez|"
        r"vizcaino|mexic(o|a)|tembabiche|montemorelos|paz|cl[ií]nica|adventista|"
        r"nuevo le[oó]n|para[ií]so|iglesia|ampliaci|felipe|angeles|mateos|carlos|"
        r"com[óo]m?and[úu]|l[oó]pez|san carlos|puerto |colonia )",
        sl,
        re.I,
    ):
        return True
    if re.match(
        r"^(limpieza|extracci[oó]n|extraccion|consulta|anteojos|lentes\s*\d?|"
        r"rayos? x|toma de presi[óo]?n|resultados de lab|plan de trat\w*|"
        r"tratamiento$|ejercicio|insumos de|kid dental|inhalador|toma 1$|toma 2$|icd$|inhal\w*)$",
        sl,
        re.I,
    ):
        return True
    if re.match(
        r"^(horas?|hrs?|h\)?\s*x|tab diaria|tomar 1$|tomar 2$|dosis unica|ml dosis|"
        r"h\)?\s*x\s*\d|hrs? por|horas? por|horas? x|hrs? x).*$",
        sl,
        re.I,
    ) and "mg" not in sl and "albendazol" not in sl and "cilocid" not in sl and "verm" not in sl:
        return True
    return False


# Marcas / presentaciones frecuentes en brigadas (complementa KNOWN_INN)
BRAND_OFT = re.compile(
    r"vermisen|vermisan|vermisa|vermiso|cilocid|cilac|ciloci|convi?fer|conviter|kamina|"
    r"rocavi|rocavit|lumboxen|ligui|kamic|gotas? oft|lagri",
    re.I,
)


def is_likely_medication_for_summary(s: str) -> bool:
    """Para listado 'solo fármacos': DCI, marca frecuente, o frase corta con mg/tab."""
    if is_obvious_noise(s) or len(s) > 220:
        return False
    if KNOWN_INN.search(s):
        return True
    if BRAND_OFT.search(s) and "colonia" not in s.lower():
        return True
    if HAS_PRESENTATION.search(s) and re.search(
        r"[a-záéíóúñ]{4,}", s, re.I
    ) and not re.search(r"^(inhalador vacío|toma 1$|tomar 1$|hrs)", s, re.I):
        return not is_obvious_noise(s)
    return False


def segment_is_medication(s: str) -> bool:
    c = clean_segment(s)
    if len(c) < 3:
        return False
    if is_obvious_noise(c):
        return False
    if STOP_SEGMENTS.search(c) and not HAS_PRESENTATION.search(c) and not KNOWN_INN.search(c):
        return False
    if re.match(r"^(Sí|SI|No|N/?D|N/A|NA)\b", c) and "mg" not in c:
        return False
    if c.isdigit():
        return False
    if re.match(r"^medicamentos?\s*[\(]?\d", c, re.I) and "mg" not in c and "tab" not in c.lower():
        return False
    if KNOWN_INN.search(c):
        return True
    if HAS_PRESENTATION.search(c):
        return not is_obvious_noise(c)
    # "Marca (principio) 20 tab 5 mg" sin palabra de la regex INN
    if re.search(
        r"\b\d+\s+tab|\d+\s*mg|tabs?\.?|compr(imi|imidos?)?\b", c, re.I
    ) and len(c) < 200:
        return not is_obvious_noise(re.sub(r"\b\d+[\s.,/]*\d*\s*mg.*", "", c, flags=re.I).strip() or c)
    if MED_SKELETON.match(c):
        return True
    return False


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, encoding=enc, newline="") as f:
                sample = f.read(16384)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
                except csv.Error:
                    dialect = csv.excel
                r = csv.reader(f, dialect)
                rows = [row for row in r]
        except (UnicodeDecodeError, OSError):
            continue
        if not rows:
            continue
        return rows[0] if rows else [], rows[1:]
    return [], []


def iter_xlsx_treatment_columns(path: Path) -> Iterator[str]:
    if not load_workbook:
        return
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return
    try:
        for sheet in wb.worksheets:
            try:
                rows = sheet.iter_rows()
                header_row = next(rows, None)
            except Exception:
                continue
            if not header_row:
                continue
            headers: list[str] = []
            for c in header_row:
                v = c.value
                if v is None:
                    headers.append("")
                else:
                    headers.append(str(v).replace("\n", " "))
            med_idx = {
                i for i, h in enumerate(headers) if h and column_looks_med_relevant(h)
            }
            if not med_idx:
                continue
            for row in rows:
                for i in med_idx:
                    if i >= len(row):
                        continue
                    cell = row[i]
                    v = cell.value
                    if v is None or str(v).strip() == "":
                        continue
                    yield str(v)
    finally:
        wb.close()


def process() -> tuple[Counter, int, int, int]:
    """Retorna (conteos, files_ok, files_skipped, files_err)"""
    counts: Counter = Counter()
    files_ok = 0
    files_skipped = 0
    files_err = 0

    for path in sorted(iter_data_files(ROOTS)):
        try:
            if path.suffix.lower() == ".csv":
                headers, rows = read_csv(path)
                if not headers:
                    files_err += 1
                    continue
                med_cols = {i for i, h in enumerate(headers) if h and column_looks_med_relevant(str(h))}
                if not med_cols:
                    files_skipped += 1
                    continue
                for row in rows:
                    for i in med_cols:
                        if i >= len(row):
                            continue
                        raw = row[i]
                        if raw is None or str(raw).strip() == "":
                            continue
                        for seg in split_segments(str(raw)):
                            c = clean_segment(seg)
                            if not segment_is_medication(c):
                                continue
                            key = re.sub(r"\s+", " ", c).strip()
                            if key:
                                counts[key.lower()] += 1
                files_ok += 1
            else:
                n = 0
                for val in iter_xlsx_treatment_columns(path):
                    n += 1
                    for seg in split_segments(str(val)):
                        c = clean_segment(seg)
                        if not segment_is_medication(c):
                            continue
                        key = re.sub(r"\s+", " ", c).strip()
                        if key:
                            counts[key.lower()] += 1
                if n == 0:
                    files_skipped += 1
                else:
                    files_ok += 1
        except Exception:
            files_err += 1
            continue

    return counts, files_ok, files_skipped, files_err


def display_name(key_lower: str) -> str:
    if not key_lower:
        return key_lower
    return key_lower[0].upper() + key_lower[1:]


def main() -> None:
    out_path = None
    if len(sys.argv) >= 3 and sys.argv[1] == "--out":
        out_path = Path(sys.argv[2])
    print("Recopilando medicamentos (columnas de tratamiento/insumos)…", file=sys.stderr)
    counts, ok, skipped, err = process()
    items = sorted(counts.items(), key=lambda x: (-x[1], x[0]))
    lines = [
        f"# Archivos con columnas de medicamento/tratamiento leídos: {ok}",
        f"# Archivos omitidos (sin columna adecuada o vacío): {skipped}",
        f"# Archivos con error: {err}",
        f"# Términos distintos: {len(items)}",
        "# Orden: frecuencia desc, luego alfabeto",
        "# Incluye cadenas libres, marcas, pautas y algunos ruidos residuales; ver *_solo_dci_marcas* para lista más limpia.",
        "",
    ]
    for k, n in items:
        lines.append(f"{n:6d} | {display_name(k)}")

    text = "\n".join(lines) + "\n"

    clean_items = [(k, n) for k, n in items if is_likely_medication_for_summary(k)]
    clean_lines = [
        f"# Subconjunto: DCI conocidos, marcas frecuentes o texto con mg/ml/tab/amp (sin puro ruido geográfico/limpieza)",
        f"# Términos distintos: {len(clean_items)}",
        "",
    ]
    for k, n in clean_items:
        clean_lines.append(f"{n:6d} | {display_name(k)}")
    text_clean = "\n".join(clean_lines) + "\n"

    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(text, encoding="utf-8")
        clean_path = out_path.with_name(
            out_path.stem + "_solo_dci_marcas" + out_path.suffix
        )
        clean_path.write_text(text_clean, encoding="utf-8")
        print(f"Guardado: {out_path}", file=sys.stderr)
        print(f"Listado DCI/marcas: {clean_path}", file=sys.stderr)
    else:
        print(text[:500000], end="")
        if len(text) > 500000:
            print("\n# … use --out", file=sys.stderr)
    print(
        f"Términos: listado amplio {len(items)} | DCI/marcas filtrado {len(clean_items)}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
