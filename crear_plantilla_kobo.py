#!/usr/bin/env python3
"""
Genera la plantilla Excel optimizada para el sistema de llenado KoboToolbox.
Cada columna coincide exactamente con EXCEL_TO_INTERNAL en excel_loader.py
y con las reglas de filling_rules.py.
"""

from pathlib import Path
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = Path(__file__).resolve().parent / "Plantilla_Llenado_Kobo.xlsx"

# Definición de columnas: (encabezado_excel, campo_interno, requerido, valores_validos, ancho, nota)
COLUMNS = [
    # --- OBLIGATORIOS ---
    ("Nombre del Paciente", "NAME", True, None, 28,
     "Nombre completo. Se convierte a MAYÚSCULAS. Si no autoriza, poner 'NA'."),
    ("Fecha de atención", "Fecha_de_atenci_n", True, None, 18,
     "Formato: AAAA-MM-DD (ej. 2026-03-12). Si se deja vacío se usa la fecha de hoy."),
    ("Sexo", "SEX", True, ["F", "M"], 10,
     "F = Femenino, M = Masculino. Se traduce automáticamente."),
    ("Servicio que se brinda", "Servicio_que_se_brinda", True,
     ["Medicina General", "Dental", "Oftalmología", "Fisioterapia", "Laboratorios"], 24,
     "Valores exactos del formulario. 'Odontología' se convierte a 'Dental'."),

    # --- RECOMENDADOS ---
    ("Edad", "AGE", False, None, 8,
     "Edad en años. Si no hay Fecha de nacimiento, se calcula DOB aproximado."),
    ("Fecha de nacimiento", "DOB", False, None, 18,
     "Formato: AAAA-MM-DD. Si está vacía y hay Edad, se calcula automáticamente."),
    ("Estado", "Estado_brigada", False,
     ["Baja California Sur", "Chihuahua", "Sonora", "Baja California", "Nuevo León"], 22,
     "Estado donde se realiza la brigada. Determina el 'Lugar de atención' (POC) del formulario."),
    ("Lugar de atención", "Lugar", False, None, 28,
     "Nombre del lugar específico (ej. 'Clínica Adventista', 'Escuela X'). Se usa para PLACE y OTH."),

    # --- OPCIONALES (con defaults automáticos) ---
    ("Modalidad", "Modalidad_de_la_atenci_n", False,
     ["Móvil", "Fija"], 14,
     "Default: 'Móvil'. Tipo de atención."),
    ("Primera vez o Seguimiento", "followup", False,
     ["Primera vez", "Seguimiento", "Atención Única"], 24,
     "Default: 'Primera vez'. Valores exactos del formulario."),
    ("Nacionalidad", "NAT", False, None, 16,
     "Default: 'México'. Solo cambiar si el paciente no es mexicano."),
    ("Estatus migratorio", "estatus_migra", False,
     ["Ciudadano Mexicano", "Solicitante de asilo", "Refugiado",
      "Migrante", "Retornado Mexicano", "Desplazado Interno",
      "Apátrida", "Comunidad de Acogida", "Otro"], 22,
     "Default: 'Ciudadano Mexicano'."),
    ("Minoría", "_Pertenece_a_alguna_minor_a_t", False,
     ["No", "Sí"], 10,
     "Default: 'No'. ¿Pertenece a alguna minoría étnica?"),

    # --- MEDIDAS ---
    ("Talla (cm)", "HEI", False, None, 12,
     "Estatura en centímetros (ej. 167). Usar punto decimal (167.5)."),
    ("Peso (kg)", "WEI", False, None, 12,
     "Peso en kilogramos (ej. 63). Usar punto decimal (63.5)."),

    # --- DIAGNÓSTICO ---
    ("Padecimiento médico actual", "Diagnostico_Motivo", False, None, 32,
     "Texto libre. Se usa para HPI (padecimiento), dxesp (diagnóstico específico) y TX (tratamiento)."),

    # --- TRATAMIENTO / ENTREGA ---
    ("¿Entrega Tratamiento?", "entrega_tx", False,
     ["Sí", "No"], 20,
     "Default: 'Sí' si hay Insumos Entregados, 'No' si no. Valores: Sí / No."),
    ("Insumos Entregados", "Resultados_Lab_Insumos", False, None, 28,
     "Descripción de lo que se entrega (medicamentos, resultados lab, etc.)."),

    # --- REFERENCIA ---
    ("¿Se hizo referencia?", "Referencia", False,
     ["Sí", "No"], 20,
     "¿Se refiere al paciente? Default: 'No'."),
    ("¿A dónde?", "Referencia_donde", False,
     ["Clínica", "Segundo Nivel", "ONG", "Ministerio público", "Otro"], 18,
     "Solo si ¿Se hizo referencia? = Sí. Default: 'Clínica'."),
    ("Motivo Ref.", "Motivo_referencia", False,
     ["Desnutrición", "Seguimiento embarazo", "Valoración y tratamiento",
      "PB Neumonía", "Cirugía", "Otro"], 22,
     "Solo si ¿Se hizo referencia? = Sí. Motivo de la referencia."),

    # --- ACOMPAÑANTE ---
    ("Acompañante", "CGR", False,
     ["Cuidadora mujer", "Cuidador hombre", "Solo/a", "Otro familiar"], 18,
     "Auto: 'Cuidadora mujer' si es menor de 18 años. ¿Quién lo acompaña?"),
]

FILL_REQUIRED = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
FILL_RECOMMENDED = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
FILL_OPTIONAL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
FILL_HEADER_REQ = PatternFill(start_color="FF548235", end_color="FF548235", fill_type="solid")
FILL_HEADER_REC = PatternFill(start_color="FFBF8F00", end_color="FFBF8F00", fill_type="solid")
FILL_HEADER_OPT = PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid")
FILL_INSTRUCCIONES_HEADER = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFFFF", size=11)
FONT_NORMAL = Font(size=10)
FONT_NOTE = Font(size=9, italic=True, color="FF666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFCCCCCC"),
    right=Side(style="thin", color="FFCCCCCC"),
    top=Side(style="thin", color="FFCCCCCC"),
    bottom=Side(style="thin", color="FFCCCCCC"),
)

EXAMPLE_ROWS = [
    {
        "Nombre del Paciente": "María García López",
        "Fecha de atención": date.today().isoformat(),
        "Sexo": "F",
        "Servicio que se brinda": "Medicina General",
        "Edad": "35",
        "Fecha de nacimiento": "1991-05-14",
        "Estado": "Baja California Sur",
        "Lugar de atención": "Clínica Adventista La Paz",
        "Modalidad": "Móvil",
        "Primera vez o Seguimiento": "Primera vez",
        "Nacionalidad": "México",
        "Estatus migratorio": "Ciudadano Mexicano",
        "Minoría": "No",
        "Talla (cm)": "162",
        "Peso (kg)": "58",
        "Padecimiento médico actual": "Cefalea tensional",
        "¿Entrega Tratamiento?": "Sí",
        "Insumos Entregados": "Paracetamol 500mg",
        "¿Se hizo referencia?": "No",
        "¿A dónde?": "",
        "Motivo Ref.": "",
        "Acompañante": "Solo/a",
    },
    {
        "Nombre del Paciente": "Juan Pérez Martínez",
        "Fecha de atención": date.today().isoformat(),
        "Sexo": "M",
        "Servicio que se brinda": "Dental",
        "Edad": "8",
        "Fecha de nacimiento": "2018-01-20",
        "Estado": "Baja California Sur",
        "Lugar de atención": "Escuela Primaria Benito Juárez",
        "Modalidad": "Móvil",
        "Primera vez o Seguimiento": "Primera vez",
        "Nacionalidad": "México",
        "Estatus migratorio": "Ciudadano Mexicano",
        "Minoría": "No",
        "Talla (cm)": "128",
        "Peso (kg)": "27",
        "Padecimiento médico actual": "Caries dental múltiple",
        "¿Entrega Tratamiento?": "Sí",
        "Insumos Entregados": "Kit dental, flúor",
        "¿Se hizo referencia?": "Sí",
        "¿A dónde?": "Clínica",
        "Motivo Ref.": "Valoración y tratamiento",
        "Acompañante": "Cuidadora mujer",
    },
    {
        "Nombre del Paciente": "Ana Torres Vega",
        "Fecha de atención": date.today().isoformat(),
        "Sexo": "F",
        "Servicio que se brinda": "Oftalmología",
        "Edad": "62",
        "Fecha de nacimiento": "",
        "Estado": "Chihuahua",
        "Lugar de atención": "Centro Comunitario",
        "Modalidad": "Móvil",
        "Primera vez o Seguimiento": "Primera vez",
        "Nacionalidad": "México",
        "Estatus migratorio": "Ciudadano Mexicano",
        "Minoría": "No",
        "Talla (cm)": "155",
        "Peso (kg)": "68",
        "Padecimiento médico actual": "Disminución de agudeza visual",
        "¿Entrega Tratamiento?": "Sí",
        "Insumos Entregados": "Anteojos graduados",
        "¿Se hizo referencia?": "No",
        "¿A dónde?": "",
        "Motivo Ref.": "",
        "Acompañante": "Solo/a",
    },
]


def create_template():
    wb = Workbook()

    # ============ HOJA 1: DATOS ============
    ws = wb.active
    ws.title = "Datos"
    ws.sheet_properties.tabColor = "548235"

    headers = [c[0] for c in COLUMNS]
    required_indices = set()
    recommended_indices = set()

    for col_idx, col_def in enumerate(COLUMNS, 1):
        name, internal, required, valid_values, width, note = col_def
        letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[letter].width = width

        if required:
            cell.fill = FILL_HEADER_REQ
            required_indices.add(col_idx)
        elif col_idx <= 8:
            cell.fill = FILL_HEADER_REC
            recommended_indices.add(col_idx)
        else:
            cell.fill = FILL_HEADER_OPT

        if valid_values:
            formula = '"' + ",".join(valid_values) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = f"Valor no válido para '{name}'"
            dv.errorTitle = "Valor inválido"
            dv.prompt = f"Selecciona un valor para '{name}'"
            dv.promptTitle = name
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}1000")

    ws.row_dimensions[1].height = 30

    for row_idx, example in enumerate(EXAMPLE_ROWS, 2):
        for col_idx, col_def in enumerate(COLUMNS, 1):
            name = col_def[0]
            val = example.get(name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if col_idx in required_indices:
                cell.fill = FILL_REQUIRED
            elif col_idx in recommended_indices:
                cell.fill = FILL_RECOMMENDED
            else:
                cell.fill = FILL_OPTIONAL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ============ HOJA 2: INSTRUCCIONES ============
    wi = wb.create_sheet("Instrucciones")
    wi.sheet_properties.tabColor = "2F5496"

    inst_headers = ["Columna Excel", "Campo Interno", "¿Requerido?", "Valores Válidos", "Default Automático", "Descripción"]
    inst_widths = [28, 30, 14, 40, 24, 50]
    for col_idx, (header, w) in enumerate(zip(inst_headers, inst_widths), 1):
        cell = wi.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_INSTRUCCIONES_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        wi.column_dimensions[get_column_letter(col_idx)].width = w

    defaults_map = {
        "Fecha_de_atenci_n": "Fecha de hoy",
        "SEX": "Femenino",
        "Modalidad_de_la_atenci_n": "Móvil",
        "followup": "Primera vez",
        "NAT": "México",
        "estatus_migra": "Ciudadano Mexicano",
        "_Pertenece_a_alguna_minor_a_t": "No",
        "entrega_tx": "Sí (si hay insumos)",
        "Referencia": "No",
        "Diagnostico_Motivo": "—",
        "CGR": "Cuidadora mujer (<18 años)",
    }

    auto_fields = [
        ("CONS1 (Consentimiento inicial)", "Se marca 'Sí' automáticamente"),
        ("CONS (Consentimiento verbal)", "Se marca 'Sí' automáticamente"),
        ("NAT / NATOT (Nacionalidad)", "'México' por defecto"),
        ("POC (Estado formulario)", "Se calcula a partir de la columna 'Estado'"),
        ("PLACE / OTH", "Se llenan desde 'Lugar de atención'"),
        ("DOB (Fecha nacimiento)", "Se calcula desde 'Edad' si no se proporciona"),
        ("HPI (Padecimiento)", "Se copia de 'Padecimiento médico actual'"),
        ("dxesp (Diagnóstico específico)", "Se copia de 'Padecimiento médico actual'"),
        ("TX (Tratamiento)", "Se copia de 'Padecimiento médico actual'"),
        ("ASESPREV (Asesoría previa)", "Se copia de 'Servicio que se brinda'"),
        ("Especifique_qu_se_entrega", "'Medicamento/suplemento' por defecto"),
        ("Unidades_entregadas", "'1' por defecto"),
        ("REF / REFORG / MEDREF", "Se calculan desde columnas de referencia"),
        ("Estado (origen persona)", "Se copia de la columna 'Estado'"),
    ]

    for row_idx, col_def in enumerate(COLUMNS, 2):
        name, internal, required, valid_values, _, note = col_def
        wi.cell(row=row_idx, column=1, value=name).font = FONT_NORMAL
        wi.cell(row=row_idx, column=2, value=internal).font = Font(size=10, name="Consolas")
        req_text = "SÍ" if required else ("Recomendado" if row_idx <= 9 else "Opcional")
        cell_req = wi.cell(row=row_idx, column=3, value=req_text)
        cell_req.font = FONT_NORMAL
        if required:
            cell_req.fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
        wi.cell(row=row_idx, column=4, value=", ".join(valid_values) if valid_values else "Texto libre").font = FONT_NORMAL
        wi.cell(row=row_idx, column=5, value=defaults_map.get(internal, "—")).font = FONT_NORMAL
        wi.cell(row=row_idx, column=6, value=note).font = FONT_NORMAL
        for c in range(1, 7):
            wi.cell(row=row_idx, column=c).border = THIN_BORDER
            wi.cell(row=row_idx, column=c).alignment = Alignment(vertical="center", wrap_text=True)

    sep_row = len(COLUMNS) + 3
    cell_sep = wi.cell(row=sep_row, column=1, value="CAMPOS AUTOMÁTICOS (no necesitas ponerlos en el Excel)")
    cell_sep.font = Font(bold=True, size=11, color="FF2F5496")
    wi.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=6)

    for i, (field, desc) in enumerate(auto_fields, sep_row + 1):
        wi.cell(row=i, column=1, value=field).font = Font(size=10, name="Consolas")
        wi.cell(row=i, column=2, value=desc).font = FONT_NORMAL
        for c in range(1, 7):
            wi.cell(row=i, column=c).border = THIN_BORDER

    wi.freeze_panes = "A2"

    # ============ HOJA 3: VALORES DEL FORMULARIO ============
    wv = wb.create_sheet("Valores Formulario")
    wv.sheet_properties.tabColor = "BF8F00"

    form_values = [
        ("Servicio que se brinda", ["Medicina General", "Dental", "Oftalmología", "Fisioterapia", "Laboratorios"]),
        ("Sexo (se traduce)", ["F → Femenino", "M → Masculino"]),
        ("Estado (POC)", ["Baja California Sur", "Chihuahua", "Sonora", "Baja California", "Nuevo León"]),
        ("Modalidad", ["Móvil", "Fija"]),
        ("Primera vez / Seguimiento", ["Primera vez", "Seguimiento", "Atención Única"]),
        ("Estatus migratorio", [
            "Ciudadano Mexicano", "Solicitante de asilo", "Refugiado",
            "Migrante", "Retornado Mexicano", "Desplazado Interno",
            "Apátrida", "Comunidad de Acogida", "Otro"
        ]),
        ("¿Entrega Tratamiento?", ["Sí → 1", "No → 0"]),
        ("¿Se hizo referencia?", ["Sí → 1", "No → 0"]),
        ("¿A dónde? (referencia)", [
            "Clínica", "Segundo Nivel de Atención/Especialidad",
            "ONG", "Ministerio público", "Otro"
        ]),
        ("Motivo referencia", [
            "Desnutrición", "Seguimiento embarazo",
            "Valoración y tratamiento", "PB Neumonía", "Cirugía", "Otro"
        ]),
        ("Acompañante", ["Cuidadora mujer", "Cuidador hombre", "Solo/a", "Otro familiar"]),
        ("Minoría étnica", ["Sí → 1", "No → 0"]),
        ("Especifique qué se entrega", [
            "Anteojos", "Medicamento/suplemento",
            "Plan de Tratamiento", "Resultados de Laboratorio", "Otro"
        ]),
    ]

    max_vals = max(len(v[1]) for v in form_values)
    for col_idx, (campo, valores) in enumerate(form_values, 1):
        letter = get_column_letter(col_idx)
        cell = wv.cell(row=1, column=col_idx, value=campo)
        cell.font = FONT_HEADER
        cell.fill = FILL_INSTRUCCIONES_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        wv.column_dimensions[letter].width = 28
        for val_idx, val in enumerate(valores, 2):
            c = wv.cell(row=val_idx, column=col_idx, value=val)
            c.font = FONT_NORMAL
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")

    wv.freeze_panes = "A2"

    wb.save(str(OUTPUT_PATH))
    print(f"Plantilla generada: {OUTPUT_PATH}")
    print(f"  - Hoja 'Datos': {len(COLUMNS)} columnas, {len(EXAMPLE_ROWS)} filas de ejemplo")
    print(f"  - Hoja 'Instrucciones': descripción de cada campo y campos automáticos")
    print(f"  - Hoja 'Valores Formulario': todos los valores válidos del formulario KoboToolbox")


if __name__ == "__main__":
    create_template()
